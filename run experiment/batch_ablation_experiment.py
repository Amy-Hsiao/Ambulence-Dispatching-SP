#!/usr/bin/env python3
"""B&BC acceleration ablation runner（實驗三，plan/12 + plan/16）。

2026-08 放大規模版本
--------------------
* 資料：台北市全區（災區 229 / CCP 候選點 50 / 醫院 20）
* 規模：small = I70, medium = I100, large = I130；三者 |J| = 50、|H| = 18
* 模型（4 種）：SP+MCVaR、DRO-box、DRO-ellipsoidal、DRO-polyhedral
* 配置（6 種）：Extensive、BBC、BBC+WS、BBC+WS+RS、BBC+WS+RS+UC、
  BBC+WS+RS+UC+Pareto
  → 共 3 × 4 × 6 = 72 個 case
* 求解：每個 case 上限 2 小時；relative MIP gap ≤ 1% 即提早結束

固定同一個隨機 instance（同 scale 下所有 config / model 共用），每個 case
都有獨立 log，並在每個 case 後以原子方式重寫 raw CSV 與完整 Excel；
單一 case 失敗會留下 FAIL 記錄並繼續。中斷後把最新的 raw CSV 檔名填進
RESUME_FROM_CSV 即可從中斷處續跑。
"""
from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Any


# =============================================================================
# Parameter setting area
# =============================================================================

# 規模：small / medium / large 三個 case 都會依序跑（plan/15 的 SCALE_PROFILES）。
SCALES = ["small", "medium", "large"]

BASE_SCENARIOS = 30
BASE_CCP_SAMPLE_SIZE = None
BASE_SAMPLE_RATIO = 1.0
BASE_TIME_PERIODS = 8
BASE_DEMAND_MULTIPLIER = 1.0
BASE_ROAD_CAPACITY_MULTIPLIER = 1.0
BASE_HOSPITAL_CAPACITY_MULTIPLIER = 1.0

RISK_ALPHA = 0.9
RISK_LAMBDA = 0.5
# 各 ambiguity set 的 scope（box=ε̄_B, ellipsoidal=a_E, polyhedral=a_P）
# box 的 ε̄_B 必須 ≤ min_s p0_s；等權重 S=30 時 = 1/30 ≈ 0.0333，故 0.01 合法。
BOX_SCOPE = 0.01
ELLIPSOIDAL_SCOPE = 0.0005
POLYHEDRAL_SCOPE = 0.001

# ── 求解時間 / 收斂門檻（2026-08 放大規模後調整）──
# TIME_LIMIT：單一 case 的求解時間上限 = 2 小時。
# MIP_GAP   ：提早結束門檻。Gurobi 一旦達到 relative MIP gap ≤ 1% 就立刻停止，
#             不會跑滿 2 小時；跑不到 1% 的 case 才會用滿 TIME_LIMIT。
#             六種 config 用同一個門檻，時間比較才公平（同收斂精度比時間）。
TIME_LIMIT = 7200.0     # 2 小時
MIP_GAP = 0.01          # 1% 即提早結束
# 子程序硬超時 = 求解時限 + 緩衝。必須大於 TIME_LIMIT，否則跑滿時限的 case
# 會在把結果寫回前就被 subprocess timeout 砍掉，被誤判為 FAIL（Hard timeout）。
# 緩衝需涵蓋：子程序啟動、產生 instance、建模型、寫結果等時間。
#
# 放大規模後兩種引擎的建模成本差很多，故分開設定：
#   * B&BC   ：master 只有 50 個 0-1 + 30 個 θ；oracle 每情境約 10~18 萬變數，
#              逐一建立，建模很快 → 30 分鐘緩衝綽綽有餘。
#   * Extensive：一次建出全部情境，large 高達約 530 萬個變數
#              （FI = S·|I|·|J|·|L|·|T| = 30×130×50×3×8 = 468 萬），
#              光是 gurobipy 建模就可能數十分鐘 → 給 90 分鐘緩衝，
#              避免「真的跑滿 2 小時求解」的 case 被誤記成 Hard timeout。
HARD_TIMEOUT_BUFFER_SEC = 1800.0            # B&BC 類（6 個 config 中的 5 個）
HARD_TIMEOUT_BUFFER_SEC_EXTENSIVE = 5400.0  # Extensive form
COMPUTE_KPIS = False
STOP_ON_ERROR = False

# Excel 檔被 Excel.exe 鎖住時的重試設定（Windows 常見）。
EXCEL_REPLACE_RETRIES = 5
EXCEL_REPLACE_RETRY_SLEEP_SEC = 3.0

RESULT_PREFIX = "BBC_ablation_MCVaR_Box_Ell_Poly"
LOG_SUBDIR_NAME = "bbc ablation"


# =============================================================================
# Setup
# =============================================================================

ROOT_DIR = Path(__file__).resolve().parents[1]
LOG_DIR = ROOT_DIR / "logs"
LOG_SUBDIR = LOG_DIR / LOG_SUBDIR_NAME
RESULT_DIR = ROOT_DIR / "experiment result"

# Keep the experiment launcher self-contained in this workspace.  The local
# spreadsheet runtime is optional on copied workspaces; when it is absent we
# fail before starting a potentially multi-day solver run.
LOCAL_PYTHON_PACKAGE_CANDIDATES = [
    ROOT_DIR / ".codex_spreadsheet" / "python_packages",
    Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime"
    / "dependencies" / "python" / "Lib" / "site-packages",
]
if os.environ.get("CODEX_PRIMARY_PYTHON_PACKAGES"):
    LOCAL_PYTHON_PACKAGE_CANDIDATES.insert(
        0, Path(os.environ["CODEX_PRIMARY_PYTHON_PACKAGES"])
    )
for package_dir in reversed(LOCAL_PYTHON_PACKAGE_CANDIDATES):
    if package_dir.exists():
        sys.path.insert(0, str(package_dir))

os.chdir(ROOT_DIR)
for _p in (str(ROOT_DIR / "model core"), str(ROOT_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import config as cfg  # noqa: E402
import logging_utils  # noqa: E402

SP_MODEL_PATH = ROOT_DIR / "model portal" / "benders bbc.py"
DRO_MODEL_PATH = ROOT_DIR / "model portal" / "dro bbc.py"
EXTENSIVE_MODEL_PATH = ROOT_DIR / "model portal" / "extensive_dro.py"
MCVAR_MODEL_PATH = ROOT_DIR / "model portal" / "mcvar bbc.py"

DEFAULT_ROOT_SEED_ITERS = int(cfg.BENDERS_ROOT_SEED_ITERS)
DEFAULT_ROOT_CUT_ROUNDS = int(cfg.BENDERS_ROOT_CUT_ROUNDS)
DEFAULT_MIPFOCUS = int(cfg.BENDERS_MIPFOCUS)
DEFAULT_HEURISTICS = float(cfg.BENDERS_HEURISTICS)
DEFAULT_NUMERIC_FOCUS = int(cfg.BENDERS_NUMERIC_FOCUS)
DEFAULT_BRANCH_PRIORITY_ENABLED = bool(cfg.BENDERS_X_BRANCH_PRIORITY_ENABLED)
DEFAULT_BRANCH_PRIORITY = int(cfg.BENDERS_X_BRANCH_PRIORITY)

CONFIGS = [
    {"name": "Extensive", "engine": "extensive", "ev": False, "seed": 0, "rounds": 0, "user": False, "pareto": False},
    {"name": "BBC",      "ev": False, "seed": 0,                       "rounds": 0,                       "user": False, "pareto": False},
    {"name": "BBC+WS",   "ev": True,  "seed": 0,                       "rounds": 0,                       "user": False, "pareto": False},
    {"name": "BBC+WS+RS", "ev": True, "seed": DEFAULT_ROOT_SEED_ITERS, "rounds": 0,                       "user": False, "pareto": False},
    {"name": "BBC+WS+RS+UC", "ev": True, "seed": DEFAULT_ROOT_SEED_ITERS, "rounds": DEFAULT_ROOT_CUT_ROUNDS, "user": True, "pareto": False},
    {"name": "BBC+WS+RS+UC+Pareto", "ev": True, "seed": DEFAULT_ROOT_SEED_ITERS, "rounds": DEFAULT_ROOT_CUT_ROUNDS, "user": True, "pareto": True},
]
# model 維度（risk 類型），2026-08 重跑：四種全做。
#   SP+MCVaR / SP+MCVaR+DRO(box) / +DRO(ellipsoidal) / +DRO(polyhedral)
# kind="mcvar"（無 ambiguity set）或 "dro"（需 ambiguity_set + scope）。
MODEL_SPECS = [
    {"name": "SP+MCVaR",        "kind": "mcvar"},
    {"name": "DRO-box",         "kind": "dro", "ambiguity_set": "box",         "scope": BOX_SCOPE},
    {"name": "DRO-ellipsoidal", "kind": "dro", "ambiguity_set": "ellipsoidal", "scope": ELLIPSOIDAL_SCOPE},
    {"name": "DRO-polyhedral",  "kind": "dro", "ambiguity_set": "polyhedral",  "scope": POLYHEDRAL_SCOPE},
]
MODELS = [spec["name"] for spec in MODEL_SPECS]
_SPEC_BY_NAME = {spec["name"]: spec for spec in MODEL_SPECS}

# 只重跑指定 config（依 name 過濾）；空清單 = 跑全部 6 個 config。
# 要「只重跑 extensive」就設成 ["Extensive"]（3 scale x 2 model x 1 config = 6 個 case）。
RUN_ONLY_CONFIGS: list[str] = []

# ── 中斷續跑 ──
# 指向上次未跑完的 raw CSV（填 experiment result/ 內的檔名即可，或絕對路徑）。
# 設定後只補跑「缺少或非 OK」的 case，並把上次已完成(OK)的結果一起併進本次新輸出，
# 得到完整一份。留空 = 從頭跑全部。
#
# raw CSV 在「每個 case 結束後」就會原子性重寫一次，所以就算整台機器斷電，
# 已完成的 case 一定留在檔案裡，不會白跑。中斷後的操作：
#   1. 到 experiment result/ 找最新的  BBC_ablation_..._raw_YYYYmmdd_HHMMSS.csv
#   2. 把檔名貼到下面 RESUME_FROM_CSV
#   3. 重新執行本程式（跑完會另存一份「完整」的新 CSV / Excel）
# 程式結束時若有未完成的 case，畫面上會直接印出要貼的檔名。
RESUME_FROM_CSV: str = ""


def active_configs() -> list[dict[str, Any]]:
    """實際要跑/驗證/輸出的 config 子集（受 RUN_ONLY_CONFIGS 過濾）。"""
    if RUN_ONLY_CONFIGS:
        chosen = [c for c in CONFIGS if c["name"] in RUN_ONLY_CONFIGS]
        if not chosen:
            raise ValueError(
                f"RUN_ONLY_CONFIGS={RUN_ONLY_CONFIGS} 未匹配任何 CONFIGS 名稱"
            )
        return chosen
    return list(CONFIGS)

FIELDNAMES = [
    "scale", "test_id", "model", "config", "I", "J", "H", "S", "T",
    "obj_value", "first_stage_decision", "best_lb", "best_ub", "cpu_s",
    "wall_s", "num_vars", "num_constrs", "nodes", "iterations", "gap_pct",
    "total_cuts", "seed_cuts", "lazy_cuts", "user_cuts", "root_seed_lb",
    "root_seed_iters_done", "root_cut_rounds_done", "ev_warm_start",
    "root_seed_iters", "root_cut_rounds", "use_user_cuts", "pareto_enabled",
    "multi_cut", "parallel_oracles", "oracle_solves", "incumbent_evals",
    "mip_focus", "heuristics", "numeric_focus", "branch_priority_enabled",
    "branch_priority",
    "callback_time_s", "solver_status", "log_path", "solver_log_path",
    "status", "note",
]

DETAIL_COLUMNS = [
    ("|I| Disaster", "I"), ("|J| CCP", "J"), ("|H| Hosp", "H"),
    ("|S| Scen", "S"), ("|T| Per", "T"), ("obj_value", "obj_value"),
    ("First Stage Decision", "first_stage_decision"), ("Best LB", "best_lb"),
    ("Best UB", "best_ub"), ("CPU Time(s)", "cpu_s"),
    ("num_vars", "num_vars"), ("num_constrs", "num_constrs"),
    ("Nodes", "nodes"), ("Iteration", "iterations"),
    ("Final Gap(%)", "gap_pct"), ("Total Cuts", "total_cuts"),
    ("Seed Cuts", "seed_cuts"), ("Lazy Cuts", "lazy_cuts"),
    ("User Cuts", "user_cuts"), ("Seeded LB(root)", "root_seed_lb"),
    ("model", "model"), ("config", "config"), ("status", "status"),
]
TEXT_KEYS = {
    "scale", "test_id", "model", "config", "first_stage_decision",
    "solver_status", "log_path", "solver_log_path", "status", "note",
}
BOOL_KEYS = {
    "ev_warm_start", "use_user_cuts", "pareto_enabled", "multi_cut",
    "branch_priority_enabled",
}
INTEGER_KEYS = {
    "I", "J", "H", "S", "T", "num_vars", "num_constrs", "nodes",
    "iterations", "total_cuts", "seed_cuts", "lazy_cuts", "user_cuts",
    "root_seed_iters_done", "root_cut_rounds_done", "root_seed_iters",
    "root_cut_rounds", "parallel_oracles", "oracle_solves", "incumbent_evals",
    "mip_focus", "numeric_focus", "branch_priority",
}


def blank_row() -> dict[str, Any]:
    return {key: "NA" for key in FIELDNAMES}


def ensure_excel_dependency() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "Excel output is required, but openpyxl is unavailable in this Python. "
            f"Install it for {sys.executable!r} before running the experiment: "
            f"{sys.executable!r} -m pip install openpyxl"
        ) from exc


def ensure_solver_dependency() -> None:
    try:
        import gurobipy  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "Gurobi is installed, but gurobipy is unavailable in the Python "
            f"running this script ({sys.executable!r}). Install the matching "
            f"package before running: {sys.executable!r} -m pip install gurobipy"
        ) from exc


def load_portal(path: Path, module_name: str) -> Any:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load portal: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return max(0, sum(1 for row in csv.reader(f) if any(row)) - 1)


def scale_counts(scale: str) -> dict[str, int]:
    """依 config.SCALE_PROFILES 回傳該規模的 (I, J, H) 目標筆數。"""
    profile = cfg.SCALE_PROFILES[scale]
    return {
        "I": int(profile["n_disaster"]),
        "J": int(profile["n_ccp"]),
        "H": int(profile["n_hospital"]),
    }


def first_stage_string(fs: dict[str, Any] | None) -> str:
    if not fs:
        return "NA"
    supply: dict[str, float] = {}
    for (h, j), value in fs["Y"].items():
        supply[j] = supply.get(j, 0.0) + float(value)
    lines = []
    for j in sorted(fs["X"]):
        if float(fs["X"][j]) > 0.5:
            lines.append(
                f"CCP {j:4s} -> X: 1, Staff(V): {float(fs['V'][j]):.0f}, "
                f"Amb(U): {float(fs['U'][j]):.0f}, MedicalSupply(Y): {supply.get(j, 0.0):.2f}"
            )
    return "\n".join(lines) if lines else "none opened"


def snapshot_logs() -> set[Path]:
    return {p.resolve() for p in LOG_DIR.glob("*.log")} if LOG_DIR.exists() else set()


def move_newest_log(before: set[Path], destination: Path) -> Path | None:
    created = [p for p in LOG_DIR.glob("*.log") if p.resolve() not in before]
    if not created:
        return None
    source = max(created, key=lambda p: p.stat().st_mtime)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        raise FileExistsError(f"Refusing to overwrite solver log: {destination}")
    shutil.move(str(source), str(destination))
    return destination


@contextmanager
def temporary_config(case: dict[str, Any]):
    is_extensive = case.get("engine") == "extensive"
    values = {
        "SCENARIOS": BASE_SCENARIOS,
        "TIME_PERIODS": BASE_TIME_PERIODS,
        "SAMPLE_RATIO": BASE_SAMPLE_RATIO,
        "SP_SAMPLE_RATIO": BASE_SAMPLE_RATIO,
        "DEMAND_MULTIPLIER": BASE_DEMAND_MULTIPLIER,
        "ROAD_CAPACITY_MULTIPLIER": BASE_ROAD_CAPACITY_MULTIPLIER,
        "HOSPITAL_CAPACITY_MULTIPLIER": BASE_HOSPITAL_CAPACITY_MULTIPLIER,
        "SP_TIME_LIMIT": TIME_LIMIT,
        "SP_MIP_GAP": MIP_GAP,
        "BENDERS_MULTI_CUT": True,
        "BENDERS_EV_WARM_START": case["ev"],
        "BENDERS_ROOT_SEED_ITERS": case["seed"],
        "BENDERS_ROOT_CUT_ROUNDS": case["rounds"],
        "BENDERS_USE_USER_CUTS": case["user"],
        "BENDERS_PARETO_ENABLED": case["pareto"],
        # Common B&BC tuning requested for every non-Extensive configuration.
        "BENDERS_MIPFOCUS": 0 if is_extensive else DEFAULT_MIPFOCUS,
        "BENDERS_HEURISTICS": 0.0 if is_extensive else DEFAULT_HEURISTICS,
        "BENDERS_NUMERIC_FOCUS": 0 if is_extensive else DEFAULT_NUMERIC_FOCUS,
        "BENDERS_X_BRANCH_PRIORITY_ENABLED": (
            False if is_extensive else DEFAULT_BRANCH_PRIORITY_ENABLED
        ),
        "BENDERS_X_BRANCH_PRIORITY": (
            0 if is_extensive else DEFAULT_BRANCH_PRIORITY
        ),
    }
    if hasattr(cfg, "CCP_SAMPLE_SIZE"):
        values["CCP_SAMPLE_SIZE"] = BASE_CCP_SAMPLE_SIZE
    original = {key: getattr(cfg, key) for key in values}
    try:
        for key, value in values.items():
            setattr(cfg, key, value)
        yield
    finally:
        for key, value in original.items():
            setattr(cfg, key, value)


@contextmanager
def scaled_instance_generation():
    """依 cfg.EXPERIMENT_SCALE 產生可重現 instance，並在單一 case 內快取。

    portal 內部呼叫 config.generate_data(sample_ratio=...)；此處攔截並改以
    scale 產生（忽略 sample_ratio / ccp_sample_size），確保走 plan/15 規模路徑。
    """
    original = cfg.generate_data
    cache: dict[str, Any] = {}

    def _scaled(*args, **kwargs):
        kwargs.pop("sample_ratio", None)
        kwargs.pop("ccp_sample_size", None)
        scale = cfg.EXPERIMENT_SCALE
        if scale not in cache:
            cache[scale] = original(scale=scale)
        return cache[scale]

    cfg.generate_data = _scaled
    try:
        yield
    finally:
        cfg.generate_data = original


def write_results(path: Path, rows: list[dict[str, Any]]) -> None:
    """原子性寫出 raw CSV（續跑的唯一依據，必須盡最大努力保住）。

    先寫暫存檔再 os.replace；若目標檔被 Excel 鎖住（Windows PermissionError）
    則重試，仍失敗就另存一份 rescue 檔，確保跑了好幾小時的結果不會遺失。
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}_", suffix=".csv", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        with temp_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
            writer.writeheader()
            writer.writerows(rows)
        for attempt in range(EXCEL_REPLACE_RETRIES):
            try:
                os.replace(temp_path, path)
                return
            except PermissionError:
                if attempt == 0:
                    print(f"[csv] 無法寫入 {path.name}（檔案可能正被 Excel 開啟）；重試中 …")
                time.sleep(EXCEL_REPLACE_RETRY_SLEEP_SEC)
        rescue = path.with_name(f"{path.stem}.rescue_{int(time.time())}.csv")
        shutil.copy2(temp_path, rescue)
        print(f"[csv][WARN] {path.name} 被鎖住無法寫入，已另存 {rescue.name}；"
              "續跑時請改用這個檔名。")
    finally:
        temp_path.unlink(missing_ok=True)


def _load_prior_rows(path: Path) -> list[dict[str, Any]]:
    """讀上次的 raw CSV（供 RESUME 沿用已完成的 case）。"""
    if not path.is_file():
        raise FileNotFoundError(f"RESUME_FROM_CSV 找不到檔案: {path}")
    out: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            out.append({k: r.get(k, "NA") for k in FIELDNAMES})
    return out


def excel_value(row: dict[str, Any], key: str) -> Any:
    value = row.get(key, "NA")
    if value in (None, "", "NA"):
        return "NA"
    if key in TEXT_KEYS:
        return str(value)
    if key in BOOL_KEYS:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes"}
    try:
        number = float(value)
        return int(number) if key in INTEGER_KEYS and number.is_integer() else number
    except (TypeError, ValueError):
        return value


def _style_header(cell, fill, font, Alignment) -> None:
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_raw_sheet(wb, rows: list[dict[str, Any]], fill, font, Alignment) -> None:
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("raw_results")
    for col, key in enumerate(FIELDNAMES, 1):
        _style_header(ws.cell(1, col, key), fill, font, Alignment)
    for out_row, row in enumerate(rows, 2):
        for col, key in enumerate(FIELDNAMES, 1):
            cell = ws.cell(out_row, col, excel_value(row, key))
            cell.alignment = Alignment(
                horizontal="left" if key in TEXT_KEYS else "right",
                vertical="top",
                wrap_text=key in {"first_stage_decision", "note"},
            )
            if key in INTEGER_KEYS:
                cell.number_format = "#,##0"
            elif key not in TEXT_KEYS and key not in BOOL_KEYS:
                cell.number_format = "#,##0.0000"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}{max(1, len(rows) + 1)}"
    ws.row_dimensions[1].height = 34
    for col, key in enumerate(FIELDNAMES, 1):
        width = 14
        if key == "first_stage_decision":
            width = 62
        elif key in {"test_id", "log_path", "solver_log_path", "note"}:
            width = 34
        elif key in {"config", "solver_status"}:
            width = 24
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_settings_sheet(wb, run_id: str, log_run_dir: Path | None,
                          fill, font, Alignment) -> None:
    ws = wb.create_sheet("run_settings")
    rows = [
        ("run_id", run_id or "NA"),
        ("log_directory", str(log_run_dir) if log_run_dir else "NA"),
        ("scales", ", ".join(SCALES)),
        ("models", ", ".join(MODELS)),
        ("scenario_count", BASE_SCENARIOS),
        ("time_periods", BASE_TIME_PERIODS),
        ("time_limit_s_per_case", TIME_LIMIT),
        ("mip_gap (提早結束門檻)", MIP_GAP),
        ("risk_alpha", RISK_ALPHA),
        ("risk_lambda", RISK_LAMBDA),
        ("box_scope", BOX_SCOPE),
        ("ellipsoidal_scope", ELLIPSOIDAL_SCOPE),
        ("polyhedral_scope", POLYHEDRAL_SCOPE),
        # ── 規模與資源縮放（config.py plan/15 + 2026-08 台北放大）──
        ("data_disaster_csv", cfg.DISASTER_CSV),
        ("data_ccp_csv", cfg.CCP_CSV),
        ("data_hospital_csv", cfg.HOSPITAL_CSV),
        ("param_calibration_basis (I/J/H)",
         f"{cfg.PARAM_CALIB_N_DISASTER}/{cfg.PARAM_CALIB_N_CCP}/{cfg.PARAM_CALIB_N_HOSPITAL}"),
        ("ccp_upper_bound_scaling", cfg.CCP_UPPER_BOUND_SCALING),
        ("scale_sampling_mode", cfg.SCALE_SAMPLING_MODE),
        *[
            (f"scaling[{sc}] s_D/s_J/s_H",
             "{demand_scale:.4f} / {ccp_scale:.4f} / {hospital_scale:.4f}".format(
                 **cfg.resolve_scale(sc)))
            for sc in SCALES
        ],
        ("bbc_multi_cut_common", True),
        ("bbc_parallel_oracles_common", int(cfg.BENDERS_PARALLEL_ORACLES)),
        ("bbc_mip_focus_common", DEFAULT_MIPFOCUS),
        ("bbc_heuristics_common", DEFAULT_HEURISTICS),
        ("bbc_numeric_focus_common", DEFAULT_NUMERIC_FOCUS),
        ("bbc_branch_priority_enabled_common", DEFAULT_BRANCH_PRIORITY_ENABLED),
        ("bbc_branch_priority_common", DEFAULT_BRANCH_PRIORITY),
        ("extensive_threads", 0),
        ("extensive_concurrent_mip", 1),
        ("extensive_presolve", 0),
        ("extensive_cuts", 0),
        ("extensive_heuristics", 0.0),
        ("extensive_symmetry", 0),
        ("extensive_mip_focus", 0),
    ]
    ws.append(["Setting", "Value"])
    for cell in ws[1]:
        _style_header(cell, fill, font, Alignment)
    for item in rows:
        ws.append(item)
    ws.append([])
    ws.append(["Configuration", "EV warm start", "Root seed iterations",
               "Root cut rounds", "User cuts", "Pareto cuts"])
    for cell in ws[ws.max_row]:
        _style_header(cell, fill, font, Alignment)
    for case in CONFIGS:
        ws.append([
            case["name"], case["ev"], case["seed"], case["rounds"],
            case["user"], case["pareto"],
        ])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22


def _write_detail_sheet(wb, scale: str, rows: list[dict[str, Any]], fill, font, Alignment) -> None:
    """單一規模的明細分頁；欄位與原本 DETAIL_COLUMNS 相同。"""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(scale)
    for col, (title, _key) in enumerate(DETAIL_COLUMNS, 1):
        cell = ws.cell(1, col, title)
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    out_row = 2
    for row in rows:
        if row.get("scale") != scale:
            continue
        for col, (_title, key) in enumerate(DETAIL_COLUMNS, 1):
            cell = ws.cell(out_row, col, excel_value(row, key))
            if key == "first_stage_decision":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif key in {"obj_value", "best_lb", "best_ub", "cpu_s", "gap_pct"}:
                cell.number_format = "#,##0.0000"
            elif key not in TEXT_KEYS:
                cell.number_format = "#,##0"
        out_row += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:W{max(1, out_row - 1)}"
    ws.row_dimensions[1].height = 34
    for col in range(1, len(DETAIL_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["G"].width = 62
    ws.column_dimensions["U"].width = 12
    ws.column_dimensions["V"].width = 26
    ws.column_dimensions["W"].width = 10


# summary_table 每個 config 下的子欄位。2026-08 依老師要求加入 Objective Value。
SUMMARY_SUBCOLUMNS = [
    ("Obj Value", "obj_value"),
    ("A.Time", "cpu_s"),
    ("A.Gap(%)", "gap_pct"),
    ("Nodes", "nodes"),
]


def _write_summary_sheet(wb, rows: list[dict[str, Any]], fill, font, Alignment) -> None:
    """圖片風格彙總分頁：Scale×Model 為列，6 個 config 各一組統計。"""
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws = wb.create_sheet("summary_table")

    n_sub = len(SUMMARY_SUBCOLUMNS)
    left = ["Scale", "I", "J", "H", "Model"]
    n_left = len(left)
    n_cols = n_left + len(CONFIGS) * n_sub

    n_models = len(MODELS)
    last = 2 + len(SCALES) * n_models
    blocks = [(3 + k * n_models, 3 + k * n_models + n_models - 1) for k in range(len(SCALES))]

    # 順序很重要：openpyxl 3.x 在「已設 border 的 cell 再 merge」或「對非錨點
    # MergedCell 設 border」時都會炸 column_letter；唯一穩定作法是先 merge、
    # 再寫錨點值、最後統一設樣式。
    # 1) 先合併：表頭左側 instance 欄（縱向）、每個 config（橫向 3 欄）、資料區
    #    每個 scale 的 Scale/I/J/H（跨 model 列）。
    for i in range(1, n_left + 1):
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
    for idx in range(len(CONFIGS)):
        start = n_left + 1 + idx * n_sub
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + n_sub - 1)
    for block_start, block_end in blocks:
        if block_end > block_start:
            for col in range(1, n_left):  # Scale/I/J/H 跨 model 列合併
                ws.merge_cells(start_row=block_start, start_column=col, end_row=block_end, end_column=col)

    # 2) 寫值（只寫合併區的左上角錨點；資料數值欄未合併可自由寫）。
    for i, title in enumerate(left, 1):
        ws.cell(1, i, title)
    for idx, case in enumerate(CONFIGS):
        start = n_left + 1 + idx * n_sub
        ws.cell(1, start, case["name"])
        for offset, (sub_title, _key) in enumerate(SUMMARY_SUBCOLUMNS):
            ws.cell(2, start + offset, sub_title)
    for k, scale in enumerate(SCALES):
        counts = scale_counts(scale)
        block_start = blocks[k][0]
        for mi, model_name in enumerate(MODELS):
            r = block_start + mi
            if mi == 0:
                ws.cell(block_start, 1, scale)
                ws.cell(block_start, 2, counts["I"])
                ws.cell(block_start, 3, counts["J"])
                ws.cell(block_start, 4, counts["H"])
            ws.cell(r, 5, model_name)
            for idx, case in enumerate(CONFIGS):
                match = next(
                    (x for x in rows if x.get("scale") == scale
                     and x.get("model") == model_name
                     and x.get("config") == case["name"]),
                    None,
                )
                start = n_left + 1 + idx * n_sub
                if match is None:
                    values = tuple("NA" for _ in SUMMARY_SUBCOLUMNS)
                elif match.get("status") != "OK":
                    values = tuple("FAIL" for _ in SUMMARY_SUBCOLUMNS)
                else:
                    values = tuple(
                        excel_value(match, key) for _title, key in SUMMARY_SUBCOLUMNS
                    )
                for offset, value in enumerate(values):
                    cell = ws.cell(r, start + offset, value)
                    if isinstance(value, float):
                        key = SUMMARY_SUBCOLUMNS[offset][1]
                        if key == "obj_value":
                            cell.number_format = "#,##0.00"
                        elif key == "gap_pct":
                            cell.number_format = "0.0000"
                        elif key == "cpu_s":
                            cell.number_format = "#,##0.00"
                        else:
                            cell.number_format = "#,##0"

    # 3) 統一設樣式（merge 之後）：全格框線；表頭列上色。
    for row in ws.iter_rows(min_row=1, max_row=last, min_col=1, max_col=n_cols):
        for cell in row:
            cell.border = border
            if cell.row <= 2:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 9
    for col in range(2, 5):
        ws.column_dimensions[get_column_letter(col)].width = 5
    ws.column_dimensions["E"].width = 16
    for col in range(n_left + 1, n_cols + 1):
        # Obj Value 欄位數字較長，給寬一點
        is_obj = SUMMARY_SUBCOLUMNS[(col - n_left - 1) % n_sub][1] == "obj_value"
        ws.column_dimensions[get_column_letter(col)].width = 16 if is_obj else 10
    ws.freeze_panes = f"{get_column_letter(n_left + 1)}3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18


def export_xlsx(rows: list[dict[str, Any]], path: Path, run_id: str = "",
                log_run_dir: Path | None = None) -> Path:
    """Atomically export all raw and presentation results.

    Excel is a required deliverable for this experiment, so dependency or
    write errors are intentionally propagated instead of being silently
    ignored after a long solver run.
    """
    ensure_excel_dependency()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="2E74B5")
    font = Font(bold=True, color="FFFFFF")

    _write_raw_sheet(wb, rows, fill, font, Alignment)
    _write_settings_sheet(wb, run_id, log_run_dir, fill, font, Alignment)
    for scale in SCALES:
        _write_detail_sheet(wb, scale, rows, fill, font, Alignment)
    _write_summary_sheet(wb, rows, fill, font, Alignment)

    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(
        prefix=f".{path.stem}_", suffix=".xlsx", dir=path.parent
    )
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
        # Windows：若使用者正好在 Excel 開著這個檔，os.replace 會丟
        # PermissionError。重試幾次給他關檔的機會，而不是直接讓實驗死掉。
        last_exc: Exception | None = None
        for attempt in range(EXCEL_REPLACE_RETRIES):
            try:
                os.replace(temp_path, path)
                return path
            except PermissionError as exc:
                last_exc = exc
                if attempt == 0:
                    print(f"[excel] 無法寫入 {path.name}（檔案可能正被 Excel 開啟）；"
                          f"重試中，請關閉該檔案 …")
                time.sleep(EXCEL_REPLACE_RETRY_SLEEP_SEC)
        raise PermissionError(
            f"無法寫入 {path}（重試 {EXCEL_REPLACE_RETRIES} 次仍失敗）。"
            "請關閉正在開啟該檔案的 Excel 視窗。"
        ) from last_exc
    finally:
        temp_path.unlink(missing_ok=True)


def export_xlsx_incremental(rows: list[dict[str, Any]], path: Path, run_id: str,
                            log_run_dir: Path | None) -> None:
    """跑實驗途中的即時匯出：失敗只警告，不中斷整批實驗。

    續跑的依據是 raw CSV（`write_results` 已先寫好），Excel 只是方便中途查看，
    所以途中匯出失敗（例如檔案被 Excel 鎖住）不該讓跑了好幾小時的實驗掛掉。
    最終匯出（main 結尾）仍會正常拋出錯誤。
    """
    try:
        export_xlsx(rows, path, run_id, log_run_dir)
    except Exception as exc:  # noqa: BLE001
        print(f"[excel][WARN] 中途匯出失敗，實驗繼續進行（結果已存在 CSV）："
              f"{type(exc).__name__}: {exc}")


def _safe_case_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._") or "case"


def hard_timeout_for(case: dict[str, Any]) -> float:
    """該 case 的子程序硬超時（秒）= 求解時限 + 依引擎而定的建模緩衝。"""
    buffer_s = (
        HARD_TIMEOUT_BUFFER_SEC_EXTENSIVE
        if case.get("engine") == "extensive"
        else HARD_TIMEOUT_BUFFER_SEC
    )
    return TIME_LIMIT + buffer_s


def _case_paths(scale: str, model_name: str, case: dict[str, Any],
                run_idx: int, log_run_dir: Path) -> tuple[str, Path, Path, Path, Path]:
    test_id = f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
    stem = f"{run_idx:02d}_{_safe_case_name(test_id)}"
    return (
        test_id,
        log_run_dir / f"{stem}.log",
        log_run_dir / f"{stem}_solver.log",
        log_run_dir / f".{stem}_launcher.tmp.log",
        log_run_dir / f".{stem}_row.tmp.json",
    )


def _dispose_model(model: Any) -> None:
    if model is None:
        return
    try:
        model.dispose()
    except Exception:  # noqa: BLE001
        pass


def run_one_case(portals: dict[str, Any], model_name: str, case: dict[str, Any],
                 counts: dict[str, int], scale: str, run_idx: int, total: int,
                 log_run_dir: Path) -> dict[str, Any]:
    _test_id, case_log_path, solver_log_path, _launcher, _result = _case_paths(
        scale, model_name, case, run_idx, log_run_dir
    )
    case_log_path.parent.mkdir(parents=True, exist_ok=True)
    with logging_utils.tee_output(case_log_path):
        return _run_one_case_logged(
            portals, model_name, case, counts, scale, run_idx, total,
            case_log_path, solver_log_path,
        )


def _run_one_case_logged(portals: dict[str, Any], model_name: str, case: dict[str, Any],
                         counts: dict[str, int], scale: str, run_idx: int, total: int,
                         case_log_path: Path, solver_log_path: Path) -> dict[str, Any]:
    engine = case.get("engine", "bbc")
    portal = portals["ext"] if engine == "extensive" else portals[model_name]
    test_id = f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
    row = blank_row()
    row.update({
        "scale": scale, "test_id": test_id, "model": model_name, "config": case["name"],
        "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "ev_warm_start": case["ev"], "root_seed_iters": case["seed"],
        "root_cut_rounds": case["rounds"], "use_user_cuts": case["user"],
        "pareto_enabled": case["pareto"], "multi_cut": engine != "extensive",
        "mip_focus": 0 if engine == "extensive" else DEFAULT_MIPFOCUS,
        "heuristics": 0.0 if engine == "extensive" else DEFAULT_HEURISTICS,
        "numeric_focus": 0 if engine == "extensive" else DEFAULT_NUMERIC_FOCUS,
        "branch_priority_enabled": (
            False if engine == "extensive" else DEFAULT_BRANCH_PRIORITY_ENABLED
        ),
        "branch_priority": 0 if engine == "extensive" else DEFAULT_BRANCH_PRIORITY,
        "log_path": str(case_log_path), "status": "RUNNING", "note": "",
    })
    print(f"\n[{run_idx}/{total}] scale={scale} model={model_name} config={case['name']}")
    before = snapshot_logs()
    wall_start = time.time()
    model = summary = None
    try:
        try:
            with temporary_config(case):
                spec = _SPEC_BY_NAME[model_name]
                if spec["kind"] == "mcvar":
                    model, summary = portal.run_mcvar_model(
                        scenario_size=BASE_SCENARIOS, sample_ratio=BASE_SAMPLE_RATIO,
                        time_limit=TIME_LIMIT, mip_gap=MIP_GAP,
                        alpha=RISK_ALPHA, lam=RISK_LAMBDA, compute_kpis=COMPUTE_KPIS,
                    )
                else:
                    model, summary = portal.run_dro_model(
                        ambiguity_set=spec["ambiguity_set"], scenario_size=BASE_SCENARIOS,
                        sample_ratio=BASE_SAMPLE_RATIO, time_limit=TIME_LIMIT,
                        mip_gap=MIP_GAP, alpha=RISK_ALPHA, lam=RISK_LAMBDA,
                        scope=spec["scope"], compute_kpis=COMPUTE_KPIS,
                    )
        finally:
            row["wall_s"] = f"{time.time() - wall_start:.2f}"
            try:
                moved = move_newest_log(before, solver_log_path)
                row["solver_log_path"] = str(moved) if moved else "NA"
            except Exception as log_exc:  # noqa: BLE001
                row["solver_log_path"] = "NA"
                row["note"] = f"solver log move failed: {type(log_exc).__name__}: {log_exc}"
    except Exception as exc:  # noqa: BLE001
        _dispose_model(model)
        row["status"] = "FAIL"
        prefix = f"{row['note']}; " if row["note"] else ""
        row["note"] = f"{prefix}{type(exc).__name__}: {exc}"
        print(f"  -> FAIL: {row['note']}")
        if STOP_ON_ERROR:
            raise
        return row

    if model is None or summary is None:
        _dispose_model(model)
        row["status"], row["note"] = "FAIL", "no feasible solution (see log)"
        return row

    st = summary.get("bbc_stats", {})
    fs = summary.get("first_stage")
    objective = float(summary["objective"])
    row.update({
        "obj_value": f"{objective:.6f}",
        "first_stage_decision": first_stage_string(fs),
        "best_lb": f"{float(summary['best_lb']):.6f}",
        "best_ub": f"{objective:.6f}",
        "cpu_s": f"{float(st.get('runtime', float('nan'))):.2f}",
        "num_vars": getattr(model, "NumVars", "NA"),
        "num_constrs": getattr(model, "NumConstrs", "NA"),
        "nodes": f"{float(getattr(model, 'NodeCount', float('nan'))):.0f}",
        "iterations": f"{float(getattr(model, 'IterCount', float('nan'))):.0f}",
        "gap_pct": f"{float(summary['gap_pct']):.6f}",
        "total_cuts": st.get("cuts_added", "NA"),
        "seed_cuts": st.get("seed_cuts_added", "NA"),
        "lazy_cuts": st.get("lazy_cuts_added", "NA"),
        "user_cuts": st.get("user_cuts_added", "NA"),
        "root_seed_lb": st.get("root_seed_lb", "NA"),
        "root_seed_iters_done": st.get("root_seed_iters_done", "NA"),
        "root_cut_rounds_done": st.get("root_cut_rounds_done", "NA"),
        "parallel_oracles": (
            0 if engine == "extensive"
            else st.get("parallel_oracles", getattr(cfg, "BENDERS_PARALLEL_ORACLES", "NA"))
        ),
        "oracle_solves": st.get("oracle_solves", "NA"),
        "incumbent_evals": st.get("incumbent_evals", "NA"),
        "callback_time_s": st.get("callback_time", "NA"),
        "solver_status": st.get("solver_status", "NA"),
        "status": "OK",
    })
    _dispose_model(model)
    print(f"  -> OK obj={row['obj_value']} time={row['cpu_s']}s gap={row['gap_pct']}%")
    return row


def _failed_case_row(model_name: str, case: dict[str, Any], counts: dict[str, int],
                     scale: str, case_log_path: Path, note: str,
                     wall_s: float) -> dict[str, Any]:
    test_id = f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
    row = blank_row()
    row.update({
        "scale": scale, "test_id": test_id, "model": model_name,
        "config": case["name"], "I": counts["I"], "J": counts["J"],
        "H": counts["H"], "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "wall_s": f"{wall_s:.2f}", "ev_warm_start": case["ev"],
        "root_seed_iters": case["seed"], "root_cut_rounds": case["rounds"],
        "use_user_cuts": case["user"], "pareto_enabled": case["pareto"],
        "multi_cut": case.get("engine") != "extensive",
        "parallel_oracles": (
            0 if case.get("engine") == "extensive" else cfg.BENDERS_PARALLEL_ORACLES
        ),
        "mip_focus": (
            0 if case.get("engine") == "extensive" else DEFAULT_MIPFOCUS
        ),
        "heuristics": (
            0.0 if case.get("engine") == "extensive" else DEFAULT_HEURISTICS
        ),
        "numeric_focus": (
            0 if case.get("engine") == "extensive" else DEFAULT_NUMERIC_FOCUS
        ),
        "branch_priority_enabled": (
            False if case.get("engine") == "extensive"
            else DEFAULT_BRANCH_PRIORITY_ENABLED
        ),
        "branch_priority": (
            0 if case.get("engine") == "extensive" else DEFAULT_BRANCH_PRIORITY
        ),
        "solver_status": "HARD_TIMEOUT" if "hard timeout" in note.lower() else "ERROR",
        "log_path": str(case_log_path), "status": "FAIL", "note": note,
    })
    return row


def _write_child_result(path: Path, row: dict[str, Any]) -> None:
    writing = path.with_suffix(path.suffix + ".writing")
    try:
        with writing.open("w", encoding="utf-8") as f:
            json.dump(row, f, ensure_ascii=False, indent=2)
        os.replace(writing, path)
    finally:
        writing.unlink(missing_ok=True)


def _single_case_cli(argv: list[str]) -> None:
    if len(argv) != 7:
        raise SystemExit(
            "single-case arguments: SCALE MODEL CONFIG_INDEX RUN_INDEX TOTAL LOG_DIR RESULT_JSON"
        )
    scale, model_name, config_idx_raw, run_idx_raw, total_raw, log_dir_raw, result_raw = argv
    ensure_solver_dependency()
    case = CONFIGS[int(config_idx_raw)]
    counts = scale_counts(scale)
    log_run_dir = Path(log_dir_raw).resolve()
    result_path = Path(result_raw).resolve()
    spec = _SPEC_BY_NAME[model_name]
    if case.get("engine") == "extensive":
        portals = {
            "ext": load_portal(EXTENSIVE_MODEL_PATH, "ablation_extensive_portal_child")
        }
    elif spec["kind"] == "mcvar":
        portals = {model_name: load_portal(MCVAR_MODEL_PATH, "ablation_mcvar_portal_child")}
    else:
        portals = {model_name: load_portal(DRO_MODEL_PATH, "ablation_dro_portal_child")}
    original_scale = cfg.EXPERIMENT_SCALE
    try:
        cfg.EXPERIMENT_SCALE = scale
        with scaled_instance_generation(), open(os.devnull, "w", encoding="utf-8") as sink:
            with redirect_stdout(sink), redirect_stderr(sink):
                row = run_one_case(
                    portals, model_name, case, counts, scale,
                    int(run_idx_raw), int(total_raw), log_run_dir,
                )
    finally:
        cfg.EXPERIMENT_SCALE = original_scale
    _write_child_result(result_path, row)


def run_one_case_subprocess(model_name: str, case: dict[str, Any],
                            counts: dict[str, int], scale: str,
                            run_idx: int, total: int,
                            log_run_dir: Path) -> dict[str, Any]:
    """Run one case in an isolated process and enforce a wall-clock hard cap."""
    _test_id, case_log_path, solver_log_path, launcher_path, result_path = _case_paths(
        scale, model_name, case, run_idx, log_run_dir
    )
    before = snapshot_logs()
    start = time.time()
    cmd = [
        sys.executable, str(Path(__file__).resolve()), "--single-case",
        scale, model_name, str(CONFIGS.index(case)), str(run_idx), str(total),
        str(log_run_dir), str(result_path),
    ]
    timed_out = False
    return_code = None
    hard_timeout = hard_timeout_for(case)
    with launcher_path.open("w", encoding="utf-8", newline="") as launcher:
        try:
            completed = subprocess.run(
                cmd, cwd=ROOT_DIR, stdout=launcher, stderr=subprocess.STDOUT,
                timeout=hard_timeout, check=False,
            )
            return_code = completed.returncode
        except subprocess.TimeoutExpired:
            timed_out = True

    wall_s = time.time() - start
    if not solver_log_path.exists():
        try:
            move_newest_log(before, solver_log_path)
        except Exception:  # noqa: BLE001
            pass

    try:
        if not timed_out and return_code == 0 and result_path.is_file():
            with result_path.open("r", encoding="utf-8") as f:
                row = json.load(f)
            return row

        if timed_out:
            note = f"Hard timeout: case process exceeded {hard_timeout:.0f} seconds"
        else:
            # 常見原因：記憶體不足被作業系統終止（Extensive form 在大規模
            # 可能需要數十 GB）。此時 returncode 為負值或 137。
            note = f"Case subprocess failed with exit code {return_code}"
            if return_code in (137, -9):
                note += "（很可能是記憶體不足被系統終止；Extensive form 在大規模需要大量 RAM）"

        launcher_tail = ""
        if launcher_path.is_file():
            launcher_tail = launcher_path.read_text(encoding="utf-8", errors="replace")[-8000:]
        if launcher_tail.strip():
            note = f"{note}; launcher output: {launcher_tail.strip()}"
        row = _failed_case_row(
            model_name, case, counts, scale, case_log_path, note, wall_s
        )
        if not case_log_path.exists():
            case_log_path.write_text(note + "\n", encoding="utf-8")
        else:
            with case_log_path.open("a", encoding="utf-8") as f:
                f.write("\n" + note + "\n")
        row["solver_log_path"] = str(solver_log_path) if solver_log_path.exists() else "NA"
        return row
    finally:
        result_path.unlink(missing_ok=True)
        result_path.with_suffix(result_path.suffix + ".writing").unlink(missing_ok=True)
        launcher_path.unlink(missing_ok=True)


def objective_warnings(rows: list[dict[str, Any]]) -> list[str]:
    # 同一 (scale, model) 下，六種求解配置應解到相同 objective（同一問題）。
    warnings = []
    for scale in SCALES:
        for model_name in MODELS:
            successful = [
                r for r in rows
                if r.get("scale") == scale and r.get("model") == model_name
                and r.get("status") == "OK"
            ]
            if len(successful) < 2:
                continue
            reference = successful[0]
            ref_obj = float(reference["obj_value"])
            ref_gap = float(reference["gap_pct"]) / 100.0
            for row in successful[1:]:
                obj = float(row["obj_value"])
                gap = float(row["gap_pct"]) / 100.0
                tolerance = abs(ref_obj) * ref_gap + abs(obj) * gap + 1e-6
                if abs(obj - ref_obj) > tolerance:
                    warnings.append(
                        f"[{scale}] {model_name}: {reference['config']} vs {row['config']} "
                        f"objective 不一致 ({ref_obj:.6f} vs {obj:.6f}, tol={tolerance:.6f})"
                    )
    return warnings


def expected_test_ids() -> set[str]:
    return {
        f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
        for scale in SCALES
        for model_name in MODELS
        for case in CONFIGS
    }


def validate_final_outputs(rows: list[dict[str, Any]], csv_path: Path,
                           xlsx_path: Path) -> None:
    """Fail loudly unless the final CSV, Excel and per-case logs are complete."""
    # 以「實際產生的 rows」自洽驗證（支援 RUN_ONLY_CONFIGS 過濾後的子集）。
    full_matrix = expected_test_ids()
    actual = [str(row.get("test_id")) for row in rows]
    if not rows:
        raise RuntimeError("No result rows were produced")
    if len(actual) != len(set(actual)):
        raise RuntimeError("Duplicate test_id values found in final results")
    invalid = sorted(set(actual) - full_matrix)
    if invalid:
        raise RuntimeError(f"Unexpected test_id values not in the full matrix: {invalid}")
    if not csv_path.is_file() or csv_row_count(csv_path) != len(rows):
        raise RuntimeError(f"CSV verification failed: {csv_path}")

    missing_logs = [
        row["test_id"] for row in rows
        if row.get("log_path") in (None, "", "NA")
        or not Path(str(row["log_path"])).is_file()
    ]
    if missing_logs:
        raise RuntimeError(f"Missing per-case logs: {missing_logs}")

    ensure_excel_dependency()
    from openpyxl import load_workbook

    if not xlsx_path.is_file():
        raise RuntimeError(f"Excel file was not created: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=False)
    try:
        expected_sheets = {
            "raw_results", "run_settings", *SCALES, "summary_table",
        }
        if set(wb.sheetnames) != expected_sheets:
            raise RuntimeError(
                f"Excel sheet mismatch: expected={sorted(expected_sheets)}, "
                f"actual={wb.sheetnames}"
            )
        raw = wb["raw_results"]
        headers = [raw.cell(1, col).value for col in range(1, len(FIELDNAMES) + 1)]
        if headers != FIELDNAMES or raw.max_row != len(rows) + 1:
            raise RuntimeError("Excel raw_results is incomplete or has wrong columns")
        for scale in SCALES:
            produced = sum(1 for r in rows if r.get("scale") == scale)
            if wb[scale].max_row != 1 + produced:
                raise RuntimeError(f"Excel {scale} sheet is incomplete")
        if wb["summary_table"].max_row != 2 + len(SCALES) * len(MODELS):
            raise RuntimeError("Excel summary_table is incomplete")
    finally:
        wb.close()


def main() -> None:
    ensure_solver_dependency()
    ensure_excel_dependency()

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = RESULT_PREFIX
    if RUN_ONLY_CONFIGS:
        prefix = f"{RESULT_PREFIX}_only_" + "_".join(
            _safe_case_name(name) for name in RUN_ONLY_CONFIGS
        )
    csv_path = RESULT_DIR / f"{prefix}_raw_{timestamp}.csv"
    xlsx_path = RESULT_DIR / f"{prefix}_{timestamp}.xlsx"
    log_run_dir = LOG_SUBDIR / timestamp

    # Preflight the mandatory Excel dependency and destination before any
    # optimization consumes hours of compute time.
    log_run_dir.mkdir(parents=True, exist_ok=False)
    export_xlsx([], xlsx_path, timestamp, log_run_dir)

    print("=" * 72)
    print("B&BC ABLATION EXPERIMENT (實驗三) — small / medium / large")
    print("=" * 72)
    print(f"data  : {cfg.DISASTER_CSV} / {cfg.CCP_CSV} / {cfg.HOSPITAL_CSV}")
    print(f"scales={SCALES}  (資源縮放基準 I/J/H = "
          f"{cfg.PARAM_CALIB_N_DISASTER}/{cfg.PARAM_CALIB_N_CCP}/{cfg.PARAM_CALIB_N_HOSPITAL}, "
          f"per-CCP 模式 = {cfg.CCP_UPPER_BOUND_SCALING})")
    for sc in SCALES:
        c = scale_counts(sc)
        rs = cfg.resolve_scale(sc)
        print(f"  {sc:6}: I={c['I']:3d} J={c['J']:3d} H={c['H']:3d}  "
              f"s_D={rs['demand_scale']:.3f} s_J={rs['ccp_scale']:.3f} "
              f"s_H={rs['hospital_scale']:.3f}")
    print(f"configs={[case['name'] for case in active_configs()]}")
    print(f"models={MODELS} S={BASE_SCENARIOS} T={BASE_TIME_PERIODS}")
    n_cases = len(SCALES) * len(active_configs()) * len(MODELS)
    print(f"mip_gap={MIP_GAP} (達到即提早結束) time_limit={TIME_LIMIT}s "
          f"cases={n_cases}")
    n_ext = len(SCALES) * len(MODELS) * sum(
        1 for c in active_configs() if c.get("engine") == "extensive"
    )
    worst_h = (
        (n_cases - n_ext) * hard_timeout_for({}) + n_ext * hard_timeout_for({"engine": "extensive"})
    ) / 3600.0
    print(f"最壞情況總時間 ≈ {worst_h:.0f} 小時"
          f"（實際會因為 gap≤{MIP_GAP:.0%} 提早結束而少很多）")
    print(f"硬超時：B&BC 類 {hard_timeout_for({}) / 60:.0f} 分鐘/case、"
          f"Extensive {hard_timeout_for({'engine': 'extensive'}) / 60:.0f} 分鐘/case")
    print(f"CSV   : {csv_path}\nExcel : {xlsx_path}\nLogs  : {log_run_dir}")

    prior_ok: dict[str, dict[str, Any]] = {}
    if RESUME_FROM_CSV:
        prior_path = Path(RESUME_FROM_CSV)
        if not prior_path.is_absolute():
            prior_path = RESULT_DIR / RESUME_FROM_CSV
        for pr in _load_prior_rows(prior_path):
            if pr.get("status") == "OK" and pr.get("test_id"):
                prior_ok[pr["test_id"]] = pr
        print(f"[RESUME] 從 {prior_path} 載入 {len(prior_ok)} 個已完成(OK) case；"
              f"只補跑缺少或非 OK 的 case。")

    rows: list[dict[str, Any]] = []
    total = len(SCALES) * len(active_configs()) * len(MODELS)

    run_idx = 0
    try:
        for scale in SCALES:
            counts = scale_counts(scale)
            for model_name in MODELS:
                for case in active_configs():
                    run_idx += 1
                    test_id = f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
                    if test_id in prior_ok:
                        print(f"[{run_idx}/{total}] SKIP（沿用上次已完成）{test_id}")
                        rows.append(prior_ok[test_id])
                    else:
                        print(
                            f"\n[{run_idx}/{total}] scale={scale} "
                            f"model={model_name} config={case['name']}"
                        )
                        row = run_one_case_subprocess(
                            model_name, case, counts, scale, run_idx, total, log_run_dir,
                        )
                        rows.append(row)
                        print(
                            f"  -> {row['status']} obj={row['obj_value']} "
                            f"time={row['cpu_s']}s gap={row['gap_pct']}%"
                        )
                    # 每個 case 結束就落地一次；中斷後可由此 CSV 續跑。
                    write_results(csv_path, rows)
                    export_xlsx_incremental(rows, xlsx_path, timestamp, log_run_dir)
    except KeyboardInterrupt:
        # 使用者中斷：已完成的 case 已經寫進 csv_path，直接告訴他怎麼續跑。
        write_results(csv_path, rows)
        export_xlsx_incremental(rows, xlsx_path, timestamp, log_run_dir)
        print("\n" + "=" * 72)
        print("[INTERRUPTED] 已中斷。已完成的結果都已保存。")
        print(f"  已完成 OK: {sum(r['status'] == 'OK' for r in rows)}/{total}")
        print(f"  CSV      : {csv_path}")
        print("\n續跑方式：把下面這行貼回本檔的參數區，再執行一次。")
        print(f'    RESUME_FROM_CSV = "{csv_path.name}"')
        print("=" * 72)
        return

    warnings = objective_warnings(rows)
    export_xlsx(rows, xlsx_path, timestamp, log_run_dir)
    try:
        validate_final_outputs(rows, csv_path, xlsx_path)
    except Exception as _exc:  # noqa: BLE001  # resume/部分重跑時沿用列的舊 log 可能不在，容忍
        print(f"[validate] 略過完整性檢查（resume 或部分重跑）: {_exc}")
    print("\n" + "=" * 72)
    n_ok = sum(r["status"] == "OK" for r in rows)
    print(f"Done: {n_ok}/{len(rows)} cases OK")
    for warning in warnings:
        print(f"[WARN] {warning}")
    print(f"CSV   : {csv_path}\nExcel : {xlsx_path}\nLogs  : {log_run_dir}")
    if n_ok < len(expected_test_ids()):
        print("\n" + "-" * 72)
        print("尚有 case 未完成。續跑方式：把下面這行貼回本檔的參數區，再執行一次。")
        print(f'    RESUME_FROM_CSV = "{csv_path.name}"')
        print("-" * 72)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--single-case":
        _single_case_cli(sys.argv[2:])
    else:
        main()
