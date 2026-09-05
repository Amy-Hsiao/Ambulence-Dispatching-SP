"""batch_bbc_full_stack.py — BBC + WS + RS + UC + Pareto + LBF + VI 完整堆疊實驗。

模型：SP + MCVaR（risk_core 的 "mcvar"，λ 混合期望值與 CVaR）。不含 DRO。
規模：small / medium / large 各一次，共三個實驗，每個 2 小時。

堆疊內容（全部同時開啟）
-----------------------
  BBC     Branch-and-Benders-Cut（integer incumbent 上加 lazy cut）
  WS      EV warm start：以期望值問題的一階解當 master 初始 incumbent
  RS      Root seeding：正式分支前先在 LP 鬆弛 master 上墊 ordinary cut
  UC      User cuts：root 節點分數解上加使用者切割
  Pareto  Papadakos core point，產生 Pareto-optimal cut
  LBF     Lower Bounding Functional：平均情境的二階 LP 內嵌進 master（Jensen）
  VI      八條有效不等式（docs/有效不等式_實作規格.docx，config.VI_* 控制）

執行
----
    python "run experiment/batch_bbc_full_stack.py"
    python "run experiment/batch_bbc_full_stack.py" --scales small --time-limit 600
    python "run experiment/batch_bbc_full_stack.py" --no-vi        # 對照組（關掉 VI）

輸出
----
  experiment result/BBC_full_stack_<timestamp>.xlsx
      Summary  規定的五欄：UB、LB、Time、Gap(%)、Nodes（外加規模與狀態）
      Detail   切割數、root seeding、VI 診斷、一階解等完整欄位
      Config   本次使用的每一項設定，供重現

  experiment result/BBC_full_stack_raw_<timestamp>.csv
      每跑完一個規模就立刻追加。三個規模共約六小時，中途若掛掉，
      已完成的結果不會跟著消失 —— Excel 也會在每個規模結束後重寫一次。
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

for _st in (sys.stdout, sys.stderr):
    try:
        _st.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

ROOT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT_DIR / "model core"))
RESULT_DIR = ROOT_DIR / "experiment result"

import gurobipy as gp

import config
import lshaped_core
import risk_core

RESULT_PREFIX = "BBC_full_stack"
DEFAULT_SCALES = ("small", "medium", "large")
DEFAULT_TIME_LIMIT = 7200.0          # 兩小時／規模


# ===================================================================== #
# 單一規模                                                               #
# ===================================================================== #
def run_one(scale: str, time_limit: float, mip_gap: float, n_scen: int | None,
            vi_on: bool, risk_type: str) -> dict:
    """跑一個規模，回傳一列結果。"""
    print("\n" + "=" * 100)
    print(f" 規模 {scale}　模型 {risk_type}　時限 {time_limit / 3600:.2f} 小時"
          f"　VI {'ON' if vi_on else 'OFF'}")
    print("=" * 100, flush=True)

    t_build = time.time()
    instance = config.generate_data(scale=scale)
    sets = instance["sets"]
    S_all = sets["S"]
    S_selected = S_all if n_scen is None else S_all[:n_scen]
    build_secs = time.time() - t_build

    dims = dict(scale=scale, n_I=len(sets["I"]), n_J=len(sets["J"]),
                n_H=len(sets["H"]), n_T=len(sets["T"]), n_S=len(S_selected))
    print(f" |I|={dims['n_I']} |J|={dims['n_J']} |H|={dims['n_H']} "
          f"|T|={dims['n_T']} |S|={dims['n_S']}　(資料生成 {build_secs:.1f}s)", flush=True)

    risk_cfg = risk_core.make_risk_cfg(risk_type)
    vi_cfg = None if vi_on else {"all": False}

    row = dict(dims)
    row.update(model=risk_type, vi=("on" if vi_on else "off"),
               time_limit=time_limit, mip_gap=mip_gap,
               started_at=datetime.now().isoformat(timespec="seconds"))
    if risk_cfg is not None:
        row.update(risk_alpha=risk_cfg["alpha"], risk_lambda=risk_cfg["lambda"])

    t0 = time.time()
    try:
        res = lshaped_core.solve_bbc(
            instance, S_selected,
            time_limit=time_limit,
            mip_gap=mip_gap,
            risk_cfg=risk_cfg,
            vi_cfg=vi_cfg,
            # 以下五項都讀 config 預設（全部為開），明示是為了讓實驗紀錄自我說明
            multi_cut=config.BENDERS_MULTI_CUT,
            ev_warm_start=config.BENDERS_EV_WARM_START,          # WS
            root_seed_iters=config.BENDERS_ROOT_SEED_ITERS,      # RS
            use_user_cuts=config.BENDERS_USE_USER_CUTS,          # UC
            root_cut_rounds=config.BENDERS_ROOT_CUT_ROUNDS,
            pareto_enabled=config.BENDERS_PARETO_ENABLED,        # Pareto
            lbf_enabled=config.BENDERS_LBF_ENABLED,              # LBF
            verbose=True,
        )
    except gp.GurobiError as exc:
        row.update(status=f"GurobiError", error=str(exc),
                   runtime=time.time() - t0)
        print(f"\n [!] Gurobi 錯誤：{exc}", flush=True)
        return row
    except Exception as exc:                                  # noqa: BLE001
        row.update(status="Exception", error=f"{type(exc).__name__}: {exc}",
                   runtime=time.time() - t0)
        traceback.print_exc()
        return row

    fs = res.get("first_stage") or {}
    row.update(
        # ---- 規定的五欄 ----
        UB=res.get("best_ub"),
        LB=res.get("best_lb"),
        Time=res.get("runtime"),
        Gap_pct=res.get("gap_pct"),
        Nodes=res.get("nodes"),
        # ---- 診斷 ----
        status=res.get("status"),
        iterations=res.get("iterations"),
        cuts_added=res.get("cuts_added"),
        seed_cuts=res.get("seed_cuts_added"),
        user_cuts=res.get("user_cuts_added"),
        lazy_cuts=res.get("lazy_cuts_added"),
        root_seed_iters_done=res.get("root_seed_iters_done"),
        root_seed_lb=res.get("root_seed_lb"),
        root_seed_stop=res.get("root_seed_stop_reason"),
        root_seed_time=res.get("root_seed_time"),
        root_cut_rounds_done=res.get("root_cut_rounds_done"),
        oracle_solves=res.get("oracle_solves"),
        cache_hits=res.get("cache_hits"),
        callback_time=res.get("callback_time"),
        vi8_rows=res.get("vi8_rows"),
        vi_flags=json.dumps(res.get("vi_flags") or {}, ensure_ascii=False),
        vi_theta_lb_min=(min(res["vi_theta_lb"].values())
                         if res.get("vi_theta_lb") else None),
        vi_theta_lb_n=len(res.get("vi_theta_lb") or {}),
        sum_X=sum(fs.get("X", {}).values()) if fs else None,
        sum_V=sum(fs.get("V", {}).values()) if fs else None,
        sum_U=sum(fs.get("U", {}).values()) if fs else None,
        sum_Y=sum(fs.get("Y", {}).values()) if fs else None,
        finished_at=datetime.now().isoformat(timespec="seconds"),
    )
    print(f"\n [完成] {scale}: UB={row['UB']} LB={row['LB']} "
          f"Gap={row['Gap_pct']}% Time={row['Time']:.1f}s Nodes={row['Nodes']:,.0f}",
          flush=True)
    return row


# ===================================================================== #
# 輸出                                                                   #
# ===================================================================== #
SUMMARY_COLS = [
    ("scale", "Scale"), ("n_I", "|I|"), ("n_J", "|J|"), ("n_H", "|H|"),
    ("n_T", "|T|"), ("n_S", "|S|"),
    ("UB", "UB"), ("LB", "LB"), ("Time", "Time (s)"),
    ("Gap_pct", "Gap (%)"), ("Nodes", "Nodes"), ("status", "Status"),
]
DETAIL_SKIP = {"vi_flags"}


def write_csv(rows: list[dict], path: Path) -> None:
    if not rows:
        return
    keys: list[str] = []
    for r in rows:
        for k in r:
            if k not in keys:
                keys.append(k)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(rows: list[dict], path: Path, meta: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor="E8ECEA")
    head_font = Font(bold=True)
    thin = Side(style="thin", color="C8CFCB")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws.append([lbl for _, lbl in SUMMARY_COLS])
    for r in rows:
        ws.append([r.get(k) for k, _ in SUMMARY_COLS])
    for c in range(1, len(SUMMARY_COLS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.border = head_fill, head_font, box
        cell.alignment = Alignment(horizontal="center")
    for ri in range(2, len(rows) + 2):
        for ci, (key, _) in enumerate(SUMMARY_COLS, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = box
            if key in ("UB", "LB"):
                cell.number_format = "#,##0.00"
            elif key == "Time":
                cell.number_format = "#,##0.0"
            elif key == "Gap_pct":
                cell.number_format = "0.000"
            elif key == "Nodes":
                cell.number_format = "#,##0"
    widths = [10, 6, 6, 6, 6, 6, 18, 18, 11, 10, 12, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    # ---------------- Detail ----------------
    ws2 = wb.create_sheet("Detail")
    keys: list[str] = []
    for r in rows:
        for k in r:
            if k not in keys and k not in DETAIL_SKIP:
                keys.append(k)
    ws2.append(keys)
    for r in rows:
        ws2.append([r.get(k) for k in keys])
    for c in range(1, len(keys) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill, cell.font = head_fill, head_font
        ws2.column_dimensions[get_column_letter(c)].width = max(12, min(26, len(keys[c - 1]) + 4))
    ws2.freeze_panes = "A2"

    # ---------------- Config ----------------
    ws3 = wb.create_sheet("Config")
    ws3.append(["設定", "值"])
    for k, v in meta.items():
        ws3.append([k, v if isinstance(v, (int, float, str)) else json.dumps(v, ensure_ascii=False)])
    for c in (1, 2):
        cell = ws3.cell(row=1, column=c)
        cell.fill, cell.font = head_fill, head_font
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 62
    ws3.freeze_panes = "A2"

    wb.save(path)


def collect_meta(args, scales) -> dict:
    vi_names = ["VI_ENABLED", "VI_1_ROADCAP_IJ", "VI_2_ROADCAP_JH", "VI_3_HOSP_MERGE",
                "VI_4_STAFF_UB", "VI_5_OPEN_USE", "VI_6_THETA_LB", "VI_7_THETA_UB",
                "VI_8_AGG_RELAX", "VI_APPLY_TO_LBF", "VI_THETA_LB_TIME_LIMIT"]
    ben_names = ["BENDERS_MULTI_CUT", "BENDERS_EV_WARM_START", "BENDERS_ROOT_SEED_ITERS",
                 "BENDERS_ROOT_SEED_STALL_ROUNDS", "BENDERS_ROOT_SEED_LB_REL_TOL",
                 "BENDERS_USE_USER_CUTS", "BENDERS_ROOT_CUT_ROUNDS",
                 "BENDERS_PARETO_ENABLED", "BENDERS_PAPADAKOS_BLEND",
                 "BENDERS_LBF_ENABLED", "BENDERS_PARALLEL_ORACLES",
                 "BENDERS_MIPFOCUS", "BENDERS_HEURISTICS", "BENDERS_NUMERIC_FOCUS",
                 "BENDERS_X_BRANCH_PRIORITY_ENABLED", "BENDERS_X_BRANCH_PRIORITY",
                 "BENDERS_INCUMBENT_EARLY_TERMINATION"]
    meta = {
        "執行時間": datetime.now().isoformat(timespec="seconds"),
        "堆疊": "BBC + WS + RS + UC + Pareto + LBF + VI",
        "模型": args.risk,
        "規模": ", ".join(scales),
        "每個規模時限(秒)": args.time_limit,
        "MIPGap": args.mip_gap,
        "情境數": args.scenarios if args.scenarios else config.SCENARIOS,
        "VI 總開關(本次)": "on" if not args.no_vi else "off（對照組）",
        "gurobi": ".".join(str(v) for v in gp.gurobi.version()),
        "python": sys.version.split()[0],
        "MASTER_SEED": getattr(config, "MASTER_SEED", None),
        "RISK_ALPHA": getattr(config, "RISK_ALPHA", None),
        "RISK_LAMBDA": getattr(config, "RISK_LAMBDA", None),
    }
    for k in vi_names + ben_names:
        meta[k] = getattr(config, k, None)
    return meta


# ===================================================================== #
# main                                                                  #
# ===================================================================== #
def main(argv=None):
    ap = argparse.ArgumentParser(
        description="BBC + WS + RS + UC + Pareto + LBF + VI 完整堆疊實驗（SP+MCVaR）")
    ap.add_argument("--scales", default=",".join(DEFAULT_SCALES),
                    help="要跑的規模，逗號分隔（預設 small,medium,large）")
    ap.add_argument("--risk", default="mcvar", choices=["mcvar", "sp"],
                    help="模型；預設 mcvar（= SP+MCVaR）")
    ap.add_argument("--time-limit", type=float, default=DEFAULT_TIME_LIMIT,
                    help="每個規模的時限（秒），預設 7200 = 2 小時")
    ap.add_argument("--mip-gap", type=float, default=None,
                    help="預設讀 config.SP_MIP_GAP")
    ap.add_argument("--scenarios", type=int, default=None, help="只取前 N 個情境")
    ap.add_argument("--no-vi", action="store_true",
                    help="關閉全部 VI（跑對照組用；其餘堆疊不變）")
    ap.add_argument("--tag", default="", help="附加在輸出檔名後的標記")
    args = ap.parse_args(argv)

    scales = [x.strip() for x in args.scales.split(",") if x.strip()]
    bad = [x for x in scales if x not in config.SCALE_PROFILES]
    if bad:
        ap.error(f"未知的規模 {bad}；可用：{', '.join(config.SCALE_PROFILES)}")
    if args.time_limit <= 0:
        ap.error("--time-limit 必須大於 0")
    if args.scenarios is not None and args.scenarios < 1:
        ap.error("--scenarios 至少要 1")
    mip_gap = config.SP_MIP_GAP if args.mip_gap is None else args.mip_gap

    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = f"_{args.tag}" if args.tag else ("_noVI" if args.no_vi else "")
    csv_path = RESULT_DIR / f"{RESULT_PREFIX}_raw{tag}_{stamp}.csv"
    xlsx_path = RESULT_DIR / f"{RESULT_PREFIX}{tag}_{stamp}.xlsx"
    meta = collect_meta(args, scales)

    print("=" * 100)
    print(" BBC + WS + RS + UC + Pareto + LBF + VI　完整堆疊實驗")
    print("=" * 100)
    print(f" 模型      : {args.risk}"
          + (f"（α={config.RISK_ALPHA}, λ={config.RISK_LAMBDA}）"
             if args.risk == "mcvar" else ""))
    print(f" 規模      : {', '.join(scales)}（共 {len(scales)} 個實驗）")
    print(f" 每個時限  : {args.time_limit:,.0f} 秒 = {args.time_limit / 3600:.2f} 小時"
          f"　　預估總時長 {len(scales) * args.time_limit / 3600:.1f} 小時")
    print(f" MIPGap    : {mip_gap}")
    print(f" VI        : {'關閉（對照組）' if args.no_vi else '開啟'}")
    print(f" 輸出      : {xlsx_path.name}")
    print(f"             {csv_path.name}（每個規模跑完就寫一次）")
    print("=" * 100, flush=True)

    rows: list[dict] = []
    t_all = time.time()
    for k, scale in enumerate(scales, start=1):
        print(f"\n>>> [{k}/{len(scales)}] {scale}", flush=True)
        row = run_one(scale, args.time_limit, mip_gap, args.scenarios,
                      vi_on=not args.no_vi, risk_type=args.risk)
        rows.append(row)
        # 每跑完一個規模就落地一次：三個規模共約六小時，中途掛掉不能全部白跑
        try:
            write_csv(rows, csv_path)
            write_xlsx(rows, xlsx_path, meta)
            print(f" [已寫入] {csv_path.name} / {xlsx_path.name}", flush=True)
        except Exception as exc:                              # noqa: BLE001
            print(f" [!] 寫檔失敗（{type(exc).__name__}: {exc}），結果仍在記憶體中，"
                  f"下一個規模結束時會再試一次。", flush=True)

    # ---------------- 收尾 ----------------
    print("\n" + "=" * 100)
    print(" 全部完成　總計 {:.2f} 小時".format((time.time() - t_all) / 3600))
    print("=" * 100)
    hdr = f" {'Scale':<9}{'UB':>20}{'LB':>20}{'Time(s)':>11}{'Gap(%)':>9}{'Nodes':>12}  Status"
    print(hdr)
    print("-" * len(hdr))
    for r in rows:
        ub = f"{r['UB']:,.2f}" if isinstance(r.get("UB"), (int, float)) else "—"
        lb = f"{r['LB']:,.2f}" if isinstance(r.get("LB"), (int, float)) else "—"
        tm = f"{r['Time']:,.1f}" if isinstance(r.get("Time"), (int, float)) else "—"
        gp_ = f"{r['Gap_pct']:.3f}" if isinstance(r.get("Gap_pct"), (int, float)) else "—"
        nd = f"{r['Nodes']:,.0f}" if isinstance(r.get("Nodes"), (int, float)) else "—"
        print(f" {r.get('scale',''):<9}{ub:>20}{lb:>20}{tm:>11}{gp_:>9}{nd:>12}"
              f"  {r.get('status','')}")
    print("=" * 100)
    print(f" Excel : {xlsx_path}")
    print(f" CSV   : {csv_path}")
    failed = [r for r in rows if r.get("UB") is None]
    if failed:
        print(f" 注意：{len(failed)} 個規模沒有取得可行解，詳見 Detail 分頁的 error 欄。")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
