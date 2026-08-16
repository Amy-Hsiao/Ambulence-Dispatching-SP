#!/usr/bin/env python3
"""候選 CCP 數量 ablation：Extensive form vs B&BC(全部加速技術) × |J| = 50/40/30/20。

目的（承接 plan/17 診斷）
------------------------
plan/17 診斷指出：放大規模後全部 case 跑滿 2 小時、gap 45~74%，主因之一是
候選 CCP 從 10 個變成 50 個，一階 0-1 決策空間從 2^10 爆炸到 2^50。
本實驗直接驗證這個假設 —— 固定所有其他條件，只掃描候選點數量 |J|，
看求解難度如何隨一階搜尋空間變化。

⚠️ 重要澄清：模型**沒有**任何「必須開幾座 CCP」的限制式。
   一階的限制式只有這四條：
       Σ V[j] ≤ total_available_staff             （全域醫護池）
       Σ U[j] ≤ total_available_ccp_ambulances    （全域救護車池）
       V[j] ≤ ccp_staff_upper_bound[j] · X[j]     （連結式）
       U[j] ≤ ccp_ambulance_upper_bound[j] · X[j] （連結式）
   求解器可以自由開 0 到 |J| 座，實際開幾座是「固定開設成本 vs 未服務罰金」
   權衡出來的結果（上一輪實測開了 5~6 座）。
   每個 case 實際開了幾座記錄在 raw_results 的 opened_ccps 欄位。

實驗矩陣
--------
* 求解配置（2 種，不是 6 種）
    - Extensive         : 單體模型，無任何加速
    - BBC+WS+RS+UC+Pareto : B&BC 全部加速技術都加上去
* 候選 CCP 數 |J| : 50 / 40 / 30 / 20（巢狀：20 ⊂ 30 ⊂ 40 ⊂ 50）
* 規模            : small / medium / large（災區 70 / 100 / 130，醫院 18）
  → 3 × 4 × 2 = 24 個 case

哪些參數隨 |J| 改變、哪些不改變（重要）
--------------------------------------
預設 ``CCP_UPPER_BOUND_SCALING = "demand_only"``：

  **不隨 |J| 改變**（per-CCP 上限是「設施的物理屬性」，不會因為規劃時
  多列或少列幾個候選地點就變大變小）：
    ccp_staff_upper_bound、ccp_ambulance_upper_bound、
    ccp_supply_upper_bound、ccp_physical_capacity_by_severity
  **不隨 |J| 改變**（全域資源池由災區數 |I| 決定）：
    total_available_staff、total_available_ccp_ambulances
  **不隨 |J| 改變**（單價 / 罰金 / 速率 / 機率）：全部維持

  → 於是資源條件在 |J| = 50/40/30/20 完全相同，
    **唯一改變的是一階 0-1 決策的搜尋空間**：

        |J|=50 → 2^50 ≈ 1.13e15
        |J|=40 → 2^40 ≈ 1.10e12
        |J|=30 → 2^30 ≈ 1.07e09
        |J|=20 → 2^20 ≈ 1.05e06

    這正是本實驗要隔離的變因。

  另一種語意 ``per_ccp_load``（把 per-CCP 上限 ×(10/|J|)）會讓每座 CCP 的
  容量隨候選點數變小，等於同時改變了「候選點數量」與「每座設施規模」兩件事，
  且在物理上難以解釋（設施容量不該因為多列幾個候選地點就縮水）—— 故不採用。
  若要對照可改 ``CCP_SCALING_MODE`` 這個開關。

  ⚠️ 不論用哪個模式，程式在跑任何求解前都會檢查兩個設計不變量：
     (1) 全域醫護池 < Σ per-CCP 醫護上限（全域池必須是綁定約束）
     (2) Σ per-CCP 救護車上限 ≥ 全域救護車池（否則池子永遠用不完）

輸出
----
* experiment result/CCP_count_ablation_raw_<ts>.csv
* experiment result/CCP_count_ablation_<ts>.xlsx
    - raw_results   : 全部欄位
    - run_settings  : 實驗設定（可追溯）
    - parameters    : 各 |J| 的縮放參數對照 + 不變量檢查
    - summary_table : **UB / LB / Time / Gap / Nodes**（依需求）
* logs/ccp count ablation/<ts>/  每個 case 兩份 log

用法
----
    python "run experiment/batch_ccp_count_ablation.py"

中斷後把最新的 raw CSV 檔名填進 RESUME_FROM_CSV 即可續跑；留空 = 從頭跑。
"""
from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Any


# =============================================================================
# 參數設定區
# =============================================================================

SCALES = ["small", "medium", "large"]      # 災區 70 / 100 / 130（醫院固定 18）
CCP_COUNTS = [50, 40, 30, 20]              # 候選 CCP 數（巢狀：20 ⊂ 30 ⊂ 40 ⊂ 50）

MODEL_NAME = "SP+MCVaR"                    # 只跑一個 model 維度，聚焦 |J| 效應

BASE_SCENARIOS = 50                        # 與主實驗一致（老師指示 S=50）
BASE_TIME_PERIODS = 8
BASE_SAMPLE_RATIO = 1.0
BASE_DEMAND_MULTIPLIER = 1.0
BASE_ROAD_CAPACITY_MULTIPLIER = 1.0
BASE_HOSPITAL_CAPACITY_MULTIPLIER = 1.0

RISK_ALPHA = 0.9
RISK_LAMBDA = 0.5

TIME_LIMIT = 7200.0    # 每個 case 求解時間上限 = 2 小時（依需求）
MIP_GAP = 0.01         # 達到 1% 即提早結束

# per-CCP 上限的縮放語意（見檔頭說明）。"demand_only" = 不隨 |J| 稀釋。
CCP_SCALING_MODE = "demand_only"           # "demand_only" | "per_ccp_load"

# 子程序硬超時 = 求解時限 + 建模緩衝（Extensive 建模遠慢於 B&BC）
HARD_TIMEOUT_BUFFER_SEC = 1800.0
HARD_TIMEOUT_BUFFER_SEC_EXTENSIVE = 5400.0

COMPUTE_KPIS = False
STOP_ON_ERROR = False

# Excel 檔被 Excel.exe 鎖住時的重試設定
EXCEL_REPLACE_RETRIES = 5
EXCEL_REPLACE_RETRY_SLEEP_SEC = 3.0

RESULT_PREFIX = "CCP_count_ablation"
LOG_SUBDIR_NAME = "ccp count ablation"

# 續跑：留空 = 從頭跑全部。中斷後把最新的 raw CSV 檔名貼進來即可續跑。
RESUME_FROM_CSV: str = ""


# =============================================================================
# Setup
# =============================================================================

ROOT_DIR = Path(__file__).resolve().parents[1]
LOG_DIR = ROOT_DIR / "logs"
LOG_SUBDIR = LOG_DIR / LOG_SUBDIR_NAME
RESULT_DIR = ROOT_DIR / "experiment result"

LOCAL_PYTHON_PACKAGE_CANDIDATES = [
    ROOT_DIR / ".codex_spreadsheet" / "python_packages",
    Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime"
    / "dependencies" / "python" / "Lib" / "site-packages",
]
if os.environ.get("CODEX_PRIMARY_PYTHON_PACKAGES"):
    LOCAL_PYTHON_PACKAGE_CANDIDATES.insert(
        0, Path(os.environ["CODEX_PRIMARY_PYTHON_PACKAGES"])
    )
for package_dir in reversed(LOCAL_PYTHON_PACKAGE_CANDIDATES):
    if package_dir.exists():
        sys.path.insert(0, str(package_dir))

os.chdir(ROOT_DIR)
for _p in (str(ROOT_DIR / "model core"), str(ROOT_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import config as cfg  # noqa: E402
import logging_utils  # noqa: E402

MCVAR_MODEL_PATH = ROOT_DIR / "model portal" / "mcvar bbc.py"
EXTENSIVE_MODEL_PATH = ROOT_DIR / "model portal" / "extensive_dro.py"

DEFAULT_ROOT_SEED_ITERS = int(cfg.BENDERS_ROOT_SEED_ITERS)
DEFAULT_ROOT_CUT_ROUNDS = int(cfg.BENDERS_ROOT_CUT_ROUNDS)
DEFAULT_MIPFOCUS = int(cfg.BENDERS_MIPFOCUS)
DEFAULT_HEURISTICS = float(cfg.BENDERS_HEURISTICS)
DEFAULT_NUMERIC_FOCUS = int(cfg.BENDERS_NUMERIC_FOCUS)
DEFAULT_BRANCH_PRIORITY_ENABLED = bool(cfg.BENDERS_X_BRANCH_PRIORITY_ENABLED)
DEFAULT_BRANCH_PRIORITY = int(cfg.BENDERS_X_BRANCH_PRIORITY)

# 只比較「完全沒加速」與「全部加速都加上去」兩端
CONFIGS = [
    {"name": "Extensive", "engine": "extensive",
     "ev": False, "seed": 0, "rounds": 0, "user": False, "pareto": False},
    {"name": "BBC+WS+RS+UC+Pareto", "engine": "bbc",
     "ev": True, "seed": DEFAULT_ROOT_SEED_ITERS, "rounds": DEFAULT_ROOT_CUT_ROUNDS,
     "user": True, "pareto": True},
]

FIELDNAMES = [
    "scale", "test_id", "ccp_count", "config", "I", "J", "H", "S", "T",
    "first_stage_space", "obj_value", "best_lb", "best_ub", "gap_pct", "cpu_s",
    "wall_s", "nodes", "iterations", "num_vars", "num_constrs",
    "opened_ccps", "first_stage_decision", "first_stage_cost",
    "total_cuts", "seed_cuts", "lazy_cuts", "user_cuts",
    "root_seed_lb", "root_seed_iters_done", "root_seed_stop_reason",
    "root_seed_time_s", "root_cut_rounds_done", "oracle_solves",
    "incumbent_evals", "callback_time_s", "parallel_oracles",
    "ccp_staff_ub", "ccp_ambulance_ub", "total_staff_pool", "total_amb_pool",
    "solver_status", "log_path", "solver_log_path", "status", "note",
]

TEXT_KEYS = {
    "scale", "test_id", "config", "first_stage_decision", "solver_status",
    "log_path", "solver_log_path", "status", "note", "root_seed_stop_reason",
}
INTEGER_KEYS = {
    "ccp_count", "I", "J", "H", "S", "T", "first_stage_space", "nodes", "iterations",
    "num_vars", "num_constrs", "opened_ccps", "total_cuts", "seed_cuts",
    "lazy_cuts", "user_cuts", "root_seed_iters_done", "root_cut_rounds_done",
    "oracle_solves", "incumbent_evals", "parallel_oracles",
}


def blank_row() -> dict[str, Any]:
    return {key: "NA" for key in FIELDNAMES}


def ensure_excel_dependency() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "需要 openpyxl 才能輸出 Excel，但目前的 Python 沒有安裝。"
            f"請先執行： {sys.executable!r} -m pip install openpyxl"
        ) from exc


def ensure_solver_dependency() -> None:
    try:
        import gurobipy  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            f"gurobipy 在目前的 Python ({sys.executable!r}) 無法載入。"
        ) from exc


def load_portal(path: Path, module_name: str) -> Any:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"無法載入 portal：{path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


# =============================================================================
# 規模 / 參數
# =============================================================================

def scale_counts(scale: str, ccp_count: int) -> dict[str, int]:
    profile = cfg.SCALE_PROFILES[scale]
    return {
        "I": int(profile["n_disaster"]),
        "J": int(ccp_count),
        "H": int(profile["n_hospital"]),
    }


def first_stage_space(ccp_count: int) -> int:
    """一階 0-1 決策的搜尋空間大小 = 2^|J|。

    ⚠️ 模型**沒有**任何「必須開幾座」的限制式。一階只有：
          Σ V[j] ≤ total_available_staff            （全域醫護池）
          Σ U[j] ≤ total_available_ccp_ambulances   （全域救護車池）
          V[j] ≤ ccp_staff_upper_bound[j] · X[j]    （連結式）
          U[j] ≤ ccp_ambulance_upper_bound[j] · X[j]（連結式）
       求解器可以自由開 0 到 |J| 座；實際開幾座由固定開設成本與資源池
       自行權衡決定（上一輪實測開了 5~6 座）。
       因此這裡用不含任何假設的 2^|J| 表示一階搜尋空間，
       實際開設座數則逐 case 記錄在 raw_results 的 opened_ccps 欄位。
    """
    return 2 ** int(ccp_count)


@contextmanager
def temporary_scale(scale: str, ccp_count: int):
    """暫時把該規模的候選 CCP 數改成 ccp_count（離開後還原）。"""
    profile = cfg.SCALE_PROFILES[scale]
    original_n_ccp = profile["n_ccp"]
    original_mode = cfg.CCP_UPPER_BOUND_SCALING
    original_scenarios = cfg.SCENARIOS
    original_periods = cfg.TIME_PERIODS
    try:
        profile["n_ccp"] = int(ccp_count)
        cfg.CCP_UPPER_BOUND_SCALING = CCP_SCALING_MODE
        cfg.SCENARIOS = BASE_SCENARIOS
        cfg.TIME_PERIODS = BASE_TIME_PERIODS
        yield
    finally:
        profile["n_ccp"] = original_n_ccp
        cfg.CCP_UPPER_BOUND_SCALING = original_mode
        cfg.SCENARIOS = original_scenarios
        cfg.TIME_PERIODS = original_periods


@contextmanager
def temporary_config(case: dict[str, Any], scale: str, ccp_count: int):
    """套用單一 case 的所有設定（含 |J| 覆寫），離開後完整還原。"""
    is_extensive = case.get("engine") == "extensive"
    values = {
        "SCENARIOS": BASE_SCENARIOS,
        "TIME_PERIODS": BASE_TIME_PERIODS,
        "SAMPLE_RATIO": BASE_SAMPLE_RATIO,
        "SP_SAMPLE_RATIO": BASE_SAMPLE_RATIO,
        "DEMAND_MULTIPLIER": BASE_DEMAND_MULTIPLIER,
        "ROAD_CAPACITY_MULTIPLIER": BASE_ROAD_CAPACITY_MULTIPLIER,
        "HOSPITAL_CAPACITY_MULTIPLIER": BASE_HOSPITAL_CAPACITY_MULTIPLIER,
        "SP_TIME_LIMIT": TIME_LIMIT,
        "SP_MIP_GAP": MIP_GAP,
        "EXPERIMENT_SCALE": scale,
        "CCP_UPPER_BOUND_SCALING": CCP_SCALING_MODE,
        "BENDERS_MULTI_CUT": True,
        "BENDERS_EV_WARM_START": case["ev"],
        "BENDERS_ROOT_SEED_ITERS": case["seed"],
        "BENDERS_ROOT_CUT_ROUNDS": case["rounds"],
        "BENDERS_USE_USER_CUTS": case["user"],
        "BENDERS_PARETO_ENABLED": case["pareto"],
        "BENDERS_MIPFOCUS": 0 if is_extensive else DEFAULT_MIPFOCUS,
        "BENDERS_HEURISTICS": 0.0 if is_extensive else DEFAULT_HEURISTICS,
        "BENDERS_NUMERIC_FOCUS": 0 if is_extensive else DEFAULT_NUMERIC_FOCUS,
        "BENDERS_X_BRANCH_PRIORITY_ENABLED": (
            False if is_extensive else DEFAULT_BRANCH_PRIORITY_ENABLED
        ),
        "BENDERS_X_BRANCH_PRIORITY": 0 if is_extensive else DEFAULT_BRANCH_PRIORITY,
    }
    if hasattr(cfg, "CCP_SAMPLE_SIZE"):
        values["CCP_SAMPLE_SIZE"] = None
    original = {key: getattr(cfg, key) for key in values}
    profile = cfg.SCALE_PROFILES[scale]
    original_n_ccp = profile["n_ccp"]
    try:
        profile["n_ccp"] = int(ccp_count)
        for key, value in values.items():
            setattr(cfg, key, value)
        yield
    finally:
        profile["n_ccp"] = original_n_ccp
        for key, value in original.items():
            setattr(cfg, key, value)


@contextmanager
def scaled_instance_generation():
    """portal 內部呼叫 generate_data(sample_ratio=...)；改導向 scale 路徑並快取。"""
    original = cfg.generate_data
    cache: dict[Any, Any] = {}

    def _scaled(*args, **kwargs):
        kwargs.pop("sample_ratio", None)
        kwargs.pop("ccp_sample_size", None)
        scale = cfg.EXPERIMENT_SCALE
        key = (scale, cfg.SCALE_PROFILES[scale]["n_ccp"], cfg.SCENARIOS)
        if key not in cache:
            cache[key] = original(scale=scale)
        return cache[key]

    cfg.generate_data = _scaled
    try:
        yield
    finally:
        cfg.generate_data = original


_PARAM_TABLE_CACHE: list[dict[str, Any]] | None = None


def parameter_table() -> list[dict[str, Any]]:
    """各 (scale, |J|) 的縮放參數與設計不變量檢查。

    產生一次 instance 在 large/S=50 要約 13 秒，12 個組合共約 2 分鐘，
    故快取結果（同一次執行內參數不會變）。
    """
    global _PARAM_TABLE_CACHE
    if _PARAM_TABLE_CACHE is not None:
        return _PARAM_TABLE_CACHE
    rows = []
    for scale in SCALES:
        for ccp_count in CCP_COUNTS:
            with temporary_scale(scale, ccp_count):
                instance = cfg.generate_data(scale=scale)
                cfg.validate_instance(instance)
                p = instance["deterministic_parameters"]
                sets = instance["sets"]
                one = lambda k: float(next(iter(p[k].values())))  # noqa: E731
                staff_pool = float(p["total_available_staff"])
                amb_pool = float(p["total_available_ccp_ambulances"])
                ccp_staff = one("ccp_staff_upper_bound")
                ccp_amb = one("ccp_ambulance_upper_bound")
                rows.append({
                    "scale": scale,
                    "I": len(sets["I"]), "J": len(sets["J"]), "H": len(sets["H"]),
                    "total_staff_pool": staff_pool,
                    "total_amb_pool": amb_pool,
                    "ccp_staff_ub": ccp_staff,
                    "ccp_amb_ub": ccp_amb,
                    "ccp_supply_ub": one("ccp_supply_upper_bound"),
                    "hospital_supply_ub": one("hospital_supply_upper_bound"),
                    "hospital_fleet": one("hospital_ambulance_fleet"),
                    "first_stage_space": first_stage_space(ccp_count),
                    # 設計不變量
                    "inv_pool_binding": staff_pool < ccp_staff * ccp_count,
                    "inv_amb_deployable": ccp_amb * ccp_count >= amb_pool,
                })
    _PARAM_TABLE_CACHE = rows
    return rows


def preflight_parameters() -> list[dict[str, Any]]:
    """跑任何求解前先檢查所有 (scale, |J|) 組合的參數合理性。"""
    rows = parameter_table()
    print("\n--- 各 |J| 的參數對照（demand_only：per-CCP 上限不隨 |J| 改變）---")
    print("    註：模型沒有「必須開幾座」的限制式，求解器可自由開 0~|J| 座。")
    header = (f"{'scale':8}{'I':>5}{'J':>4}{'H':>4}{'醫護池':>8}{'救護車池':>9}"
              f"{'單CCP醫護':>10}{'單CCP車':>9}{'一階搜尋空間 2^J':>20}")
    print(header)
    print("-" * 88)
    for r in rows:
        print(f"{r['scale']:8}{r['I']:>5}{r['J']:>4}{r['H']:>4}"
              f"{r['total_staff_pool']:>8.0f}{r['total_amb_pool']:>9.0f}"
              f"{r['ccp_staff_ub']:>10.0f}{r['ccp_amb_ub']:>9.0f}"
              f"{r['first_stage_space']:>20.3e}")

    bad = [r for r in rows if not r["inv_pool_binding"] or not r["inv_amb_deployable"]]
    if bad:
        for r in bad:
            print(f"  [X] {r['scale']} J={r['J']}: "
                  f"全域池<Σper-CCP={r['inv_pool_binding']}, "
                  f"救護車可部署={r['inv_amb_deployable']}")
        raise RuntimeError("參數不變量檢查失敗，請先修正 config 再跑實驗")
    print("  [OK] 所有組合都通過：全域池 < Σ per-CCP 上限、且救護車池可完全部署")
    return rows


# =============================================================================
# 執行單一 case
# =============================================================================

def _safe_case_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._") or "case"


def hard_timeout_for(case: dict[str, Any]) -> float:
    buffer_s = (HARD_TIMEOUT_BUFFER_SEC_EXTENSIVE
                if case.get("engine") == "extensive" else HARD_TIMEOUT_BUFFER_SEC)
    return TIME_LIMIT + buffer_s


def make_test_id(scale: str, ccp_count: int, case_name: str) -> str:
    return f"{scale}_J{ccp_count}_{case_name}".replace("+", "_plus_")


def _case_paths(scale: str, ccp_count: int, case: dict[str, Any],
                run_idx: int, log_run_dir: Path):
    test_id = make_test_id(scale, ccp_count, case["name"])
    stem = f"{run_idx:02d}_{_safe_case_name(test_id)}"
    return (
        test_id,
        log_run_dir / f"{stem}.log",
        log_run_dir / f"{stem}_solver.log",
        log_run_dir / f".{stem}_launcher.tmp.log",
        log_run_dir / f".{stem}_row.tmp.json",
    )


def snapshot_logs() -> set[Path]:
    return {p.resolve() for p in LOG_DIR.glob("*.log")} if LOG_DIR.exists() else set()


def move_newest_log(before: set[Path], destination: Path) -> Path | None:
    created = [p for p in LOG_DIR.glob("*.log") if p.resolve() not in before]
    if not created:
        return None
    source = max(created, key=lambda p: p.stat().st_mtime)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        raise FileExistsError(f"不覆寫既有的 solver log：{destination}")
    shutil.move(str(source), str(destination))
    return destination


def first_stage_string(fs: dict[str, Any] | None) -> str:
    if not fs:
        return "NA"
    supply: dict[str, float] = {}
    for (h, j), value in fs["Y"].items():
        supply[j] = supply.get(j, 0.0) + float(value)
    lines = []
    for j in sorted(fs["X"]):
        if float(fs["X"][j]) > 0.5:
            lines.append(
                f"CCP {j:4s} -> X: 1, Staff(V): {float(fs['V'][j]):.0f}, "
                f"Amb(U): {float(fs['U'][j]):.0f}, MedicalSupply(Y): {supply.get(j, 0.0):.2f}"
            )
    return "\n".join(lines) if lines else "none opened"


def _dispose_model(model: Any) -> None:
    if model is None:
        return
    try:
        model.dispose()
    except Exception:  # noqa: BLE001
        pass


def run_one_case(portal: Any, case: dict[str, Any], scale: str, ccp_count: int,
                 run_idx: int, total: int, log_run_dir: Path) -> dict[str, Any]:
    _tid, case_log_path, solver_log_path, _l, _r = _case_paths(
        scale, ccp_count, case, run_idx, log_run_dir
    )
    case_log_path.parent.mkdir(parents=True, exist_ok=True)
    with logging_utils.tee_output(case_log_path):
        return _run_one_case_logged(
            portal, case, scale, ccp_count, run_idx, total,
            case_log_path, solver_log_path,
        )


def _run_one_case_logged(portal: Any, case: dict[str, Any], scale: str,
                         ccp_count: int, run_idx: int, total: int,
                         case_log_path: Path, solver_log_path: Path) -> dict[str, Any]:
    engine = case.get("engine", "bbc")
    counts = scale_counts(scale, ccp_count)
    test_id = make_test_id(scale, ccp_count, case["name"])
    row = blank_row()
    row.update({
        "scale": scale, "test_id": test_id, "ccp_count": ccp_count,
        "config": case["name"], "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "log_path": str(case_log_path), "status": "RUNNING", "note": "",
    })
    print(f"\n[{run_idx}/{total}] scale={scale} |J|={ccp_count} config={case['name']}")
    before = snapshot_logs()
    wall_start = time.time()
    model = summary = None
    try:
        try:
            with temporary_config(case, scale, ccp_count):
                model, summary = portal.run_mcvar_model(
                    scenario_size=BASE_SCENARIOS, sample_ratio=BASE_SAMPLE_RATIO,
                    time_limit=TIME_LIMIT, mip_gap=MIP_GAP,
                    alpha=RISK_ALPHA, lam=RISK_LAMBDA, compute_kpis=COMPUTE_KPIS,
                )
        finally:
            row["wall_s"] = f"{time.time() - wall_start:.2f}"
            try:
                moved = move_newest_log(before, solver_log_path)
                row["solver_log_path"] = str(moved) if moved else "NA"
            except Exception as log_exc:  # noqa: BLE001
                row["solver_log_path"] = "NA"
                row["note"] = f"solver log move failed: {type(log_exc).__name__}: {log_exc}"
    except Exception as exc:  # noqa: BLE001
        _dispose_model(model)
        row["status"] = "FAIL"
        prefix = f"{row['note']}; " if row["note"] else ""
        row["note"] = f"{prefix}{type(exc).__name__}: {exc}"
        print(f"  -> FAIL: {row['note']}")
        if STOP_ON_ERROR:
            raise
        return row

    if model is None or summary is None:
        _dispose_model(model)
        row["status"] = "FAIL"
        row["note"] = "portal 沒有回傳模型或摘要"
        return row

    st = summary.get("bbc_stats", {}) or {}
    fs = summary.get("first_stage")
    opened = [j for j, v in ((fs or {}).get("X") or {}).items() if float(v) > 0.5]
    objective = summary.get("objective")
    if objective is None:
        _dispose_model(model)
        row["status"] = "FAIL"
        row["note"] = "時限內未找到任何可行解（objective 為 None）"
        print(f"  -> FAIL: {row['note']}")
        return row

    objective = float(objective)
    best_lb = summary.get("best_lb")
    gap = summary.get("gap_pct")
    with temporary_scale(scale, ccp_count):
        params = cfg.generate_data(scale=scale)["deterministic_parameters"]
    row.update({
        "first_stage_space": first_stage_space(ccp_count),
        "obj_value": f"{objective:.6f}",
        "best_ub": f"{objective:.6f}",
        "best_lb": "NA" if best_lb is None else f"{float(best_lb):.6f}",
        "gap_pct": "NA" if gap is None else f"{float(gap):.6f}",
        "cpu_s": f"{float(st.get('runtime', float('nan'))):.2f}",
        "nodes": f"{float(getattr(model, 'NodeCount', float('nan'))):.0f}",
        "iterations": f"{float(getattr(model, 'IterCount', float('nan'))):.0f}",
        "num_vars": getattr(model, "NumVars", "NA"),
        "num_constrs": getattr(model, "NumConstrs", "NA"),
        "opened_ccps": len(opened),
        "first_stage_decision": first_stage_string(fs),
        "first_stage_cost": summary.get("first_stage_cost", "NA"),
        "total_cuts": st.get("cuts_added", "NA"),
        "seed_cuts": st.get("seed_cuts_added", "NA"),
        "lazy_cuts": st.get("lazy_cuts_added", "NA"),
        "user_cuts": st.get("user_cuts_added", "NA"),
        "root_seed_lb": st.get("root_seed_lb", "NA"),
        "root_seed_iters_done": st.get("root_seed_iters_done", "NA"),
        "root_seed_stop_reason": st.get("root_seed_stop_reason", "NA"),
        "root_seed_time_s": st.get("root_seed_time", "NA"),
        "root_cut_rounds_done": st.get("root_cut_rounds_done", "NA"),
        "oracle_solves": st.get("oracle_solves", "NA"),
        "incumbent_evals": st.get("incumbent_evals", "NA"),
        "callback_time_s": st.get("callback_time", "NA"),
        "parallel_oracles": (0 if engine == "extensive"
                             else st.get("parallel_oracles",
                                         getattr(cfg, "BENDERS_PARALLEL_ORACLES", "NA"))),
        "ccp_staff_ub": float(next(iter(params["ccp_staff_upper_bound"].values()))),
        "ccp_ambulance_ub": float(next(iter(params["ccp_ambulance_upper_bound"].values()))),
        "total_staff_pool": float(params["total_available_staff"]),
        "total_amb_pool": float(params["total_available_ccp_ambulances"]),
        "solver_status": st.get("solver_status", "NA"),
        "status": "OK",
    })
    _dispose_model(model)
    return row


def _failed_case_row(case: dict[str, Any], scale: str, ccp_count: int,
                     case_log_path: Path, note: str, wall_s: float) -> dict[str, Any]:
    counts = scale_counts(scale, ccp_count)
    row = blank_row()
    row.update({
        "scale": scale, "test_id": make_test_id(scale, ccp_count, case["name"]),
        "ccp_count": ccp_count, "config": case["name"],
        "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "wall_s": f"{wall_s:.2f}",
        "solver_status": "HARD_TIMEOUT" if "hard timeout" in note.lower() else "ERROR",
        "log_path": str(case_log_path), "status": "FAIL", "note": note,
    })
    return row


def _write_child_result(path: Path, row: dict[str, Any]) -> None:
    writing = path.with_suffix(path.suffix + ".writing")
    try:
        with writing.open("w", encoding="utf-8") as f:
            json.dump(row, f, ensure_ascii=False, indent=2)
        os.replace(writing, path)
    finally:
        writing.unlink(missing_ok=True)


def _single_case_cli(argv: list[str]) -> None:
    if len(argv) != 7:
        raise SystemExit(
            "single-case arguments: SCALE CCP_COUNT CONFIG_INDEX RUN_INDEX TOTAL LOG_DIR RESULT_JSON"
        )
    scale, ccp_raw, config_idx_raw, run_idx_raw, total_raw, log_dir_raw, result_raw = argv
    ensure_solver_dependency()
    case = CONFIGS[int(config_idx_raw)]
    ccp_count = int(ccp_raw)
    log_run_dir = Path(log_dir_raw).resolve()
    result_path = Path(result_raw).resolve()
    portal = load_portal(
        EXTENSIVE_MODEL_PATH if case.get("engine") == "extensive" else MCVAR_MODEL_PATH,
        "ccp_ablation_portal_child",
    )
    original_scale = cfg.EXPERIMENT_SCALE
    try:
        cfg.EXPERIMENT_SCALE = scale
        with scaled_instance_generation(), open(os.devnull, "w", encoding="utf-8") as sink:
            with redirect_stdout(sink), redirect_stderr(sink):
                row = run_one_case(
                    portal, case, scale, ccp_count,
                    int(run_idx_raw), int(total_raw), log_run_dir,
                )
    finally:
        cfg.EXPERIMENT_SCALE = original_scale
    _write_child_result(result_path, row)


def run_one_case_subprocess(case: dict[str, Any], scale: str, ccp_count: int,
                            run_idx: int, total: int,
                            log_run_dir: Path) -> dict[str, Any]:
    """在隔離子程序中跑單一 case，並強制 wall-clock 上限。"""
    _tid, case_log_path, solver_log_path, launcher_path, result_path = _case_paths(
        scale, ccp_count, case, run_idx, log_run_dir
    )
    before = snapshot_logs()
    start = time.time()
    cmd = [
        sys.executable, str(Path(__file__).resolve()), "--single-case",
        scale, str(ccp_count), str(CONFIGS.index(case)), str(run_idx), str(total),
        str(log_run_dir), str(result_path),
    ]
    timed_out = False
    return_code = None
    hard_timeout = hard_timeout_for(case)
    with launcher_path.open("w", encoding="utf-8", newline="") as launcher:
        try:
            completed = subprocess.run(
                cmd, cwd=ROOT_DIR, stdout=launcher, stderr=subprocess.STDOUT,
                timeout=hard_timeout, check=False,
            )
            return_code = completed.returncode
        except subprocess.TimeoutExpired:
            timed_out = True

    wall_s = time.time() - start
    if not solver_log_path.exists():
        try:
            move_newest_log(before, solver_log_path)
        except Exception:  # noqa: BLE001
            pass

    try:
        if not timed_out and return_code == 0 and result_path.is_file():
            with result_path.open("r", encoding="utf-8") as f:
                return json.load(f)

        if timed_out:
            note = f"Hard timeout: 子程序超過 {hard_timeout:.0f} 秒"
        else:
            note = f"子程序結束碼 {return_code}"
            if return_code in (137, -9):
                note += "（很可能是記憶體不足被系統終止）"

        launcher_tail = ""
        if launcher_path.is_file():
            launcher_tail = launcher_path.read_text(encoding="utf-8", errors="replace")[-8000:]
        if launcher_tail.strip():
            note = f"{note}; launcher output: {launcher_tail.strip()}"
        row = _failed_case_row(case, scale, ccp_count, case_log_path, note, wall_s)
        if not case_log_path.exists():
            case_log_path.write_text(note + "\n", encoding="utf-8")
        else:
            with case_log_path.open("a", encoding="utf-8") as f:
                f.write("\n" + note + "\n")
        row["solver_log_path"] = str(solver_log_path) if solver_log_path.exists() else "NA"
        return row
    finally:
        result_path.unlink(missing_ok=True)
        launcher_path.unlink(missing_ok=True)


# =============================================================================
# 輸出
# =============================================================================

def write_results(path: Path, rows: list[dict[str, Any]]) -> None:
    """原子性寫出 raw CSV；被鎖住時重試，仍失敗就另存 rescue 檔。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}_", suffix=".csv", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        with temp_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
            writer.writeheader()
            writer.writerows(rows)
        for attempt in range(EXCEL_REPLACE_RETRIES):
            try:
                os.replace(temp_path, path)
                return
            except PermissionError:
                if attempt == 0:
                    print(f"[csv] 無法寫入 {path.name}（可能正被 Excel 開啟）；重試中 …")
                time.sleep(EXCEL_REPLACE_RETRY_SLEEP_SEC)
        rescue = path.with_name(f"{path.stem}.rescue_{int(time.time())}.csv")
        shutil.copy2(temp_path, rescue)
        print(f"[csv][WARN] {path.name} 被鎖住，已另存 {rescue.name}")
    finally:
        temp_path.unlink(missing_ok=True)


def _load_prior_rows(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"RESUME_FROM_CSV 找不到檔案: {path}")
    out: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            out.append({k: r.get(k, "NA") for k in FIELDNAMES})
    return out


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return max(0, sum(1 for row in csv.reader(f) if any(row)) - 1)


def excel_value(row: dict[str, Any], key: str) -> Any:
    value = row.get(key, "NA")
    if value in (None, "", "NA"):
        return "NA"
    if key in TEXT_KEYS:
        return str(value)
    try:
        number = float(value)
        return int(number) if key in INTEGER_KEYS and number.is_integer() else number
    except (TypeError, ValueError):
        return value


def _style_header(cell, fill, font, Alignment) -> None:
    cell.fill, cell.font = fill, font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_raw_sheet(wb, rows, fill, font, Alignment) -> None:
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("raw_results")
    for col, name in enumerate(FIELDNAMES, 1):
        _style_header(ws.cell(1, col, name), fill, font, Alignment)
    for r, row in enumerate(rows, 2):
        for col, key in enumerate(FIELDNAMES, 1):
            cell = ws.cell(r, col, excel_value(row, key))
            if key == "first_stage_decision":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    for col in range(1, len(FIELDNAMES) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _write_settings_sheet(wb, run_id, log_run_dir, fill, font, Alignment) -> None:
    ws = wb.create_sheet("run_settings")
    rows = [
        ("run_id", run_id or "NA"),
        ("log_directory", str(log_run_dir) if log_run_dir else "NA"),
        ("實驗目的", "驗證 plan/17 假設：候選 CCP 數量是求解難度的主因"),
        ("scales", ", ".join(SCALES)),
        ("ccp_counts", ", ".join(str(j) for j in CCP_COUNTS)),
        ("configs", ", ".join(c["name"] for c in CONFIGS)),
        ("model", MODEL_NAME),
        ("scenario_count", BASE_SCENARIOS),
        ("time_periods", BASE_TIME_PERIODS),
        ("time_limit_s_per_case", TIME_LIMIT),
        ("mip_gap (提早結束門檻)", MIP_GAP),
        ("risk_alpha", RISK_ALPHA),
        ("risk_lambda", RISK_LAMBDA),
        ("ccp_upper_bound_scaling", CCP_SCALING_MODE),
        ("ccp_sampling", f"{cfg.SCALE_SAMPLING_MODE}（20 ⊂ 30 ⊂ 40 ⊂ 50）"),
        ("data_disaster_csv", cfg.DISASTER_CSV),
        ("data_ccp_csv", cfg.CCP_CSV),
        ("data_hospital_csv", cfg.HOSPITAL_CSV),
        ("bbc_root_seed_iters", DEFAULT_ROOT_SEED_ITERS),
        ("bbc_root_seed_lb_rel_tol", float(cfg.BENDERS_ROOT_SEED_LB_REL_TOL)),
        ("bbc_root_seed_stall_rounds", int(cfg.BENDERS_ROOT_SEED_STALL_ROUNDS)),
        ("bbc_root_cut_rounds", DEFAULT_ROOT_CUT_ROUNDS),
        ("bbc_parallel_oracles", int(cfg.BENDERS_PARALLEL_ORACLES)),
        ("hard_timeout_bbc_s", TIME_LIMIT + HARD_TIMEOUT_BUFFER_SEC),
        ("hard_timeout_extensive_s", TIME_LIMIT + HARD_TIMEOUT_BUFFER_SEC_EXTENSIVE),
    ]
    ws.append(["Setting", "Value"])
    for cell in ws[1]:
        _style_header(cell, fill, font, Alignment)
    for item in rows:
        ws.append(item)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 56


def _write_parameters_sheet(wb, param_rows, fill, font, Alignment) -> None:
    """各 |J| 的參數對照 —— 證明「只有候選組合數在變」。"""
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("parameters")
    headers = [
        ("Scale", "scale"), ("|I|", "I"), ("|J|", "J"), ("|H|", "H"),
        ("全域醫護池", "total_staff_pool"), ("全域救護車池", "total_amb_pool"),
        ("單CCP醫護上限", "ccp_staff_ub"), ("單CCP救護車上限", "ccp_amb_ub"),
        ("單CCP物資上限", "ccp_supply_ub"), ("單醫院物資上限", "hospital_supply_ub"),
        ("單醫院車隊", "hospital_fleet"),
        ("一階搜尋空間 2^|J|", "first_stage_space"),
        ("池<Σper-CCP", "inv_pool_binding"), ("救護車可完全部署", "inv_amb_deployable"),
    ]
    for col, (title, _k) in enumerate(headers, 1):
        _style_header(ws.cell(1, col, title), fill, font, Alignment)
    for r, row in enumerate(param_rows, 2):
        for col, (_t, key) in enumerate(headers, 1):
            ws.cell(r, col, row[key])
    ws.append([])
    ws.append(["說明：per-CCP 上限與全域資源池皆不隨 |J| 改變（設施容量是物理屬性），"
               "故「必須開設座數」固定，唯一改變的是候選組合數。"])
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_summary_sheet(wb, rows, fill, font, Alignment) -> None:
    """依需求輸出 UB / LB / Time / Gap / Nodes。"""
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws = wb.create_sheet("summary_table")

    sub_columns = [
        ("UB", "best_ub"), ("LB", "best_lb"), ("Time(s)", "cpu_s"),
        ("Gap(%)", "gap_pct"), ("Nodes", "nodes"),
    ]
    n_sub = len(sub_columns)
    left = ["Scale", "I", "J", "H", "2^|J|"]
    n_left = len(left)
    n_cols = n_left + len(CONFIGS) * n_sub

    n_j = len(CCP_COUNTS)
    last = 2 + len(SCALES) * n_j
    blocks = [(3 + k * n_j, 3 + k * n_j + n_j - 1) for k in range(len(SCALES))]

    # 1) 先合併
    for i in range(1, n_left + 1):
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
    for idx in range(len(CONFIGS)):
        start = n_left + 1 + idx * n_sub
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + n_sub - 1)
    for block_start, block_end in blocks:
        if block_end > block_start:
            for col in (1, 2, 4):   # Scale / I / H 跨 |J| 列合併（J 與 2^|J| 每列不同）
                ws.merge_cells(start_row=block_start, start_column=col,
                               end_row=block_end, end_column=col)

    # 2) 寫值
    for i, title in enumerate(left, 1):
        ws.cell(1, i, title)
    for idx, case in enumerate(CONFIGS):
        start = n_left + 1 + idx * n_sub
        ws.cell(1, start, case["name"])
        for offset, (sub_title, _key) in enumerate(sub_columns):
            ws.cell(2, start + offset, sub_title)

    for k, scale in enumerate(SCALES):
        block_start = blocks[k][0]
        for ji, ccp_count in enumerate(CCP_COUNTS):
            r = block_start + ji
            counts = scale_counts(scale, ccp_count)
            if ji == 0:
                ws.cell(block_start, 1, scale)
                ws.cell(block_start, 2, counts["I"])
                ws.cell(block_start, 4, counts["H"])
            ws.cell(r, 3, ccp_count)
            match_any = next(
                (x for x in rows if x.get("scale") == scale
                 and str(x.get("ccp_count")) == str(ccp_count)), None
            )
            ws.cell(r, 5, first_stage_space(ccp_count))
            for idx, case in enumerate(CONFIGS):
                match = next(
                    (x for x in rows if x.get("scale") == scale
                     and str(x.get("ccp_count")) == str(ccp_count)
                     and x.get("config") == case["name"]),
                    None,
                )
                start = n_left + 1 + idx * n_sub
                if match is None:
                    values = tuple("NA" for _ in sub_columns)
                elif match.get("status") != "OK":
                    values = tuple("FAIL" for _ in sub_columns)
                else:
                    values = tuple(excel_value(match, key) for _t, key in sub_columns)
                for offset, value in enumerate(values):
                    cell = ws.cell(r, start + offset, value)
                    if isinstance(value, float):
                        key = sub_columns[offset][1]
                        cell.number_format = "0.0000" if key == "gap_pct" else "#,##0.00"

    # 3) 樣式
    for row in ws.iter_rows(min_row=1, max_row=last, min_col=1, max_col=n_cols):
        for cell in row:
            cell.border = border
            if cell.row <= 2:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 9
    for col in "BCD":
        ws.column_dimensions[col].width = 6
    ws.column_dimensions["E"].width = 16
    for col in range(n_left + 1, n_cols + 1):
        key = sub_columns[(col - n_left - 1) % n_sub][1]
        ws.column_dimensions[get_column_letter(col)].width = 16 if key in ("best_ub", "best_lb") else 11
    ws.freeze_panes = f"{get_column_letter(n_left + 1)}3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18


def export_xlsx(rows, path: Path, run_id: str = "",
                log_run_dir: Path | None = None,
                param_rows: list[dict[str, Any]] | None = None) -> Path:
    ensure_excel_dependency()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="2E74B5")
    font = Font(bold=True, color="FFFFFF")

    _write_raw_sheet(wb, rows, fill, font, Alignment)
    _write_settings_sheet(wb, run_id, log_run_dir, fill, font, Alignment)
    if param_rows:
        _write_parameters_sheet(wb, param_rows, fill, font, Alignment)
    _write_summary_sheet(wb, rows, fill, font, Alignment)

    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}_", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
        last_exc: Exception | None = None
        for attempt in range(EXCEL_REPLACE_RETRIES):
            try:
                os.replace(temp_path, path)
                return path
            except PermissionError as exc:
                last_exc = exc
                if attempt == 0:
                    print(f"[excel] 無法寫入 {path.name}（可能正被 Excel 開啟）；重試中 …")
                time.sleep(EXCEL_REPLACE_RETRY_SLEEP_SEC)
        raise PermissionError(f"無法寫入 {path}，請關閉開啟該檔的 Excel") from last_exc
    finally:
        temp_path.unlink(missing_ok=True)


def export_xlsx_incremental(rows, path, run_id, log_run_dir, param_rows) -> None:
    """途中匯出失敗只警告，不中斷實驗（CSV 才是續跑依據）。"""
    try:
        export_xlsx(rows, path, run_id, log_run_dir, param_rows)
    except Exception as exc:  # noqa: BLE001
        print(f"[excel][WARN] 中途匯出失敗，實驗繼續（結果已存在 CSV）："
              f"{type(exc).__name__}: {exc}")


def expected_test_ids() -> set[str]:
    return {
        make_test_id(scale, j, case["name"])
        for scale in SCALES for j in CCP_COUNTS for case in CONFIGS
    }


# =============================================================================
# main
# =============================================================================

def main() -> None:
    ensure_solver_dependency()
    ensure_excel_dependency()

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = RESULT_DIR / f"{RESULT_PREFIX}_raw_{timestamp}.csv"
    xlsx_path = RESULT_DIR / f"{RESULT_PREFIX}_{timestamp}.xlsx"
    log_run_dir = LOG_SUBDIR / timestamp
    log_run_dir.mkdir(parents=True, exist_ok=False)

    print("=" * 78)
    print("候選 CCP 數量 ABLATION：Extensive vs B&BC(全開) × |J| = " +
          "/".join(str(j) for j in CCP_COUNTS))
    print("=" * 78)
    print(f"資料 : {cfg.DISASTER_CSV} / {cfg.CCP_CSV} / {cfg.HOSPITAL_CSV}")
    print(f"模型 : {MODEL_NAME}   S={BASE_SCENARIOS} T={BASE_TIME_PERIODS}")
    print(f"配置 : {[c['name'] for c in CONFIGS]}")
    print(f"時限 : {TIME_LIMIT:.0f}s/case，gap 達 {MIP_GAP:.0%} 即提早結束")
    print(f"縮放 : CCP_UPPER_BOUND_SCALING = {CCP_SCALING_MODE}")
    print(f"抽樣 : CCP 巢狀（{CCP_COUNTS[-1]} ⊂ … ⊂ {CCP_COUNTS[0]}）")

    # 跑任何求解之前，先驗證每個 (scale, |J|) 的參數合理
    param_rows = preflight_parameters()

    n_cases = len(SCALES) * len(CCP_COUNTS) * len(CONFIGS)
    n_ext = len(SCALES) * len(CCP_COUNTS) * sum(1 for c in CONFIGS if c.get("engine") == "extensive")
    worst_h = ((n_cases - n_ext) * hard_timeout_for({})
               + n_ext * hard_timeout_for({"engine": "extensive"})) / 3600.0
    print(f"\ncases = {len(SCALES)} scale × {len(CCP_COUNTS)} |J| × "
          f"{len(CONFIGS)} config = {n_cases}")
    print(f"最壞情況總時間 ≈ {worst_h:.0f} 小時（gap 提早收斂會少很多）")
    print(f"CSV   : {csv_path}\nExcel : {xlsx_path}\nLogs  : {log_run_dir}")

    # preflight：先確認 Excel 寫得出去，避免跑完才發現
    export_xlsx([], xlsx_path, timestamp, log_run_dir, param_rows)

    prior_ok: dict[str, dict[str, Any]] = {}
    if not RESUME_FROM_CSV:
        print(f"\n[FRESH START] RESUME_FROM_CSV 為空 → 從頭跑全部 {n_cases} 個 case")
    else:
        prior_path = Path(RESUME_FROM_CSV)
        if not prior_path.is_absolute():
            prior_path = RESULT_DIR / RESUME_FROM_CSV
        for pr in _load_prior_rows(prior_path):
            if pr.get("status") == "OK" and pr.get("test_id"):
                prior_ok[pr["test_id"]] = pr
        print(f"\n[RESUME] 從 {prior_path} 載入 {len(prior_ok)} 個已完成 case")

    rows: list[dict[str, Any]] = []
    run_idx = 0
    try:
        for scale in SCALES:
            for ccp_count in CCP_COUNTS:
                for case in CONFIGS:
                    run_idx += 1
                    test_id = make_test_id(scale, ccp_count, case["name"])
                    if test_id in prior_ok:
                        print(f"[{run_idx}/{n_cases}] SKIP（沿用上次已完成）{test_id}")
                        rows.append(prior_ok[test_id])
                    else:
                        print(f"\n[{run_idx}/{n_cases}] scale={scale} |J|={ccp_count} "
                              f"config={case['name']}")
                        row = run_one_case_subprocess(
                            case, scale, ccp_count, run_idx, n_cases, log_run_dir,
                        )
                        rows.append(row)
                        print(f"  -> {row['status']} UB={row['obj_value']} "
                              f"LB={row['best_lb']} time={row['cpu_s']}s "
                              f"gap={row['gap_pct']}% nodes={row['nodes']}")
                    write_results(csv_path, rows)
                    export_xlsx_incremental(rows, xlsx_path, timestamp, log_run_dir, param_rows)
    except KeyboardInterrupt:
        write_results(csv_path, rows)
        export_xlsx_incremental(rows, xlsx_path, timestamp, log_run_dir, param_rows)
        print("\n" + "=" * 78)
        print("[INTERRUPTED] 已中斷，完成的結果都已保存。")
        print(f"  已完成 OK: {sum(r['status'] == 'OK' for r in rows)}/{n_cases}")
        print("\n續跑方式：把下面這行貼回本檔參數區，再執行一次。")
        print(f'    RESUME_FROM_CSV = "{csv_path.name}"')
        print("=" * 78)
        return

    export_xlsx(rows, xlsx_path, timestamp, log_run_dir, param_rows)
    n_ok = sum(r["status"] == "OK" for r in rows)
    print("\n" + "=" * 78)
    print(f"Done: {n_ok}/{len(rows)} cases OK")
    print_console_summary(rows)
    print(f"\nCSV   : {csv_path}\nExcel : {xlsx_path}\nLogs  : {log_run_dir}")
    if n_ok < n_cases:
        print("\n尚有 case 未完成，續跑請把下面這行貼回參數區：")
        print(f'    RESUME_FROM_CSV = "{csv_path.name}"')


def print_console_summary(rows: list[dict[str, Any]]) -> None:
    print("\n--- SUMMARY（UB / LB / Time / Gap / Nodes）---")
    head = f"{'scale':7}{'J':>4}{'config':22}{'UB':>15}{'LB':>15}{'Time(s)':>10}{'Gap(%)':>9}{'Nodes':>10}"
    print(head)
    print("-" * len(head))
    for scale in SCALES:
        for j in CCP_COUNTS:
            for case in CONFIGS:
                m = next((x for x in rows if x.get("scale") == scale
                          and str(x.get("ccp_count")) == str(j)
                          and x.get("config") == case["name"]), None)
                if m is None:
                    continue
                if m.get("status") != "OK":
                    print(f"{scale:7}{j:>4}{case['name']:22}{'FAIL':>15}"
                          f"{'':>15}{'':>10}{'':>9}{'':>10}")
                    continue
                fmt = lambda k, d=0: (  # noqa: E731
                    f"{float(m[k]):,.{d}f}" if m.get(k) not in (None, "", "NA") else "NA"
                )
                print(f"{scale:7}{j:>4}{case['name']:22}"
                      f"{fmt('best_ub'):>15}{fmt('best_lb'):>15}"
                      f"{fmt('cpu_s',1):>10}{fmt('gap_pct',2):>9}{fmt('nodes'):>10}")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--single-case":
        _single_case_cli(sys.argv[2:])
    else:
        main()
