"""batch_dro_full_matrix.py — DRO 全矩陣實驗：3 規模 × 3 模糊集 × 8 方法 = 72 格。

模型
----
三種都是 SP + MCVaR + DRO（風險層本來就含 MCVaR，DRO 只是在其上加模糊集）：
    dro_box          方塊模糊集（線性對偶，master 仍是 MILP）
    dro_ellipsoidal  橢球模糊集（含 ‖·‖₂ 二階錐，master 變成 MIQCP，明顯較慢）
    dro_polyhedral   多面體模糊集（∞-norm 線性化，master 仍是 MILP）

方法階梯（累加式，共 8 段）
-------------------------
    1  Extensive                      整體式，無任何加速
    2  BBC                            Branch-and-Benders-Cut，其餘全關
    3  BBC+WS                         ＋EV 暖啟動
    4  BBC+WS+RS                      ＋root seeding
    5  BBC+WS+RS+UC                   ＋root 節點 user cuts
    6  BBC+WS+RS+UC+Pareto            ＋Papadakos core point
    7  BBC+WS+RS+UC+Pareto+LBF        ＋平均情境下界函數
    8  BBC+WS+RS+UC+Pareto+LBF+VI     ＋八條有效不等式

執行
----
    python "run experiment/batch_dro_full_matrix.py"                # 全部 72 格
    python "run experiment/batch_dro_full_matrix.py" --scales small # 只跑 small（24 格）
    python "run experiment/batch_dro_full_matrix.py" --models dro_box
    python "run experiment/batch_dro_full_matrix.py" --methods 1,2,8
    python "run experiment/batch_dro_full_matrix.py" --dry-run      # 不求解，只驗流程與輸出

中斷續跑
-------
    **預設就會續跑。** 每跑完一格立刻寫進度檔，重新執行時已成功的格子自動跳過；
    失敗的格子（GurobiError / MemoryError / Exception）會自動重試 —— 那些多半是
    暫時性原因，不該被永久跳過。要保留失敗結果不重試請加 --keep-failed。
    要從頭重跑請加 --fresh（舊進度檔會先改名備份，不會刪除）。

    同時只允許一個實例執行（鎖檔）。六天的批次最怕的就是開兩個視窗跑，
    兩邊各自從自己的快照重寫整個進度檔，後寫的把前面幾天的結果洗掉。

輸出（都在 experiment result/）
-----------------------------
    DRO_full_matrix_raw.csv     進度檔＋原始資料，每格跑完立刻更新（原子替換）
    DRO_full_matrix.xlsx        每格跑完就重建
        Summary        平表：Scale/Model/Method + UB/LB/Time/Gap(%)/Nodes + 健檢
        Table_small    論文版面：8 列方法 × 3 個模糊集，每個 5 欄
        Table_medium   （同上）
        Table_large    （同上）
        Detail         全部診斷欄位
        Config         本次每一項設定，供重現
    DRO_full_matrix_<timestamp>.xlsx   全部跑完時另存封存副本

欄位定義（重要）
--------------
    Time        整格的實際牆鐘時間（含建模）。Extensive 與 BBC 兩條路徑用同一個
                定義才可比 —— Gurobi 的 Runtime 只算最後一次 optimize，會漏掉
                large 建模的好幾分鐘，讓 Extensive 看起來比實際快。
    solver_secs 純求解時間（Extensive 為 m.Runtime；BBC 為 solve_bbc 的 runtime）。
    Nodes       直接讀 master.NodeCount，即使沒有 incumbent 也照實記錄。
    sanity      自動健檢旗標，空白代表沒發現異常。見 sanity_flags()。
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
import traceback
from datetime import datetime, timedelta
from pathlib import Path

for _st in (sys.stdout, sys.stderr):
    try:
        _st.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

ROOT_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT_DIR / "model core"))
sys.path.insert(0, str(ROOT_DIR / "model portal"))
RESULT_DIR = ROOT_DIR / "experiment result"

import gurobipy as gp
from gurobipy import GRB

import config
import extensive_form_core as model_core
import lshaped_core
import risk_core
# Extensive 那一段直接重用 portal 既有的風險目標式建構，確保與原本的 extensive
# DRO 入口逐字相同 —— 自己重寫一份會有對不上的風險。
import extensive_dro

RESULT_PREFIX = "DRO_full_matrix"
SCALES = ("small", "medium", "large")
MODELS = ("dro_box", "dro_ellipsoidal", "dro_polyhedral")
MODEL_LABEL = {"dro_box": "DRO-box", "dro_ellipsoidal": "DRO-ellipsoidal",
               "dro_polyhedral": "DRO-polyhedral"}
DEFAULT_TIME_LIMIT = 7200.0

_ALL_VI_OFF = {"all": False}
_ALL_VI_ON = {"all": True}
_RS = int(getattr(config, "BENDERS_ROOT_SEED_ITERS", 300))
_UC_ROUNDS = int(getattr(config, "BENDERS_ROOT_CUT_ROUNDS", 15))

# 這些狀態代表「這一格真的跑完了」，續跑時才跳過。其餘（GurobiError、
# MemoryError、Exception…）一律視為未完成並自動重試。
SUCCESS_STATUS = {"OPTIMAL", "TIME_LIMIT", "SUBOPTIMAL", "INTERRUPTED",
                  "NODE_LIMIT", "SOLUTION_LIMIT", "ITERATION_LIMIT"}
STATUS_NAME = {1: "LOADED", 2: "OPTIMAL", 3: "INFEASIBLE", 4: "INF_OR_UNBD",
               5: "UNBOUNDED", 6: "CUTOFF", 7: "ITERATION_LIMIT", 8: "NODE_LIMIT",
               9: "TIME_LIMIT", 10: "SOLUTION_LIMIT", 11: "INTERRUPTED",
               12: "NUMERIC", 13: "SUBOPTIMAL", 14: "INPROGRESS",
               15: "USER_OBJ_LIMIT", 16: "WORK_LIMIT", 17: "MEM_LIMIT"}

# 累加式階梯。每一段的設定都是絕對值（不是與上一段的差異），所以
# --methods 1,2,8 這種跳著跑也不會出錯。
LADDER = [
    (1, "Extensive", dict(kind="ext")),
    (2, "BBC", dict(kind="bbc", ws=False, rs=0, uc=False, pareto=False, lbf=False, vi=False)),
    (3, "BBC+WS", dict(kind="bbc", ws=True, rs=0, uc=False, pareto=False, lbf=False, vi=False)),
    (4, "BBC+WS+RS", dict(kind="bbc", ws=True, rs=_RS, uc=False, pareto=False, lbf=False, vi=False)),
    (5, "BBC+WS+RS+UC", dict(kind="bbc", ws=True, rs=_RS, uc=True, pareto=False, lbf=False, vi=False)),
    (6, "BBC+WS+RS+UC+Pareto", dict(kind="bbc", ws=True, rs=_RS, uc=True, pareto=True, lbf=False, vi=False)),
    (7, "BBC+WS+RS+UC+Pareto+LBF", dict(kind="bbc", ws=True, rs=_RS, uc=True, pareto=True, lbf=True, vi=False)),
    (8, "BBC+WS+RS+UC+Pareto+LBF+VI", dict(kind="bbc", ws=True, rs=_RS, uc=True, pareto=True, lbf=True, vi=True)),
]
METHOD_BY_ID = {mid: (name, cfg) for mid, name, cfg in LADDER}


def sname(st) -> str:
    try:
        return STATUS_NAME.get(int(st), f"STATUS_{st}")
    except (TypeError, ValueError):
        return str(st)


# ===================================================================== #
# instance 快取（一次只留一個規模，避免三份大 instance 同時佔記憶體）        #
# ===================================================================== #
_INSTANCE_CACHE: dict[str, dict] = {}


def get_instance(scale: str) -> dict:
    if scale not in _INSTANCE_CACHE:
        for k in list(_INSTANCE_CACHE):        # 換規模就把上一個放掉
            if k != scale:
                _INSTANCE_CACHE.pop(k, None)
        t0 = time.time()
        _INSTANCE_CACHE[scale] = config.generate_data(scale=scale)
        print(f"   （生成 {scale} instance：{time.time() - t0:.1f}s）", flush=True)
    return _INSTANCE_CACHE[scale]


def norm_probs_of(instance: dict, S_sel: list[str]) -> dict[str, float]:
    sd = instance["scenario_data"]
    raw = {s: sd["probability"][s] for s in S_sel}
    tot = sum(raw.values())
    return {s: p / tot for s, p in raw.items()}


# ===================================================================== #
# 兩種求解路徑                                                           #
# ===================================================================== #
def solve_extensive(instance, S_sel, risk_cfg, time_limit, mip_gap) -> dict:
    """整體式：建完整模型 + 風險目標式，直接丟給 Gurobi。無任何加速。"""
    sets = instance["sets"]
    p0 = norm_probs_of(instance, S_sel)
    sub_sd = {k: {s: instance["scenario_data"][k][s] for s in S_sel} for k in
              ("demand", "road_availability_ij", "road_availability_jh",
               "hospital_receiving_capacity")}
    m = None
    t_build = time.time()
    try:
        m, v = model_core.build_gurobi_model(
            sets["I"], sets["J"], sets["H"], sets["L"], sets["L_transfer"],
            sets["T"], S_sel, instance["deterministic_parameters"], sub_sd, p0,
            instance["road_capacity"]["cap_ij"], instance["road_capacity"]["cap_jh"],
            instance["transport_cost"]["cost_ij"], instance["transport_cost"]["cost_jh"],
            model_name="Extensive_DRO", time_limit=time_limit, mip_gap=mip_gap,
            vi_cfg=_ALL_VI_OFF,          # Extensive 是基準線，一條 VI 都不加
        )
        extensive_dro._apply_risk_objective(m, v, S_sel, p0, risk_cfg)
        build_secs = time.time() - t_build
        m.setParam("OutputFlag", 1)
        m.setParam("DisplayInterval", int(getattr(config, "BENDERS_DISPLAY_INTERVAL", 30)))
        m.optimize()
        has = m.SolCount > 0
        out = dict(
            UB=float(m.ObjVal) if has else None,
            Gap_pct=float(m.MIPGap) * 100.0 if has else None,
            solver_secs=float(m.Runtime),
            build_secs=build_secs,
            status=sname(m.status),
            model_vars=int(m.NumVars), model_constrs=int(m.NumConstrs),
            model_qconstrs=int(m.NumQConstrs),
        )
        # 沒有界的時候讀 ObjBound 會拋，逐一保護
        try:
            out["LB"] = float(m.ObjBound)
        except (gp.GurobiError, AttributeError):
            out["LB"] = None
        try:
            out["Nodes"] = float(m.NodeCount)
        except (gp.GurobiError, AttributeError):
            out["Nodes"] = None
        return out
    finally:
        if m is not None:
            m.dispose()


def solve_ladder_bbc(instance, S_sel, risk_cfg, cfg, time_limit, mip_gap) -> dict:
    """階梯第 2~8 段：BBC，依 cfg 逐項開啟元件。"""
    # 沒有 EV 資料就不能做暖啟動；直接關掉並記錄，不要讓整格因 KeyError 陣亡
    ws = cfg["ws"] and ("deterministic_data" in instance)
    res = master = None
    try:
        res = lshaped_core.solve_bbc(
            instance, S_sel,
            time_limit=time_limit, mip_gap=mip_gap, risk_cfg=risk_cfg,
            multi_cut=True,                                   # 風險模型必須 multi-cut
            ev_warm_start=ws,                                 # WS
            root_seed_iters=cfg["rs"],                        # RS
            use_user_cuts=cfg["uc"],                          # UC
            root_cut_rounds=(_UC_ROUNDS if cfg["uc"] else 0),
            pareto_enabled=cfg["pareto"],                     # Pareto
            # rung 8 明確傳 {"all": True}，不靠 config 決定 —— 否則有人改了
            # config.VI_ENABLED，第 8 段就會在標籤寫著 +VI 的情況下沒有 VI。
            lbf_enabled=cfg["lbf"],                           # LBF
            vi_cfg=(_ALL_VI_ON if cfg["vi"] else _ALL_VI_OFF),
            verbose=True,
        )
        fs = res.get("first_stage") or {}
        master = res.pop("master", None)
        res.pop("vars", None)
        # solve_bbc 在沒有 incumbent 時把 nodes 記成 0，會讓最難的那幾格
        # 看起來像「根節點就解完」。直接讀 master 的真實節點數。
        nodes = res.get("nodes")
        nq = None
        if master is not None:
            try:
                nodes = float(master.NodeCount)
                nq = int(master.NumQConstrs)
            except (gp.GurobiError, AttributeError):
                pass
        return dict(
            UB=res.get("best_ub"), LB=res.get("best_lb"),
            solver_secs=res.get("runtime"),
            Gap_pct=res.get("gap_pct"), Nodes=nodes,
            status=res.get("status"),
            ws_effective=ws, model_qconstrs=nq,
            iterations=res.get("iterations"), cuts_added=res.get("cuts_added"),
            seed_cuts=res.get("seed_cuts_added"), user_cuts=res.get("user_cuts_added"),
            lazy_cuts=res.get("lazy_cuts_added"),
            root_seed_iters_done=res.get("root_seed_iters_done"),
            root_seed_lb=res.get("root_seed_lb"),
            root_seed_stop=res.get("root_seed_stop_reason"),
            root_seed_time=res.get("root_seed_time"),
            root_cut_rounds_done=res.get("root_cut_rounds_done"),
            oracle_solves=res.get("oracle_solves"), cache_hits=res.get("cache_hits"),
            callback_time=res.get("callback_time"),
            vi8_rows=res.get("vi8_rows"),
            vi_flags=json.dumps(res.get("vi_flags") or {}, ensure_ascii=False),
            vi_theta_lb_n=len(res.get("vi_theta_lb") or {}),
            sum_X=sum(fs.get("X", {}).values()) if fs else None,
            sum_V=sum(fs.get("V", {}).values()) if fs else None,
            sum_U=sum(fs.get("U", {}).values()) if fs else None,
            sum_Y=sum(fs.get("Y", {}).values()) if fs else None,
        )
    finally:
        # 63 格 BBC，每個 master 都帶著 VI 區塊、DRO 對偶層與數百輪 seed cut。
        # 不釋放的話第五天很可能因記憶體不足倒掉。
        if master is not None:
            try:
                master.dispose()
            except Exception:                                  # noqa: BLE001
                pass


# ===================================================================== #
# 自動健檢：抓「跑完了但這個數字沒有意義」的情況                            #
# ===================================================================== #
def sanity_flags(row: dict, time_limit: float) -> str:
    f = []
    if row.get("status") in SUCCESS_STATUS:
        if row.get("LB") is None:
            f.append("無下界")
        if row.get("UB") is None:
            f.append("無可行解")
        nd = row.get("Nodes")
        if isinstance(nd, (int, float)) and nd == 0 and row.get("method_id") != 1:
            f.append("節點數為0")
        rst = row.get("root_seed_time")
        if isinstance(rst, (int, float)) and rst > 0.15 * time_limit:
            f.append(f"root_seeding佔{rst / time_limit * 100:.0f}%時限")
        if row.get("method_id") == 8:
            vf = row.get("vi_flags")
            try:
                d = json.loads(vf) if isinstance(vf, str) else (vf or {})
                if d and not all(d.values()):
                    f.append("第8段的VI未全開")
            except (ValueError, TypeError):
                pass
        if row.get("ws_effective") is False and row.get("method_id", 0) >= 3:
            f.append("暖啟動未生效")
    return "; ".join(f)


# ===================================================================== #
# 單一格                                                                 #
# ===================================================================== #
def run_cell(scale, model_type, method_id, time_limit, mip_gap, n_scen, dry_run) -> dict:
    name, cfg = METHOD_BY_ID[method_id]
    row = dict(scale=scale, model=model_type, model_label=MODEL_LABEL[model_type],
               method_id=method_id, method=name,
               time_limit=time_limit, mip_gap=mip_gap,
               started_at=datetime.now().isoformat(timespec="seconds"))
    if dry_run:
        row.update(UB=1.0, LB=1.0, Time=0.0, Gap_pct=0.0, Nodes=0.0,
                   status="DRY_RUN", n_I=0, n_J=0, n_H=0, n_T=0, n_S=0, sanity="",
                   finished_at=datetime.now().isoformat(timespec="seconds"))
        return row

    t0 = time.time()
    try:
        # 連 instance 生成與 risk_cfg 都包進來 —— 第 49 格才第一次生成 large，
        # 那時已經跑了四天，不能因為一個 MemoryError 就讓整批陣亡。
        instance = get_instance(scale)
        sets = instance["sets"]
        S_all = sets["S"]
        S_sel = S_all if n_scen is None else S_all[:n_scen]
        row.update(n_I=len(sets["I"]), n_J=len(sets["J"]), n_H=len(sets["H"]),
                   n_T=len(sets["T"]), n_S=len(S_sel))
        risk_cfg = risk_core.make_risk_cfg(model_type)
        row.update(risk_alpha=risk_cfg["alpha"], risk_lambda=risk_cfg["lambda"],
                   ambiguity_scope=risk_core.risk_scope(risk_cfg))
        risk_core.validate_risk_cfg_for_probs(risk_cfg, norm_probs_of(instance, S_sel))

        if cfg["kind"] == "ext":
            out = solve_extensive(instance, S_sel, risk_cfg, time_limit, mip_gap)
        else:
            out = solve_ladder_bbc(instance, S_sel, risk_cfg, cfg, time_limit, mip_gap)
        row.update(out)
    except gp.GurobiError as exc:
        row.update(status="GurobiError", error=str(exc))
        print(f"\n   [!] Gurobi 錯誤：{exc}", flush=True)
    except MemoryError:
        row.update(status="MemoryError", error="記憶體不足")
        print("\n   [!] 記憶體不足", flush=True)
    except Exception as exc:                                       # noqa: BLE001
        row.update(status="Exception", error=f"{type(exc).__name__}: {exc}")
        traceback.print_exc()
    # Time 一律用整格牆鐘時間：Extensive 與 BBC 才可比（Gurobi 的 Runtime
    # 不含建模，large 的建模動輒好幾分鐘）。
    row["Time"] = time.time() - t0
    row["wall_secs"] = row["Time"]
    row["finished_at"] = datetime.now().isoformat(timespec="seconds")
    row["sanity"] = sanity_flags(row, time_limit)
    return row


# ===================================================================== #
# 進度檔                                                                 #
# ===================================================================== #
_FLOAT_COLS = {"UB", "LB", "Time", "Gap_pct", "Nodes", "wall_secs", "solver_secs",
               "build_secs", "time_limit", "mip_gap", "root_seed_lb", "callback_time",
               "root_seed_time", "risk_alpha", "risk_lambda", "ambiguity_scope",
               "sum_X", "sum_V", "sum_U", "sum_Y"}
_INT_COLS = {"method_id", "n_I", "n_J", "n_H", "n_T", "n_S", "iterations",
             "cuts_added", "seed_cuts", "user_cuts", "lazy_cuts",
             "root_seed_iters_done", "root_cut_rounds_done", "oracle_solves",
             "cache_hits", "vi8_rows", "vi_theta_lb_n", "model_vars",
             "model_constrs", "model_qconstrs"}


def cell_key(r):
    """回傳 (scale, model, method_id)；欄位缺漏或壞掉時回 None 而不是拋例外。"""
    try:
        sc, md, mid = r.get("scale"), r.get("model"), r.get("method_id")
        if sc is None or md is None or mid is None:
            return None
        return (str(sc), str(md), int(mid))
    except (TypeError, ValueError):
        return None


def load_progress(path: Path) -> list[dict]:
    if not path.exists():
        return []
    rows, dropped = [], 0
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            for r in csv.DictReader(f):
                for k, val in list(r.items()):
                    if val == "":
                        r[k] = None
                    elif k in _FLOAT_COLS:
                        try:
                            r[k] = float(val)
                        except (TypeError, ValueError):
                            pass
                    elif k in _INT_COLS:
                        try:
                            r[k] = int(float(val))
                        except (TypeError, ValueError):
                            pass
                    elif k == "ws_effective":
                        r[k] = (str(val).strip().lower() == "true")
                if cell_key(r) is None:
                    dropped += 1
                    continue
                rows.append(r)
    except (OSError, csv.Error) as exc:
        print(f" [!] 進度檔讀取失敗（{exc}），本次視為從頭開始。"
              f"舊檔仍在：{path}", flush=True)
        return []
    if dropped:
        print(f" [!] 進度檔中有 {dropped} 列缺少 scale/model/method_id，已略過"
              f"（可能來自別的腳本）。", flush=True)
    return rows


def write_csv(rows: list[dict], path: Path) -> None:
    keys: list[str] = []
    for r in rows:
        for k in r:
            if k not in keys:
                keys.append(k)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    tmp.replace(path)      # 原子替換：寫到一半斷電也不會毀掉舊進度檔


# ===================================================================== #
# Excel                                                                 #
# ===================================================================== #
METRICS = [("UB", "UB"), ("LB", "LB"), ("Time", "Time (s)"),
           ("Gap_pct", "Gap (%)"), ("Nodes", "Nodes")]
NUMFMT = {"UB": "#,##0.00", "LB": "#,##0.00", "Time": "#,##0.0",
          "Gap_pct": "0.000", "Nodes": "#,##0"}
SUMMARY_HEAD = [("scale", "Scale"), ("model_label", "Model"), ("method", "Method"),
                ("n_I", "|I|"), ("n_J", "|J|"), ("n_H", "|H|"), ("n_S", "|S|"),
                ("UB", "UB"), ("LB", "LB"), ("Time", "Time (s)"),
                ("Gap_pct", "Gap (%)"), ("Nodes", "Nodes"),
                ("status", "Status"), ("sanity", "健檢")]


def write_xlsx(rows: list[dict], path: Path, meta: dict, scales, models) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor="2F6DB5")
    sub_fill = PatternFill("solid", fgColor="D9E4F1")
    head_font = Font(bold=True, color="FFFFFF")
    sub_font = Font(bold=True)
    thin = Side(style="thin", color="B7C4D4")
    boxb = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal="center", vertical="center")

    idx = {}
    for r in rows:
        k = cell_key(r)
        if k is not None:
            idx[k] = r
    # 分頁範圍取「本次指定的」∪「進度檔裡已有的」。否則用 --scales large 續跑時，
    # 會把已完成的 small/medium 從 Summary 與表格中整個抹掉。
    have_sc = {k[0] for k in idx}
    have_md = {k[1] for k in idx}
    xl_scales = [s for s in SCALES if s in (set(scales) | have_sc)]
    xl_models = [m for m in MODELS if m in (set(models) | have_md)]

    wb = Workbook()

    # ---------------- Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws.append([lbl for _, lbl in SUMMARY_HEAD])
    for sc in xl_scales:
        for md in xl_models:
            for mid, mname, _ in LADDER:
                r = idx.get((sc, md, mid))
                if r is None:
                    ws.append([sc, MODEL_LABEL[md], mname] + [None] * 9 + ["未執行", ""])
                else:
                    ws.append([r.get(k) for k, _ in SUMMARY_HEAD])
    for c in range(1, len(SUMMARY_HEAD) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = head_fill, head_font, ctr, boxb
    for ri in range(2, ws.max_row + 1):
        for ci, (key, _) in enumerate(SUMMARY_HEAD, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = boxb
            if key in NUMFMT:
                cell.number_format = NUMFMT[key]
    for ci, w in enumerate([9, 17, 30, 6, 6, 6, 6, 18, 18, 11, 10, 12, 13, 28], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "D2"

    # ---------------- Table_<scale>：論文版面 ----------------
    for sc in xl_scales:
        wt = wb.create_sheet(f"Table_{sc}")
        wt.cell(row=1, column=1, value="Method")
        wt.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        for k, md in enumerate(xl_models):
            c0 = 2 + k * len(METRICS)
            wt.cell(row=1, column=c0, value=MODEL_LABEL[md])
            wt.merge_cells(start_row=1, start_column=c0,
                           end_row=1, end_column=c0 + len(METRICS) - 1)
            for q, (_, lbl) in enumerate(METRICS):
                wt.cell(row=2, column=c0 + q, value=lbl)
        for ri, (mid, mname, _) in enumerate(LADDER, start=3):
            wt.cell(row=ri, column=1, value=mname)
            for k, md in enumerate(xl_models):
                c0 = 2 + k * len(METRICS)
                r = idx.get((sc, md, mid))
                for q, (key, _) in enumerate(METRICS):
                    val = r.get(key) if r else None
                    # 沒有 incumbent 時只有 UB 與 Gap 不存在；LB、Time、Nodes
                    # 都是真實算出來的資料，不可以一起抹掉。
                    if r is not None and val is None and key in ("UB", "Gap_pct"):
                        val = (r.get("status") if key == "UB" else None)
                    cell = wt.cell(row=ri, column=c0 + q, value=val)
                    if isinstance(val, (int, float)):
                        cell.number_format = NUMFMT[key]
        ncol = 1 + len(xl_models) * len(METRICS)
        for c in range(1, ncol + 1):
            for rr in (1, 2):
                cell = wt.cell(row=rr, column=c)
                cell.fill = head_fill if rr == 1 else sub_fill
                cell.font = head_font if rr == 1 else sub_font
                cell.alignment, cell.border = ctr, boxb
        for rr in range(3, 3 + len(LADDER)):
            for c in range(1, ncol + 1):
                wt.cell(row=rr, column=c).border = boxb
        wt.column_dimensions["A"].width = 32
        for c in range(2, ncol + 1):
            wt.column_dimensions[get_column_letter(c)].width = 15
        wt.freeze_panes = "B3"

    # ---------------- Detail ----------------
    wd = wb.create_sheet("Detail")
    keys: list[str] = []
    for r in rows:
        for k in r:
            if k not in keys:
                keys.append(k)
    wd.append(keys)
    for r in rows:
        wd.append([r.get(k) for k in keys])
    for c in range(1, len(keys) + 1):
        cell = wd.cell(row=1, column=c)
        cell.fill, cell.font = head_fill, head_font
        wd.column_dimensions[get_column_letter(c)].width = \
            max(12, min(26, len(keys[c - 1]) + 4))
    wd.freeze_panes = "A2"

    # ---------------- Config ----------------
    wc = wb.create_sheet("Config")
    wc.append(["設定", "值"])
    for k, v in meta.items():
        wc.append([k, v if isinstance(v, (int, float, str)) else
                   json.dumps(v, ensure_ascii=False)])
    for c in (1, 2):
        cell = wc.cell(row=1, column=c)
        cell.fill, cell.font = head_fill, head_font
    wc.column_dimensions["A"].width = 40
    wc.column_dimensions["B"].width = 64
    wc.freeze_panes = "A2"

    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    tmp.replace(path)


def collect_meta(args, scales, models, method_ids, mip_gap) -> dict:
    names = ["VI_ENABLED", "VI_1_ROADCAP_IJ", "VI_2_ROADCAP_JH", "VI_3_HOSP_MERGE",
             "VI_4_STAFF_UB", "VI_5_OPEN_USE", "VI_6_THETA_LB", "VI_7_THETA_UB",
             "VI_8_AGG_RELAX", "VI_APPLY_TO_LBF", "VI_THETA_LB_TIME_LIMIT",
             "BENDERS_MULTI_CUT", "BENDERS_ROOT_SEED_ITERS", "BENDERS_ROOT_CUT_ROUNDS",
             "BENDERS_PAPADAKOS_BLEND", "BENDERS_PARALLEL_ORACLES", "BENDERS_MIPFOCUS",
             "BENDERS_HEURISTICS", "BENDERS_NUMERIC_FOCUS",
             "BENDERS_X_BRANCH_PRIORITY_ENABLED", "BENDERS_X_BRANCH_PRIORITY",
             "BENDERS_INCUMBENT_EARLY_TERMINATION", "MASTER_SEED", "SCENARIOS",
             "RISK_ALPHA", "RISK_LAMBDA", "DRO_EPSILON_BOX", "DRO_A_E", "DRO_A_P"]
    meta = {
        "產生時間": datetime.now().isoformat(timespec="seconds"),
        "矩陣": f"{len(scales)} 規模 × {len(models)} 模糊集 × {len(method_ids)} 方法"
                f" = {len(scales) * len(models) * len(method_ids)} 格",
        "規模": ", ".join(scales),
        "模型": ", ".join(MODEL_LABEL[m] for m in models) + "（皆為 SP+MCVaR+DRO）",
        "方法階梯": " → ".join(METHOD_BY_ID[i][0] for i in method_ids),
        "每格時限(秒)": args.time_limit,
        "MIPGap": mip_gap,
        "情境數": args.scenarios if args.scenarios else getattr(config, "SCENARIOS", None),
        "Time 欄定義": "整格牆鐘時間（含建模）；純求解時間見 solver_secs",
        "Nodes 欄定義": "master.NodeCount，無 incumbent 時亦照實記錄",
        "續跑規則": "只有成功狀態才跳過；失敗的格子會自動重試",
        "gurobi": ".".join(str(v) for v in gp.gurobi.version()),
        "python": sys.version.split()[0],
    }
    for k in names:
        meta[k] = getattr(config, k, None)
    return meta


# ===================================================================== #
# main                                                                  #
# ===================================================================== #
def fmt_hms(sec) -> str:
    try:
        return str(timedelta(seconds=int(max(0, float(sec)))))
    except (TypeError, ValueError):
        return "—"


def num(v, spec="", na="—") -> str:
    return format(v, spec) if isinstance(v, (int, float)) else na


class RunLock:
    """同時只允許一個實例執行。兩個實例會各自從自己的快照重寫進度檔，
    後寫的把前面幾天的結果洗掉，而且因為是原子替換連壞檔都不會留下。

    執行中每跑完一格會更新鎖檔的時間戳（心跳）。若鎖檔的心跳已經超過
    stale_after 秒沒更新，視為上次異常終止留下的陳舊鎖，直接接手 ——
    否則一次強制關機就會讓後續所有執行都被永久擋住。
    """

    def __init__(self, path: Path, stale_after: float):
        self.path = path
        self.stale_after = max(1800.0, stale_after)
        self.fd = None

    def _write(self):
        try:
            self.path.write_text(
                f"pid={os.getpid()} "
                f"start={datetime.now():%Y-%m-%d %H:%M:%S}\n", encoding="utf-8")
        except OSError:
            pass

    def heartbeat(self):
        """每跑完一格呼叫一次，證明本實例還活著。"""
        self._write()

    def __enter__(self):
        try:
            self.fd = os.open(self.path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(self.fd); self.fd = None
            self._write()
            return self
        except FileExistsError:
            pass
        try:
            age = time.time() - self.path.stat().st_mtime
            info = self.path.read_text(encoding="utf-8").strip()
        except OSError:
            age, info = 0.0, "(無法讀取)"
        if age > self.stale_after:
            print(f" [!] 發現陳舊鎖檔（{fmt_hms(age)} 未更新），視為上次異常終止，"
                  f"本次接手。內容：{info}", flush=True)
            self._write()
            return self
        print("=" * 100)
        print(" 偵測到已有一個實例正在執行，本次中止。")
        print(f" 鎖檔：{self.path}")
        print(f" 內容：{info}　（{fmt_hms(age)} 前更新）")
        print(" 兩個實例同時跑會互相覆蓋進度檔，已完成的結果會消失，故不允許。")
        print(f" 若確定沒有其他程序在跑，刪除該鎖檔後再執行；或等 "
              f"{fmt_hms(self.stale_after - age)} 後它會被自動視為陳舊。")
        print("=" * 100)
        raise SystemExit(3)

    def __exit__(self, *a):
        try:
            self.path.unlink(missing_ok=True)
        except OSError:
            pass
        return False


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="DRO 全矩陣：3 規模 × 3 模糊集 × 8 方法（預設自動續跑）")
    ap.add_argument("--scales", default=",".join(SCALES))
    ap.add_argument("--models", default=",".join(MODELS))
    ap.add_argument("--methods", default="1,2,3,4,5,6,7,8",
                    help="階梯編號，1=Extensive … 8=全堆疊")
    ap.add_argument("--time-limit", type=float, default=DEFAULT_TIME_LIMIT)
    ap.add_argument("--mip-gap", type=float, default=None)
    ap.add_argument("--scenarios", type=int, default=None)
    ap.add_argument("--tag", default="")
    ap.add_argument("--fresh", action="store_true",
                    help="不續跑，從頭開始（舊進度檔會先備份不會刪除）")
    ap.add_argument("--keep-failed", action="store_true",
                    help="失敗的格子也視為完成、不重試（預設會自動重試）")
    ap.add_argument("--dry-run", action="store_true",
                    help="不求解，只驗證流程、續跑邏輯與 Excel 版面")
    args = ap.parse_args(argv)

    scales = [x.strip() for x in args.scales.split(",") if x.strip()]
    models = [x.strip() for x in args.models.split(",") if x.strip()]
    try:
        method_ids = [int(x) for x in args.methods.replace(" ", "").split(",") if x]
    except ValueError:
        ap.error(f"--methods 只能是逗號分隔的整數，收到 {args.methods!r}")
    if not scales:
        ap.error("--scales 不能是空的")
    if not models:
        ap.error("--models 不能是空的")
    if not method_ids:
        ap.error("--methods 不能是空的")
    if bad := [x for x in scales if x not in config.SCALE_PROFILES]:
        ap.error(f"未知規模 {bad}；可用 {', '.join(config.SCALE_PROFILES)}")
    if bad := [x for x in models if x not in MODELS]:
        ap.error(f"未知模型 {bad}；可用 {', '.join(MODELS)}")
    if bad := [x for x in method_ids if x not in METHOD_BY_ID]:
        ap.error(f"未知方法編號 {bad}；可用 1–8")
    if args.time_limit <= 0:
        ap.error("--time-limit 必須大於 0")
    if args.scenarios is not None and args.scenarios < 1:
        ap.error("--scenarios 至少要 1")
    # 階梯的區辨性檢查：這兩個為 0 的話第 4~8 段會與第 3 段完全相同，
    # 表格會出現一整排一模一樣的數字而沒有任何警告。
    if 4 in method_ids and _RS <= 0:
        ap.error(f"config.BENDERS_ROOT_SEED_ITERS = {_RS}，第 4 段以後的 RS 不會生效，"
                 f"階梯會塌掉。請設為正數，或用 --methods 排除第 4 段以後。")
    if 5 in method_ids and _UC_ROUNDS <= 0:
        ap.error(f"config.BENDERS_ROOT_CUT_ROUNDS = {_UC_ROUNDS}，第 5 段以後的 UC "
                 f"不會生效。請設為正數，或用 --methods 排除第 5 段以後。")
    mip_gap = getattr(config, "SP_MIP_GAP", 0.01) if args.mip_gap is None else args.mip_gap

    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    tag = f"_{args.tag}" if args.tag else ""
    dr = "_DRYRUN" if args.dry_run else ""
    csv_path = RESULT_DIR / f"{RESULT_PREFIX}_raw{dr}{tag}.csv"
    xlsx_path = RESULT_DIR / f"{RESULT_PREFIX}{dr}{tag}.xlsx"
    lock_path = RESULT_DIR / f"{RESULT_PREFIX}{dr}{tag}.lock"

    with RunLock(lock_path, stale_after=3.0 * args.time_limit) as lock:
        return _run(args, scales, models, method_ids, mip_gap,
                    csv_path, xlsx_path, lock)


def _run(args, scales, models, method_ids, mip_gap, csv_path, xlsx_path, lock=None):
    if args.fresh and csv_path.exists():
        bak = csv_path.with_name(
            f"{csv_path.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}.csv")
        csv_path.replace(bak)
        print(f" 已把舊進度檔備份為 {bak.name}")

    rows = load_progress(csv_path)
    if args.dry_run:
        rows = []
    done, retry = set(), []
    for r in rows:
        k = cell_key(r)
        if k is None:
            continue
        st = r.get("status")
        if st in SUCCESS_STATUS or (args.keep_failed and st not in (None, "", "DRY_RUN")):
            done.add(k)
        elif st not in (None, "", "DRY_RUN"):
            retry.append((k, st))

    todo = [(sc, md, mid) for sc in scales for md in models for mid in method_ids
            if (sc, md, mid) not in done]
    total = len(scales) * len(models) * len(method_ids)

    print("=" * 104)
    print(" DRO 全矩陣實驗　3 規模 × 3 模糊集 × 8 方法")
    print("=" * 104)
    print(f" 規模      : {', '.join(scales)}")
    print(f" 模型      : {', '.join(MODEL_LABEL[m] for m in models)}（皆為 SP+MCVaR+DRO）")
    print(f" 方法      : {len(method_ids)} 段"
          f"（{METHOD_BY_ID[method_ids[0]][0]} … {METHOD_BY_ID[method_ids[-1]][0]}）")
    print(f" 每格時限  : {args.time_limit:,.0f} 秒 = {args.time_limit / 3600:.2f} 小時"
          f"　MIPGap {mip_gap}")
    print(f" 總格數    : {total}　　已完成 {total - len(todo)}　　待跑 {len(todo)}")
    print(f" 預估剩餘  : {len(todo) * args.time_limit / 3600:.1f} 小時"
          f" = {len(todo) * args.time_limit / 86400:.2f} 天（不含建模等額外開銷）")
    if retry:
        print(f" 自動重試  : {len(retry)} 格上次失敗，本次會重跑："
              f"{', '.join(f'{k[0]}/{k[1]}/{k[2]}({st})' for k, st in retry[:6])}"
              + ("…" if len(retry) > 6 else ""))
    print(f" 進度檔    : {csv_path.name}")
    print(f" 輸出      : {xlsx_path.name}")
    if args.dry_run:
        print(" *** DRY RUN：不會真的求解 ***")
    print("=" * 104, flush=True)

    meta = collect_meta(args, scales, models, method_ids, mip_gap)
    t_start = time.time()
    for k, (sc, md, mid) in enumerate(todo, start=1):
        mname = METHOD_BY_ID[mid][0]
        elapsed = time.time() - t_start
        eta = (elapsed / (k - 1) * (len(todo) - k + 1)) if k > 1 else \
              (len(todo) - k + 1) * args.time_limit
        print("\n" + "=" * 104)
        print(f" [{k}/{len(todo)}]　{sc}　{MODEL_LABEL[md]}　{mid}. {mname}")
        print(f" 已用 {fmt_hms(elapsed)}　預估剩餘 {fmt_hms(eta)}"
              f"　預計結束 {(datetime.now() + timedelta(seconds=eta)):%Y-%m-%d %H:%M}")
        print("=" * 104, flush=True)

        try:
            row = run_cell(sc, md, mid, args.time_limit, mip_gap, args.scenarios,
                           args.dry_run)
        except BaseException as exc:            # 最後一道防線，連 KeyboardInterrupt 也接
            row = dict(scale=sc, model=md, model_label=MODEL_LABEL[md],
                       method_id=mid, method=mname, status="Aborted",
                       error=f"{type(exc).__name__}: {exc}", sanity="",
                       finished_at=datetime.now().isoformat(timespec="seconds"))
            rows = [r for r in rows if cell_key(r) != (sc, md, mid)] + [row]
            try:
                write_csv(rows, csv_path)
                write_xlsx(rows, xlsx_path, meta, scales, models)
            except Exception:                                      # noqa: BLE001
                pass
            if isinstance(exc, KeyboardInterrupt):
                print("\n 已中斷。已完成的格子都在進度檔裡，直接重跑即可續跑。")
                return 130
            raise

        rows = [r for r in rows if cell_key(r) != (sc, md, mid)] + [row]

        # 先落地再印。之前把 print 排在寫檔前面，而 LB 為 None 時 f"{lb:,.2f}"
        # 會拋 TypeError —— 剛跑完兩小時的結果就這樣丟了，整批也跟著死。
        try:
            write_csv(rows, csv_path)
            write_xlsx(rows, xlsx_path, meta, scales, models)
        except Exception as exc:                                   # noqa: BLE001
            print(f" [!] 寫檔失敗（{type(exc).__name__}: {exc}）；結果仍在記憶體，"
                  f"下一格會再試。", flush=True)
        if lock is not None:
            lock.heartbeat()

        try:
            print(f"\n [完成] {sc}/{MODEL_LABEL[md]}/{mname}　狀態 {row.get('status')}"
                  f"　耗時 {fmt_hms(row.get('Time'))}")
            print(f"          UB={num(row.get('UB'), ',.2f')}"
                  f"　LB={num(row.get('LB'), ',.2f')}"
                  f"　Gap={num(row.get('Gap_pct'), '.3f')}%"
                  f"　Nodes={num(row.get('Nodes'), ',.0f')}", flush=True)
            if row.get("sanity"):
                print(f"          [健檢] {row['sanity']}", flush=True)
        except Exception:                                          # noqa: BLE001
            pass

    # ---------------- 收尾 ----------------
    print("\n" + "=" * 104)
    print(f" 全部結束　實際耗時 {fmt_hms(time.time() - t_start)}")
    print("=" * 104)
    idx = {cell_key(r): r for r in rows if cell_key(r) is not None}
    for sc in scales:
        print(f"\n── {sc} " + "─" * 92)
        print(f" {'Method':<30}" + "".join(f"{MODEL_LABEL[m]:>24}" for m in models))
        for mid, mname, _ in LADDER:
            if mid not in method_ids:
                continue
            line = f" {mname:<30}"
            for md in models:
                r = idx.get((sc, md, mid))
                if r is None:
                    line += f"{'未執行':>24}"
                elif isinstance(r.get("Gap_pct"), (int, float)):
                    line += f"{'Gap ' + format(r['Gap_pct'], '.2f') + '%':>24}"
                else:
                    line += f"{str(r.get('status') or '—'):>24}"
            print(line)
    flagged = [r for r in rows if r.get("sanity")]
    if flagged:
        print(f"\n 健檢提醒（{len(flagged)} 格）：")
        for r in flagged[:15]:
            print(f"   {r.get('scale')}/{r.get('model_label')}/{r.get('method')}"
                  f"　{r.get('sanity')}")
        if len(flagged) > 15:
            print(f"   …另有 {len(flagged) - 15} 格，詳見 Summary 的「健檢」欄")
    if not args.dry_run and rows:
        arch = RESULT_DIR / f"{RESULT_PREFIX}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        try:
            write_xlsx(rows, arch, meta, scales, models)
            print(f"\n 封存副本：{arch.name}")
        except Exception:                                          # noqa: BLE001
            pass
    print(f"\n Excel : {xlsx_path}")
    print(f" CSV   : {csv_path}")
    nofeas = [r for r in rows if r.get("UB") is None]
    if nofeas:
        print(f" 注意：{len(nofeas)} 格沒有取得可行解，"
              f"詳見 Detail 分頁的 status / error 欄。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
