#!/usr/bin/env python3
"""
Batch PDR experiment runner（實驗二，plan/11）。

PDR = (DRO* − MCVaR*) / MCVaR*（Jin et al. 2024 式 41；baseline 是同 α、λ
的 SP+MCVaR，不是純 SP）。固定 α = λ（預設 0.9），掃三種 ambiguity set 的
scope。共 1（MCVaR baseline）+ 3 sets × len(scopes) 次求解。
執行時先跑 ellipsoidal 最小 scope 作正式規模 pilot；成功才繼續其餘
ellipsoidal、box、polyhedral，失敗則保留輸出後停止。

This script is only a runner. It temporarily changes values in the imported
config module while it runs each case, then restores them. It does not
rewrite config.py and does not change the model core logic.

⚠ Gap 紀律：PDR 通常 < 1%，MIP_GAP 必須遠小於 PDR 訊號（預設 1e-4，
正式跑建議更緊）；baseline 與 DRO 用同一 gap、同 seed、同情境資料。

輸出（experiment result/）：
* raw CSV（來源真相，逐 case 重寫）
* Excel 四分頁：
    - "box" / "ellipsoidal" / "polyhedral"：明細表（欄位同 stress test，
      不含 VSS/EVPI；第一列 = MCVaR baseline，之後每列一個 scope，
      尾端附 scope、PDR(%) 與 risk 統計欄）
    - "PDR_table"：論文格式（Jin Table 6）——ω|PDR、a_E|PDR、a_P|PDR
      三組並排，可直接貼論文
* first_stage/{test_id}.json：完整一階解（供 out-of-sample 重用）
"""
from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import json
import math
import os
import shutil
import sys
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Any


# =============================================================================
# Parameter setting area
# =============================================================================

# ── 風險參數（Jin et al. Table 6 用 α = λ = 0.9）────────────────────────────
RISK_ALPHA_FIXED  = 0.9
RISK_LAMBDA_FIXED = 0.9

# ── scope 掃描值（box 需全部 ≤ 1/BASE_SCENARIOS，main() 會檢查）──────────────
SCOPE_VALUES = {
    "box":         [0.001, 0.005, 0.01, 0.02, 0.03],
    "ellipsoidal": [0.00005, 0.0001, 0.0005, 0.001, 0.01],
    "polyhedral":  [0.0005, 0.001, 0.005, 0.01, 0.05, 0.5],
}
AMBIGUITY_SETS = ["box", "ellipsoidal", "polyhedral"]

# ── Fixed base settings（須與實驗一相同，數字才可互相對照）────────────────────
BASE_SCENARIOS                    = 30
BASE_CCP_SAMPLE_SIZE              = None
BASE_SAMPLE_RATIO                 = 1.0
BASE_TIME_PERIODS                 = 8
BASE_DEMAND_MULTIPLIER            = 1.0
BASE_ROAD_CAPACITY_MULTIPLIER     = 1.0
BASE_HOSPITAL_CAPACITY_MULTIPLIER = 1.0

# ── Solver settings ──────────────────────────────────────────────────────────
TIME_LIMIT   = 3600.0
MIP_GAP      = 1e-4     # 正式論文表採 0.01% relative gap；太慢可放寬但要註記
COMPUTE_KPIS = False

# ── Output settings ──────────────────────────────────────────────────────────
RESULT_PREFIX   = "DRO_PDR"
LOG_SUBDIR_NAME = "dro pdr"
STOP_ON_ERROR   = False
ELLIPSOIDAL_PREFLIGHT_FIRST = True
REQUIRE_FIRST_ELLIPSOIDAL_SUCCESS = True


# =============================================================================
# Setup
# =============================================================================
ROOT_DIR   = Path(__file__).resolve().parents[1]
LOG_DIR    = ROOT_DIR / "logs"
LOG_SUBDIR = LOG_DIR / LOG_SUBDIR_NAME
RESULT_DIR = ROOT_DIR / "experiment result"
FS_DIR     = RESULT_DIR / "first_stage"

os.chdir(ROOT_DIR)
for _p in (str(ROOT_DIR / "model core"), str(ROOT_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import config as cfg  # noqa: E402

DRO_MODEL_PATH   = ROOT_DIR / "model portal" / "dro bbc.py"
MCVAR_MODEL_PATH = ROOT_DIR / "model portal" / "mcvar bbc.py"

SCOPE_SYMBOL = {"box": "ω (ε̄_B)", "ellipsoidal": "a_E", "polyhedral": "a_P"}

FIELDNAMES = [
    "test_id",
    "model",            # mcvar | dro_box | dro_ellipsoidal | dro_polyhedral
    "ambiguity_set",    # baseline 時為 "mcvar"
    "alpha",
    "lambda",
    "scope",
    "factor",
    "I", "J", "H", "S", "T",
    "obj_value",
    "PDR_pct",
    "first_stage_decision",
    "best_lb",
    "best_ub",
    "cpu_s",
    "wall_s",
    "num_vars",
    "num_constrs",
    "nodes",
    "iterations",
    "gap_pct",
    "n_opened_ccp",
    "sum_V",
    "sum_U",
    "sum_Y",
    "first_stage_cost",
    "expected_Q",
    "VaR_phi",
    "CVaR",
    "MCVaR",
    "WMCVaR",
    "worst_p_max_dev",
    "engine",
    "total_cuts",
    "oracle_solves",
    "solver_status",
    "log_path",
    "status",
    "note",
]


# =============================================================================
# Helpers（慣例同 batch_risk_experiment.py）
# =============================================================================
def blank_row() -> dict[str, Any]:
    return {key: "NA" for key in FIELDNAMES}


def write_results(csv_path: Path, rows: list[dict[str, Any]]) -> None:
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return max(0, sum(1 for row in csv.reader(f) if any(row)) - 1)


def estimate_counts() -> dict[str, int]:
    full_i = csv_row_count(ROOT_DIR / "data" / cfg.DISASTER_CSV)
    full_j = csv_row_count(ROOT_DIR / "data" / cfg.CCP_CSV)
    full_h = csv_row_count(ROOT_DIR / "data" / cfg.HOSPITAL_CSV)
    ratio = float(BASE_SAMPLE_RATIO)
    ccp = BASE_CCP_SAMPLE_SIZE
    return {
        "I": max(1, math.ceil(full_i * ratio)) if ratio < 1.0 else full_i,
        "J": min(ccp, full_j) if ccp is not None else full_j,
        "H": max(1, math.ceil(full_h * ratio)) if ratio < 1.0 else full_h,
    }


def first_stage_string(fs: dict[str, Any] | None) -> str:
    if not fs:
        return "NA"
    lines = []
    ys_by_j: dict[str, float] = {}
    for (h, j), val in fs["Y"].items():
        ys_by_j[j] = ys_by_j.get(j, 0.0) + float(val)
    for j in sorted(fs["X"]):
        if float(fs["X"][j]) > 0.5:
            lines.append(
                f"CCP {j:4s} -> X: 1, Staff(V): {float(fs['V'][j]):2.0f}, "
                f"Amb(U): {float(fs['U'][j]):2.0f}, "
                f"MedicalSupply(Y): {ys_by_j.get(j, 0.0):.2f}"
            )
    return "\n".join(lines) if lines else "none opened"


def first_stage_totals(fs: dict[str, Any] | None) -> dict[str, Any]:
    if not fs:
        return {"n_opened_ccp": "NA", "sum_V": "NA", "sum_U": "NA", "sum_Y": "NA"}
    return {
        "n_opened_ccp": sum(1 for v in fs["X"].values() if float(v) > 0.5),
        "sum_V": sum(float(v) for v in fs["V"].values()),
        "sum_U": sum(float(v) for v in fs["U"].values()),
        "sum_Y": sum(float(v) for v in fs["Y"].values()),
    }


def save_first_stage_json(test_id: str, fs: dict[str, Any] | None) -> None:
    if not fs:
        return
    FS_DIR.mkdir(parents=True, exist_ok=True)
    serializable = {
        "X": {j: float(v) for j, v in fs["X"].items()},
        "V": {j: float(v) for j, v in fs["V"].items()},
        "U": {j: float(v) for j, v in fs["U"].items()},
        "Y": {f"{h},{j}": float(v) for (h, j), v in fs["Y"].items()},
    }
    with (FS_DIR / f"{test_id}.json").open("w", encoding="utf-8") as f:
        json.dump(serializable, f, indent=2, ensure_ascii=False)


def snapshot_logs() -> set[Path]:
    if not LOG_DIR.exists():
        return set()
    return {p.resolve() for p in LOG_DIR.glob("*.log")}


def newest_created_log(before: set[Path]) -> Path | None:
    if not LOG_DIR.exists():
        return None
    created = [p for p in LOG_DIR.glob("*.log") if p.resolve() not in before]
    if not created:
        return None
    return max(created, key=lambda p: p.stat().st_mtime)


def move_log_to_subdir(log_path: Path | None) -> Path | None:
    if log_path is None or not log_path.exists():
        return log_path
    LOG_SUBDIR.mkdir(parents=True, exist_ok=True)
    dest = LOG_SUBDIR / log_path.name
    shutil.move(str(log_path), str(dest))
    return dest


# =============================================================================
# Context managers (temporarily patch config, restore on exit)
# =============================================================================
@contextmanager
def temporary_config():
    keys: dict[str, Any] = {
        "SCENARIOS":                    BASE_SCENARIOS,
        "TIME_PERIODS":                 BASE_TIME_PERIODS,
        "SAMPLE_RATIO":                 BASE_SAMPLE_RATIO,
        "SP_SAMPLE_RATIO":              BASE_SAMPLE_RATIO,
        "DEMAND_MULTIPLIER":            BASE_DEMAND_MULTIPLIER,
        "ROAD_CAPACITY_MULTIPLIER":     BASE_ROAD_CAPACITY_MULTIPLIER,
        "HOSPITAL_CAPACITY_MULTIPLIER": BASE_HOSPITAL_CAPACITY_MULTIPLIER,
        "SP_TIME_LIMIT":                TIME_LIMIT,
        "SP_MIP_GAP":                   MIP_GAP,
    }
    if hasattr(cfg, "CCP_SAMPLE_SIZE"):
        keys["CCP_SAMPLE_SIZE"] = BASE_CCP_SAMPLE_SIZE
    original = {key: getattr(cfg, key) for key in keys}
    try:
        for key, value in keys.items():
            setattr(cfg, key, value)
        yield
    finally:
        for key, value in original.items():
            setattr(cfg, key, value)


@contextmanager
def patched_generate_data():
    original = cfg.generate_data

    def _patched(*args, **kwargs):
        kwargs.setdefault("ccp_sample_size", BASE_CCP_SAMPLE_SIZE)
        return original(*args, **kwargs)

    cfg.generate_data = _patched
    try:
        yield
    finally:
        cfg.generate_data = original


@contextmanager
def patched_generate_scenarios():
    original = cfg.generate_scenarios

    def _patched(*args, **kwargs):
        kwargs.setdefault("demand_multiplier",            BASE_DEMAND_MULTIPLIER)
        kwargs.setdefault("road_capacity_multiplier",     BASE_ROAD_CAPACITY_MULTIPLIER)
        kwargs.setdefault("hospital_capacity_multiplier", BASE_HOSPITAL_CAPACITY_MULTIPLIER)
        kwargs.setdefault("num_periods",                  BASE_TIME_PERIODS)
        return original(*args, **kwargs)

    cfg.generate_scenarios = _patched
    try:
        yield
    finally:
        cfg.generate_scenarios = original


# =============================================================================
# Core run logic
# =============================================================================
def load_portal(path: Path, name: str) -> Any:
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _fill_common(row: dict[str, Any], model: Any, summary: dict[str, Any]) -> None:
    fs = summary.get("first_stage")
    risk = summary.get("risk", {})
    st = summary.get("bbc_stats", {})
    row.update({
        "obj_value": f"{summary['objective']:.4f}",
        "first_stage_decision": first_stage_string(fs),
        "best_lb":  f"{summary['best_lb']:.4f}",
        "best_ub":  f"{summary['objective']:.4f}",
        "cpu_s":    f"{st.get('runtime', float('nan')):.2f}",
        "num_vars":    getattr(model, "NumVars", "NA"),
        "num_constrs": getattr(model, "NumConstrs", "NA"),
        "nodes":       f"{getattr(model, 'NodeCount', float('nan')):.0f}",
        "iterations":  f"{getattr(model, 'IterCount', float('nan')):.0f}",
        "gap_pct":  f"{summary['gap_pct']:.4f}",
        "first_stage_cost": f"{summary.get('first_stage_cost', float('nan')):.2f}",
        "expected_Q": f"{risk.get('expected_Q', float('nan')):.2f}",
        "VaR_phi":    f"{risk.get('phi_star_VaR', float('nan')):.2f}",
        "CVaR":       f"{risk.get('CVaR', float('nan')):.2f}",
        "MCVaR":      f"{risk.get('MCVaR', float('nan')):.2f}",
        "WMCVaR":     f"{risk.get('WMCVaR', float('nan')):.2f}"
                      if "WMCVaR" in risk else "NA",
        "worst_p_max_dev": f"{risk.get('worst_p_max_dev', float('nan')):.6f}"
                           if "worst_p_max_dev" in risk else "NA",
        "engine":        st.get("engine", "NA"),
        "total_cuts":    st.get("cuts_added", "NA"),
        "oracle_solves": st.get("oracle_solves", "NA"),
        "solver_status": st.get("solver_status", "NA"),
        "status": "OK",
    })
    row.update(first_stage_totals(fs))


def _run_case(runner, row: dict[str, Any], test_id: str) -> dict[str, Any]:
    """共用外框：log 快照/搬移、計時、例外處理。runner() 回傳 (model, summary)。"""
    logs_before = snapshot_logs()
    wall_start = time.time()
    try:
        try:
            with (
                temporary_config(),
                patched_generate_data(),
                patched_generate_scenarios(),
            ):
                model, summary = runner()
        finally:
            log_path = move_log_to_subdir(newest_created_log(logs_before))
            row["log_path"] = str(log_path) if log_path else "NA"
            row["wall_s"] = f"{time.time() - wall_start:.2f}"
    except Exception as exc:  # noqa: BLE001
        row["status"] = "FAIL"
        row["note"] = f"{type(exc).__name__}: {exc}"
        print(f"  -> FAIL: {row['note']}")
        if STOP_ON_ERROR:
            raise
        return row
    if model is None or summary is None:
        row["status"] = "FAIL"
        row["note"] = "no feasible solution (see log)"
        print(f"  -> FAIL: {row['note']}")
        return row
    _fill_common(row, model, summary)
    save_first_stage_json(test_id, summary.get("first_stage"))
    print(f"  -> OK obj={row['obj_value']} gap={row['gap_pct']}% cpu={row['cpu_s']}s")
    return row


def run_mcvar_baseline(mcvar_module: Any, counts: dict[str, int]) -> dict[str, Any]:
    test_id = f"pdr_mcvar_a{RISK_ALPHA_FIXED}_l{RISK_LAMBDA_FIXED}".replace(".", "p")
    row = blank_row()
    row.update({
        "test_id": test_id,
        "model": "mcvar",
        "ambiguity_set": "mcvar",
        "alpha": RISK_ALPHA_FIXED,
        "lambda": RISK_LAMBDA_FIXED,
        "scope": 0.0,
        "factor": f"MCVaR baseline, alpha={RISK_ALPHA_FIXED}, lambda={RISK_LAMBDA_FIXED}",
        "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "PDR_pct": 0.0,
        "status": "RUNNING", "note": "",
    })
    print(f"\n[baseline] {test_id} | S={BASE_SCENARIOS}, gap={MIP_GAP}")
    return _run_case(
        lambda: mcvar_module.run_mcvar_model(
            scenario_size=BASE_SCENARIOS,
            sample_ratio=BASE_SAMPLE_RATIO,
            time_limit=TIME_LIMIT,
            mip_gap=MIP_GAP,
            alpha=RISK_ALPHA_FIXED,
            lam=RISK_LAMBDA_FIXED,
            compute_kpis=COMPUTE_KPIS,
        ),
        row, test_id,
    )


def run_dro_case(dro_module: Any, run_idx: int, total: int, aset: str,
                 scope: float, counts: dict[str, int],
                 baseline_obj: float | None) -> dict[str, Any]:
    test_id = f"pdr_{aset}_e{scope}".replace(".", "p")
    row = blank_row()
    row.update({
        "test_id": test_id,
        "model": f"dro_{aset}",
        "ambiguity_set": aset,
        "alpha": RISK_ALPHA_FIXED,
        "lambda": RISK_LAMBDA_FIXED,
        "scope": scope,
        "factor": f"{aset}, {SCOPE_SYMBOL[aset]}={scope:g}",
        "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "status": "RUNNING", "note": "",
    })
    print(f"\n[{run_idx}/{total}] {test_id} | scope={scope:g}")
    row = _run_case(
        lambda: dro_module.run_dro_model(
            ambiguity_set=aset,
            scenario_size=BASE_SCENARIOS,
            sample_ratio=BASE_SAMPLE_RATIO,
            time_limit=TIME_LIMIT,
            mip_gap=MIP_GAP,
            alpha=RISK_ALPHA_FIXED,
            lam=RISK_LAMBDA_FIXED,
            scope=scope,
            compute_kpis=COMPUTE_KPIS,
        ),
        row, test_id,
    )
    if row["status"] == "OK" and baseline_obj:
        pdr = (float(row["obj_value"]) - baseline_obj) / baseline_obj * 100.0
        row["PDR_pct"] = f"{pdr:.4f}"
        gap_tol = (float(row["gap_pct"]) + MIP_GAP * 100.0)
        if pdr < -gap_tol:
            row["note"] = (row["note"] + " " if row["note"] else "") + \
                "PDR<0 超出 gap 容差：請調緊 MIP_GAP 重跑"
            print(f"  [WARN] PDR={pdr:.4f}% < 0 超出 gap 容差")
        print(f"  -> PDR = {pdr:.4f}%")
    return row


# =============================================================================
# Excel export（四分頁：三明細 + PDR_table 論文格式）
# =============================================================================
XLSX_DETAIL_COLUMNS = [
    ("factor",               "factor"),
    ("|I| Disaster",         "I"),
    ("|J| CCP",              "J"),
    ("|H| Hosp",             "H"),
    ("|S| Scen",             "S"),
    ("|T| Per",              "T"),
    ("obj_value",            "obj_value"),
    ("First Stage Decision", "first_stage_decision"),
    ("Best LB",              "best_lb"),
    ("Best UB",              "best_ub"),
    ("CPU Time(s)",          "cpu_s"),
    ("num_vars",             "num_vars"),
    ("num_constrs",          "num_constrs"),
    ("Nodes",                "nodes"),
    ("Iteration",            "iterations"),
    ("Final Gap(%)",         "gap_pct"),
    ("scope",                "scope"),
    ("PDR(%)",               "PDR_pct"),
    ("alpha",                "alpha"),
    ("lambda",               "lambda"),
    ("Opened CCPs",          "n_opened_ccp"),
    ("Sum V",                "sum_V"),
    ("Sum U",                "sum_U"),
    ("Sum Y",                "sum_Y"),
    ("E[Q] (p0)",            "expected_Q"),
    ("CVaR (p0)",            "CVaR"),
    ("MCVaR (p0)",           "MCVaR"),
    ("WMCVaR",               "WMCVaR"),
    ("worst_p_max_dev",      "worst_p_max_dev"),
    ("Solver Status",        "solver_status"),
    ("Status",               "status"),
]

_TEXT_KEYS = ("factor", "first_stage_decision", "engine", "solver_status",
              "status", "log_path", "note", "ambiguity_set", "test_id", "model")


def export_xlsx(rows: list[dict[str, Any]], xlsx_path: Path) -> None:
    """CSV 為來源真相；匯出失敗（openpyxl 未裝、檔案被開啟）不中斷實驗。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [xlsx] openpyxl 未安裝，略過 Excel 匯出（pip install openpyxl）")
        return
    try:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)
        header_fill = PatternFill("solid", fgColor="2E74B5")
        header_font = Font(bold=True, color="FFFFFF")
        baseline = next((r for r in rows if r.get("model") == "mcvar"), None)

        # ---- 三個明細分頁（第一列 = MCVaR baseline）------------------------
        for aset in AMBIGUITY_SETS:
            ws = wb.create_sheet(aset)
            for col_idx, (title, _key) in enumerate(XLSX_DETAIL_COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=title)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center", wrap_text=True)
            sheet_rows = ([baseline] if baseline else []) + [
                r for r in rows if r.get("ambiguity_set") == aset
            ]
            for r_idx, row in enumerate(sheet_rows, start=2):
                for col_idx, (_title, key) in enumerate(XLSX_DETAIL_COLUMNS, start=1):
                    value = row.get(key, "NA")
                    if key not in _TEXT_KEYS:
                        try:
                            value = float(value)
                        except (TypeError, ValueError):
                            pass
                    cell = ws.cell(row=r_idx, column=col_idx, value=value)
                    if key == "first_stage_decision":
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["H"].width = 62

        # ---- PDR_table（Jin et al. Table 6 論文格式）-----------------------
        ws = wb.create_sheet("PDR_table")
        group_titles = [("Box ambiguity set", "ω"),
                        ("Ellipsoidal ambiguity set", "a_E"),
                        ("Polyhedral ambiguity set", "a_P")]
        for g_idx, (title, sym) in enumerate(group_titles):
            c0 = 1 + g_idx * 2
            ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + 1)
            cell = ws.cell(row=1, column=c0, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.cell(row=2, column=c0, value=sym).font = Font(bold=True)
            ws.cell(row=2, column=c0 + 1, value="PDR").font = Font(bold=True)
        n_max = max(len(v) for v in SCOPE_VALUES.values())
        for i in range(n_max):
            for g_idx, aset in enumerate(AMBIGUITY_SETS):
                c0 = 1 + g_idx * 2
                scopes = SCOPE_VALUES[aset]
                if i >= len(scopes):
                    continue
                scope = scopes[i]
                ws.cell(row=3 + i, column=c0, value=scope)
                match = next(
                    (r for r in rows
                     if r.get("ambiguity_set") == aset and r.get("scope") == scope),
                    None,
                )
                if match is None:
                    val = None
                elif match.get("status") != "OK":
                    val = match.get("status", "NA")
                else:
                    try:
                        val = f"{float(match['PDR_pct']):.4f}%"
                    except (TypeError, ValueError):
                        val = "NA"
                ws.cell(row=3 + i, column=c0 + 1, value=val)
        note_row = 4 + n_max
        base_txt = "NA" if baseline is None else baseline.get("obj_value", "NA")
        ws.cell(row=note_row, column=1, value=(
            f"MCVaR baseline obj = {base_txt} "
            f"(alpha={RISK_ALPHA_FIXED}, lambda={RISK_LAMBDA_FIXED}, "
            f"S={BASE_SCENARIOS}, mip_gap={MIP_GAP}); "
            f"PDR = (DRO* - MCVaR*) / MCVaR*"
        ))
        for c in range(1, 7):
            ws.column_dimensions[get_column_letter(c)].width = 14

        wb.save(xlsx_path)
    except Exception as exc:  # noqa: BLE001
        print(f"  [xlsx] Excel 匯出失敗（不影響 CSV）: {exc}")


# =============================================================================
# Main
# =============================================================================
def main() -> None:
    bad = [s for s in SCOPE_VALUES["box"] if s > 1.0 / BASE_SCENARIOS + 1e-12]
    if bad:
        raise ValueError(
            f"box scopes {bad} > 1/S = {1.0 / BASE_SCENARIOS:.6f}；"
            f"等權重下 box ambiguity set 需 ε̄_B ≤ 1/S。"
        )

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path  = RESULT_DIR / f"{RESULT_PREFIX}_raw_{timestamp}.csv"
    xlsx_path = RESULT_DIR / f"{RESULT_PREFIX}_{timestamp}.xlsx"

    execution_sets = list(AMBIGUITY_SETS)
    if ELLIPSOIDAL_PREFLIGHT_FIRST:
        execution_sets = ["ellipsoidal"] + [
            aset for aset in execution_sets if aset != "ellipsoidal"
        ]
    cases = [(aset, scope)
             for aset in execution_sets for scope in SCOPE_VALUES[aset]]
    counts = estimate_counts()

    print("=" * 70)
    print("BATCH PDR EXPERIMENT (實驗二)")
    print("=" * 70)
    print(f"alpha=lambda={RISK_ALPHA_FIXED}/{RISK_LAMBDA_FIXED}")
    print(f"scopes={SCOPE_VALUES}")
    print(f"S={BASE_SCENARIOS} T={BASE_TIME_PERIODS} mip_gap={MIP_GAP} "
          f"time_limit={TIME_LIMIT}")
    print(f"cases=1 baseline + {len(cases)} DRO")
    print(f"CSV   : {csv_path}")
    print(f"Excel : {xlsx_path}")

    mcvar_module = load_portal(MCVAR_MODEL_PATH, "mcvar_bbc_portal")
    dro_module   = load_portal(DRO_MODEL_PATH, "dro_bbc_portal")

    rows: list[dict[str, Any]] = []
    base_row = run_mcvar_baseline(mcvar_module, counts)
    rows.append(base_row)
    write_results(csv_path, rows)
    export_xlsx(rows, xlsx_path)

    baseline_obj = None
    if base_row["status"] == "OK":
        baseline_obj = float(base_row["obj_value"])
    else:
        print("[WARN] MCVaR baseline 失敗，後續 PDR 無法計算（只記 DRO obj）。")

    for idx, (aset, scope) in enumerate(cases, start=1):
        row = run_dro_case(dro_module, idx, len(cases), aset, scope,
                           counts, baseline_obj)
        rows.append(row)
        write_results(csv_path, rows)
        export_xlsx(rows, xlsx_path)
        if (idx == 1 and ELLIPSOIDAL_PREFLIGHT_FIRST
                and REQUIRE_FIRST_ELLIPSOIDAL_SUCCESS
                and row.get("status") != "OK"):
            print("\n[ABORT] 第一個 ellipsoidal PDR pilot 失敗；已保留 "
                  "CSV/Excel/log，不執行後續 DRO cases。")
            break
        if (idx == 1 and ELLIPSOIDAL_PREFLIGHT_FIRST
                and REQUIRE_FIRST_ELLIPSOIDAL_SUCCESS):
            print("\n[PILOT PASS] 第一個 ellipsoidal PDR case 成功，"
                  "繼續其餘 ellipsoidal、box、polyhedral cases。")

    # ---- console PDR 表與單調性檢查 ----------------------------------------
    print("\n" + "=" * 70)
    print(f"PDR SUMMARY (alpha=lambda={RISK_ALPHA_FIXED}; "
          f"baseline obj={baseline_obj})")
    warnings = []
    for aset in AMBIGUITY_SETS:
        print(f"\n  {aset}:")
        prev = None
        for scope in SCOPE_VALUES[aset]:
            match = next((r for r in rows if r.get("ambiguity_set") == aset
                          and r.get("scope") == scope), None)
            pdr_txt = "NA"
            if match and match.get("status") == "OK":
                try:
                    pdr = float(match["PDR_pct"])
                    pdr_txt = f"{pdr:.4f}%"
                    if prev is not None and pdr < prev - (2 * MIP_GAP * 100.0):
                        warnings.append(
                            f"{aset}: PDR 隨 scope 下降超出 gap 容差 "
                            f"({prev:.4f}% -> {pdr:.4f}% @ scope={scope:g})"
                        )
                    prev = pdr
                except (TypeError, ValueError):
                    pass
            print(f"    {SCOPE_SYMBOL[aset]:>10s} = {scope:<10g} PDR = {pdr_txt}")
    n_ok = sum(1 for r in rows if r.get("status") == "OK")
    print("\n" + "-" * 70)
    print(f"Done: {n_ok}/{len(rows)} cases OK.")
    for w in warnings:
        print(f"  [WARN] {w}")
    print(f"CSV   : {csv_path}")
    print(f"Excel : {xlsx_path}")


if __name__ == "__main__":
    main()
