#!/usr/bin/env python3
"""
Batch DRO risk-parameter experiment runner（實驗一）。

全網格：scale x ambiguity set x alpha x lambda
  scales = small / medium / large（plan/15 SCALE_PROFILES）
  sets   = box / ellipsoidal / polyhedral
  alpha  = [0.5, 0.6, 0.7, 0.8, 0.9]
  lambda = [0.3, 0.5, 0.7, 0.9]
  → 3 x 3 x 5 x 4 = 180 cases

模型 = SP + MCVaR + DRO，B&BC 引擎且「所有加速全開」
（EV warm start + root seeding + user cuts + Pareto + multi-cut）。

停止條件：每個 case 跑滿 TIME_LIMIT(1 小時) 或 gap 收斂到 MIP_GAP(5%)
其一滿足即停。

輸出（experiment result/，檔名帶 timestamp、不覆寫）：
  * raw CSV：每 case 一列、逐 case 重寫（來源真相、含完整欄位）。
  * Excel：
      - 9 個 (scale x set) 分頁，順序 small_box, small_ellipsoidal,
        small_polyhedral, medium_box, ... , large_polyhedral；每頁是
        alpha(列) x lambda(欄) 的三區塊矩陣（obj_value / CPU Time / Final Gap）。
      - 最後一頁 "ALL_summary"：9 組矩陣依序堆疊，逐組標明 size 與 set。
  * 不輸出 log 檔（本 runner 執行期間停用 logging_utils.tee_output）。

本檔只是 runner：暫時改 config 的值、跑完還原；不改 config.py、不改 model core。
"""
from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import os
import sys
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Any


# =============================================================================
# Parameter setting area
# =============================================================================

SCALES         = ["small", "medium", "large"]
AMBIGUITY_SETS = ["box", "ellipsoidal", "polyhedral"]
ALPHA_VALUES   = [0.5, 0.6, 0.7, 0.8, 0.9]
LAMBDA_VALUES  = [0.3, 0.5, 0.7, 0.9]

# ambiguity scope（各 set 的 scope；box 需 <= 1/BASE_SCENARIOS，main() 會檢查）
SCOPES = {
    "box":         0.01,
    "ellipsoidal": 0.0005,
    "polyhedral":  0.001,
}

# ── Fixed base settings ─────────────────────────────────────────────────────
BASE_SCENARIOS                    = 30
BASE_TIME_PERIODS                 = 8
BASE_DEMAND_MULTIPLIER            = 1.0
BASE_ROAD_CAPACITY_MULTIPLIER     = 1.0
BASE_HOSPITAL_CAPACITY_MULTIPLIER = 1.0
BASE_SAMPLE_RATIO                 = 1.0   # 走 scale 路徑時忽略（僅為相容欄位）

# ── Solver settings ─────────────────────────────────────────────────────────
TIME_LIMIT   = 3600.0   # 每個 case 最多跑 1 小時
MIP_GAP      = 0.05     # 或收斂到 5% relative gap 即停（兩條件先到者為準）
COMPUTE_KPIS = False    # 省時；要 KPI 重解改 True
# 數值穩定度（0=關,1=預設,2/3=更穩但較慢）。ellipsoidal 是 MISOCP，較易數值失敗，
# 拉高可減少「求解成功但數值報錯」的 FAIL。
SOLVER_NUMERIC_FOCUS = 2

# ── Output settings ─────────────────────────────────────────────────────────
RESULT_PREFIX = "DRO_scale_set_alpha_lambda"
STOP_ON_ERROR = False   # 單一 case 失敗記 FAIL 續跑，不中斷整批


# =============================================================================
# Setup
# =============================================================================
ROOT_DIR   = Path(__file__).resolve().parents[1]
RESULT_DIR = ROOT_DIR / "experiment result"

os.chdir(ROOT_DIR)
for _p in (str(ROOT_DIR / "model core"), str(ROOT_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import config as cfg          # noqa: E402
import logging_utils          # noqa: E402

DRO_MODEL_PATH = ROOT_DIR / "model portal" / "dro bbc.py"

# B&BC 全加速值（讀 config 的既有設定當「開啟」值）
FULL_ROOT_SEED_ITERS = int(getattr(cfg, "BENDERS_ROOT_SEED_ITERS", 300))
FULL_ROOT_CUT_ROUNDS = int(getattr(cfg, "BENDERS_ROOT_CUT_ROUNDS", 15))

FIELDNAMES = [
    "scale", "test_id", "ambiguity_set", "alpha", "lambda", "scope", "factor",
    "I", "J", "H", "S", "T",
    "obj_value", "first_stage_decision", "best_lb", "best_ub",
    "cpu_s", "wall_s", "num_vars", "num_constrs", "nodes", "iterations", "gap_pct",
    # risk 統計
    "n_opened_ccp", "sum_V", "sum_U", "sum_Y", "first_stage_cost",
    "expected_Q", "VaR_phi", "CVaR", "MCVaR", "WMCVaR", "worst_p_max_dev",
    # B&BC 引擎統計
    "engine", "total_cuts", "seed_cuts", "lazy_cuts", "user_cuts",
    "root_seed_iters_done", "root_seed_lb", "root_cut_rounds_done",
    "parallel_oracles", "oracle_solves", "incumbent_evals",
    "solver_status", "status", "note",
]


# =============================================================================
# Helpers
# =============================================================================
def blank_row() -> dict[str, Any]:
    return {key: "NA" for key in FIELDNAMES}


def scale_counts(scale: str) -> dict[str, int]:
    p = cfg.SCALE_PROFILES[scale]
    return {"I": int(p["n_disaster"]), "J": int(p["n_ccp"]), "H": int(p["n_hospital"])}


def write_results(csv_path: Path, rows: list[dict[str, Any]]) -> None:
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def first_stage_string(fs: dict[str, Any] | None) -> str:
    if not fs:
        return "NA"
    ys_by_j: dict[str, float] = {}
    for (h, j), val in fs["Y"].items():
        ys_by_j[j] = ys_by_j.get(j, 0.0) + float(val)
    lines = []
    for j in sorted(fs["X"]):
        if float(fs["X"][j]) > 0.5:
            lines.append(
                f"CCP {j:4s} -> X: 1, Staff(V): {float(fs['V'][j]):2.0f}, "
                f"Amb(U): {float(fs['U'][j]):2.0f}, "
                f"MedicalSupply(Y): {ys_by_j.get(j, 0.0):.2f}"
            )
    return "\n".join(lines) if lines else "none opened"


def first_stage_totals(fs: dict[str, Any] | None) -> dict[str, Any]:
    if not fs:
        return {"n_opened_ccp": "NA", "sum_V": "NA", "sum_U": "NA", "sum_Y": "NA"}
    return {
        "n_opened_ccp": sum(1 for v in fs["X"].values() if float(v) > 0.5),
        "sum_V": sum(float(v) for v in fs["V"].values()),
        "sum_U": sum(float(v) for v in fs["U"].values()),
        "sum_Y": sum(float(v) for v in fs["Y"].values()),
    }


# =============================================================================
# Context managers（暫時改 config，跑完還原）
# =============================================================================
@contextmanager
def temporary_config():
    """固定基礎設定 + B&BC 全加速全開。"""
    keys: dict[str, Any] = {
        "SCENARIOS":                    BASE_SCENARIOS,
        "TIME_PERIODS":                 BASE_TIME_PERIODS,
        "DEMAND_MULTIPLIER":            BASE_DEMAND_MULTIPLIER,
        "ROAD_CAPACITY_MULTIPLIER":     BASE_ROAD_CAPACITY_MULTIPLIER,
        "HOSPITAL_CAPACITY_MULTIPLIER": BASE_HOSPITAL_CAPACITY_MULTIPLIER,
        "SP_TIME_LIMIT":                TIME_LIMIT,
        "SP_MIP_GAP":                   MIP_GAP,
        # B&BC 所有加速全開
        "BENDERS_MULTI_CUT":       True,
        "BENDERS_EV_WARM_START":   True,
        "BENDERS_ROOT_SEED_ITERS": FULL_ROOT_SEED_ITERS,
        "BENDERS_ROOT_CUT_ROUNDS": FULL_ROOT_CUT_ROUNDS,
        "BENDERS_USE_USER_CUTS":   True,
        "BENDERS_PARETO_ENABLED":  True,
        "BENDERS_NUMERIC_FOCUS":   SOLVER_NUMERIC_FOCUS,
    }
    original = {key: getattr(cfg, key) for key in keys if hasattr(cfg, key)}
    try:
        for key, value in keys.items():
            if hasattr(cfg, key):
                setattr(cfg, key, value)
        yield
    finally:
        for key, value in original.items():
            setattr(cfg, key, value)


@contextmanager
def scaled_instance_generation():
    """依 cfg.EXPERIMENT_SCALE 產生 instance；同 scale 的所有 case 共用一份快取。"""
    original = cfg.generate_data
    cache: dict[str, Any] = {}

    def _gen(*args, **kwargs):
        kwargs.pop("sample_ratio", None)
        kwargs.pop("ccp_sample_size", None)
        scale = cfg.EXPERIMENT_SCALE
        if scale not in cache:
            cache[scale] = original(scale=scale)
        return cache[scale]

    cfg.generate_data = _gen
    try:
        yield
    finally:
        cfg.generate_data = original


@contextmanager
def suppress_solver_logs():
    """停用 logging_utils.tee_output → 不產生任何 .log 檔（console 仍會顯示）。"""
    original = logging_utils.tee_output

    @contextmanager
    def _null(*_args, **_kwargs):
        yield

    logging_utils.tee_output = _null
    try:
        yield
    finally:
        logging_utils.tee_output = original


@contextmanager
def robust_risk_summary():
    """讓「求解成功、但事後 WMCVaR 風險摘要（ellipsoidal 的 SOCP）數值失敗」
    不再整個 case FAIL。

    dro bbc 的 _run 在解完後一定會呼叫 risk_core.risk_summary_from_Q；其中
    ellipsoidal 的 WMCVaR 要解一個小型 SOCP，Q~1e7 搭配極小 scope 時 barrier
    可能數值失敗並 raise → 明明主問題解好了卻被丟掉。這裡在 runner 執行期間把
    該函式包成 exception-safe：失敗時保留不需解算的統計（E[Q]/VaR/CVaR/MCVaR），
    只把 WMCVaR/worst_p 設為 NaN，讓 obj/gap/cpu 等主結果照常寫入。
    """
    import risk_core  # 延遲載入（避免 sandbox 無 gurobi 時 import 失敗）

    original = risk_core.risk_summary_from_Q

    def _safe(q_by_s, p0, risk_cfg, _orig=original):
        try:
            return _orig(q_by_s, p0, risk_cfg)
        except Exception as exc:  # noqa: BLE001
            alpha = risk_cfg["alpha"]
            lam = risk_cfg["lambda"]
            exp_q = sum(p0[s] * q_by_s[s] for s in q_by_s)
            try:
                mcvar, phi = risk_core.evaluate_mcvar(q_by_s, p0, alpha, lam)
                tail = sum(p0[s] * max(q_by_s[s] - phi, 0.0) for s in q_by_s)
                cvar = phi + tail / (1.0 - alpha) if alpha < 1.0 else float("nan")
            except Exception:  # noqa: BLE001
                mcvar = phi = cvar = float("nan")
            print(f"  [robust] WMCVaR 事後評估失敗（保留主結果）: "
                  f"{type(exc).__name__}: {exc}")
            return {
                "alpha": alpha, "lambda": lam, "expected_Q": exp_q,
                "phi_star_VaR": phi, "CVaR": cvar, "MCVaR": mcvar,
                "WMCVaR": float("nan"), "worst_p": None,
                "worst_p_max_dev": float("nan"),
                "scope": risk_core.risk_scope(risk_cfg),
                "wmcvar_note": f"WMCVaR eval failed: {type(exc).__name__}",
            }

    risk_core.risk_summary_from_Q = _safe
    try:
        yield
    finally:
        risk_core.risk_summary_from_Q = original


def classify_failure(msg: str) -> str:
    """把例外訊息歸類，寫進 solver_status 方便逐格診斷。"""
    low = msg.lower()
    if "oracle" in low and "not optimal" in low:
        return "ORACLE_LP_NOT_OPTIMAL"
    if "wmcvar" in low or "evaluate_wmcvar" in low:
        return "WMCVAR_EVAL_FAILED"
    if "ambiguity set requires" in low or "epsilon_box" in low:
        return "SCOPE_INVALID"
    if "multi-cut" in low or "single-cut" in low:
        return "NEEDS_MULTI_CUT"
    if "infeasible" in low:
        return "INFEASIBLE"
    if "out of memory" in low or "memory" in low:
        return "OUT_OF_MEMORY"
    return "ERROR"


# =============================================================================
# Core run logic
# =============================================================================
def load_dro_module() -> Any:
    spec = importlib.util.spec_from_file_location("dro_bbc_portal", DRO_MODEL_PATH)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def run_one_case(dro_module: Any, run_idx: int, total: int, scale: str,
                 aset: str, alpha: float, lam: float,
                 counts: dict[str, int]) -> dict[str, Any]:
    scope = SCOPES[aset]
    test_id = f"{scale}_{aset}_a{alpha}_l{lam}".replace(".", "p")
    row = blank_row()
    row.update({
        "scale": scale, "test_id": test_id, "ambiguity_set": aset,
        "alpha": alpha, "lambda": lam, "scope": scope,
        "factor": f"{scale}, {aset}, alpha={alpha}, lambda={lam}",
        "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "status": "RUNNING", "note": "",
    })
    print(f"\n[{run_idx}/{total}] {test_id} | scope={scope} "
          f"| I={counts['I']} J={counts['J']} H={counts['H']} S={BASE_SCENARIOS}")

    wall_start = time.time()
    model = summary = None
    try:
        with temporary_config():
            model, summary = dro_module.run_dro_model(
                ambiguity_set=aset,
                scenario_size=BASE_SCENARIOS,
                sample_ratio=BASE_SAMPLE_RATIO,
                time_limit=TIME_LIMIT,
                mip_gap=MIP_GAP,
                alpha=alpha,
                lam=lam,
                scope=scope,
                engine="bbc",
                compute_kpis=COMPUTE_KPIS,
            )
        row["wall_s"] = f"{time.time() - wall_start:.2f}"
    except Exception as exc:  # noqa: BLE001
        row["wall_s"] = f"{time.time() - wall_start:.2f}"
        row["status"] = "FAIL"
        row["note"] = f"{type(exc).__name__}: {exc}"
        row["solver_status"] = classify_failure(str(exc))
        print(f"  -> FAIL [{row['solver_status']}]: {row['note']}")
        if STOP_ON_ERROR:
            raise
        return row

    if model is None or summary is None:
        row["status"] = "FAIL"
        row["note"] = "no feasible solution / no incumbent within time limit"
        row["solver_status"] = "NO_INCUMBENT"
        print("  -> FAIL [NO_INCUMBENT]: no feasible solution")
        return row

    fs = summary.get("first_stage")
    risk = summary.get("risk", {})
    st = summary.get("bbc_stats", {})
    row.update({
        "obj_value": f"{summary['objective']:.4f}",
        "first_stage_decision": first_stage_string(fs),
        "best_lb": f"{summary['best_lb']:.4f}",
        "best_ub": f"{summary['objective']:.4f}",
        "cpu_s": f"{st.get('runtime', float('nan')):.2f}",
        "num_vars": getattr(model, "NumVars", "NA"),
        "num_constrs": getattr(model, "NumConstrs", "NA"),
        "nodes": f"{getattr(model, 'NodeCount', float('nan')):.0f}",
        "iterations": f"{getattr(model, 'IterCount', float('nan')):.0f}",
        "gap_pct": f"{summary['gap_pct']:.4f}",
        "first_stage_cost": f"{summary.get('first_stage_cost', float('nan')):.2f}",
        "expected_Q": f"{risk.get('expected_Q', float('nan')):.2f}",
        "VaR_phi": f"{risk.get('phi_star_VaR', float('nan')):.2f}",
        "CVaR": f"{risk.get('CVaR', float('nan')):.2f}",
        "MCVaR": f"{risk.get('MCVaR', float('nan')):.2f}",
        "WMCVaR": f"{risk.get('WMCVaR', float('nan')):.2f}",
        "worst_p_max_dev": f"{risk.get('worst_p_max_dev', float('nan')):.6f}",
        "engine": st.get("engine", "NA"),
        "total_cuts": st.get("cuts_added", "NA"),
        "seed_cuts": st.get("seed_cuts_added", "NA"),
        "lazy_cuts": st.get("lazy_cuts_added", "NA"),
        "user_cuts": st.get("user_cuts_added", "NA"),
        "root_seed_iters_done": st.get("root_seed_iters_done", "NA"),
        "root_seed_lb": st.get("root_seed_lb", "NA"),
        "root_cut_rounds_done": st.get("root_cut_rounds_done", "NA"),
        "parallel_oracles": st.get("parallel_oracles", "NA"),
        "oracle_solves": st.get("oracle_solves", "NA"),
        "incumbent_evals": st.get("incumbent_evals", "NA"),
        "solver_status": st.get("solver_status", "NA"),
        "status": "OK",
    })
    row.update(first_stage_totals(fs))
    print(f"  -> OK obj={row['obj_value']} gap={row['gap_pct']}% cpu={row['cpu_s']}s")
    return row


# =============================================================================
# Excel export（9 個 scale x set 矩陣分頁 + 最後 ALL_summary 總表）
# =============================================================================
MATRIX_BLOCKS = [("obj_value", "obj_value"), ("CPU Time(s)", "cpu_s"),
                 ("Final Gap(%)", "gap_pct")]


def scale_set_order() -> list[tuple[str, str]]:
    return [(scale, aset) for scale in SCALES for aset in AMBIGUITY_SETS]


def _matrix_value(rows, scale, aset, alpha, lam, key):
    for row in rows:
        if (row.get("scale") == scale and row.get("ambiguity_set") == aset
                and row.get("alpha") == alpha and row.get("lambda") == lam):
            if row.get("status") != "OK":
                return row.get("status", "NA")
            try:
                return float(row.get(key))
            except (TypeError, ValueError):
                return row.get(key, "NA")
    return None  # 尚未跑到


def _write_matrix(ws, r, rows, scale, aset, styles):
    """在 ws 從第 r 列開始寫 obj/CPU/gap 三區塊；回傳下一個可用列。"""
    fill, font, center, bold = styles
    for block_title, key in MATRIX_BLOCKS:
        c = ws.cell(r, 1, block_title)
        c.fill, c.font, c.alignment = fill, font, center
        for ci, lam in enumerate(LAMBDA_VALUES, start=2):
            c = ws.cell(r, ci, f"λ = {lam:g}")
            c.fill, c.font, c.alignment = fill, font, center
        r += 1
        for alpha in ALPHA_VALUES:
            ac = ws.cell(r, 1, f"α = {alpha:g}")
            ac.font = bold
            for ci, lam in enumerate(LAMBDA_VALUES, start=2):
                val = _matrix_value(rows, scale, aset, alpha, lam, key)
                cell = ws.cell(r, ci, "" if val is None else val)
                if isinstance(val, float):
                    cell.number_format = "0.0000" if key == "gap_pct" else "#,##0.00"
                cell.alignment = center
            r += 1
        r += 1  # 區塊間空一列
    return r


def export_xlsx(rows: list[dict[str, Any]], xlsx_path: Path) -> None:
    """CSV 為來源真相；Excel 匯出失敗不中斷實驗。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("  [xlsx] openpyxl 未安裝，略過 Excel（pip install openpyxl）")
        return
    try:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)
        fill = PatternFill("solid", fgColor="2E74B5")
        font = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        bold = Font(bold=True)
        styles = (fill, font, center, bold)

        # ---- 9 個 (scale x set) 矩陣分頁 --------------------------------
        for scale, aset in scale_set_order():
            ws = wb.create_sheet(f"{scale}_{aset}"[:31])
            c = scale_counts(scale)
            title = ws.cell(
                1, 1,
                f"size = {scale}  |  set = {aset}  |  scope = {SCOPES[aset]:g}  |  "
                f"S = {BASE_SCENARIOS}, T = {BASE_TIME_PERIODS}  |  "
                f"I={c['I']}, J={c['J']}, H={c['H']}"
            )
            title.font = bold
            _write_matrix(ws, 3, rows, scale, aset, styles)
            ws.column_dimensions["A"].width = 16
            for ci in range(2, len(LAMBDA_VALUES) + 2):
                ws.column_dimensions[chr(ord("A") + ci - 1)].width = 16

        # ---- 最後：ALL_summary 總表（9 組矩陣堆疊，逐組標明 size/set）----
        ws = wb.create_sheet("ALL_summary")
        r = 1
        for scale, aset in scale_set_order():
            label = ws.cell(
                r, 1, f"====== size = {scale}  |  set = {aset}  "
                      f"(scope = {SCOPES[aset]:g}) ======"
            )
            label.font = Font(bold=True, color="C00000")
            r += 1
            r = _write_matrix(ws, r, rows, scale, aset, styles)
            r += 1  # 組間再空一列
        ws.column_dimensions["A"].width = 20
        for ci in range(2, len(LAMBDA_VALUES) + 2):
            ws.column_dimensions[chr(ord("A") + ci - 1)].width = 16

        wb.save(xlsx_path)
    except Exception as exc:  # noqa: BLE001
        print(f"  [xlsx] Excel 匯出失敗（不影響 CSV）: {exc}")


# =============================================================================
# Main
# =============================================================================
def validate_grid() -> None:
    """跑之前先擋掉會保證 FAIL 的設定錯誤（快速失敗、附清楚訊息）。"""
    problems: list[str] = []
    for a in ALPHA_VALUES:
        if not (0.0 <= a < 1.0):
            problems.append(f"alpha={a} 不在 [0,1)（CVaR 要求 alpha<1）")
    for l in LAMBDA_VALUES:
        if not (0.0 <= l <= 1.0):
            problems.append(f"lambda={l} 不在 [0,1]")
    for aset in AMBIGUITY_SETS:
        sc = SCOPES.get(aset)
        if sc is None or sc < 0:
            problems.append(f"{aset} 的 scope 無效: {sc!r}（需 >= 0）")
    if "box" in AMBIGUITY_SETS and SCOPES.get("box", 0.0) > 1.0 / BASE_SCENARIOS + 1e-12:
        problems.append(
            f"box scope {SCOPES['box']} > 1/S={1.0/BASE_SCENARIOS:.4f}；"
            f"等權重下 box 需 scope <= 1/S，否則 worst-case 機率為負、求解會 FAIL"
        )
    if not getattr(cfg, "BENDERS_MULTI_CUT", True):
        problems.append("BENDERS_MULTI_CUT 必須為 True（risk master 需要 multi-cut）")
    if problems:
        raise ValueError("實驗設定無效，請修正後再跑：\n  - " + "\n  - ".join(problems))


def main() -> None:
    validate_grid()

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path  = RESULT_DIR / f"{RESULT_PREFIX}_raw_{timestamp}.csv"
    xlsx_path = RESULT_DIR / f"{RESULT_PREFIX}_{timestamp}.xlsx"

    cases = [(scale, aset, alpha, lam)
             for scale in SCALES
             for aset in AMBIGUITY_SETS
             for alpha in ALPHA_VALUES
             for lam in LAMBDA_VALUES]

    print("=" * 72)
    print("BATCH DRO RISK-PARAMETER EXPERIMENT — scale x set x alpha x lambda")
    print("=" * 72)
    print(f"scales={SCALES}")
    for sc in SCALES:
        c = scale_counts(sc)
        print(f"  {sc:6}: I={c['I']} J={c['J']} H={c['H']}")
    print(f"sets={AMBIGUITY_SETS} scopes={SCOPES}")
    print(f"alpha={ALPHA_VALUES} lambda={LAMBDA_VALUES}")
    print(f"S={BASE_SCENARIOS} T={BASE_TIME_PERIODS} "
          f"time_limit={TIME_LIMIT} mip_gap={MIP_GAP} (跑滿1hr 或 gap5% 即停)")
    print(f"engine=B&BC (all acceleration on)  cases={len(cases)}")
    print(f"CSV   : {csv_path}\nExcel : {xlsx_path}")

    dro_module = load_dro_module()
    rows: list[dict[str, Any]] = []
    original_scale = cfg.EXPERIMENT_SCALE

    try:
        with (suppress_solver_logs(), robust_risk_summary(),
              scaled_instance_generation()):
            for idx, (scale, aset, alpha, lam) in enumerate(cases, start=1):
                cfg.EXPERIMENT_SCALE = scale
                counts = scale_counts(scale)
                row = run_one_case(dro_module, idx, len(cases),
                                   scale, aset, alpha, lam, counts)
                rows.append(row)
                write_results(csv_path, rows)   # 逐 case 重寫（可中斷續看）
                export_xlsx(rows, xlsx_path)     # 逐 case 重寫 Excel（確保有檔）
    finally:
        cfg.EXPERIMENT_SCALE = original_scale

    n_ok = sum(1 for r in rows if r.get("status") == "OK")
    print("\n" + "-" * 72)
    print(f"Done: {n_ok}/{len(rows)} cases OK.")
    print(f"CSV   : {csv_path}\nExcel : {xlsx_path}")


if __name__ == "__main__":
    main()
