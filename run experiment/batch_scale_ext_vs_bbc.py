#!/usr/bin/env python3
"""規模 ablation：Extensive form vs B&BC(全部加速技術) × 災區 70 / 100 / 130。

實驗矩陣
--------
* 規模（3 種）：small = I70、medium = I100、large = I130
                （三者 |J| = 50、|H| = 18 完全相同，只有災區數在變）
* 求解配置（2 種）：
    - Extensive             : 單體模型，無任何加速
    - BBC+WS+RS+UC+Pareto   : B&BC 全部加速技術都加上去
  → 3 × 2 = 6 個 case
* 模型：SP+MCVaR，時限 7200 秒，gap 達 1% 提早結束

summary_table 輸出 **UB / LB / Time / Gap / Nodes**。

B&BC 參數狀態（重要）
--------------------
本檔使用 config.py 中 **commit 142520f「Add Taipei scale-up ablation workflow」
的原始 root seeding 設定**：

    BENDERS_ROOT_SEED_ITERS        = 300
    BENDERS_ROOT_SEED_LB_REL_TOL   = 5e-4
    BENDERS_ROOT_SEED_STALL_ROUNDS = 10
    BENDERS_PARALLEL_ORACLES       = 5

⚠️ 為什麼不要再把 seeding 開大：
   2026-08-16 曾把上述改成 600 / 5e-5 / 40 / 10，結果 **seeding 永遠不會判定
   停滯，把整個 7200 秒全部用光**，B&C 一個節點都沒跑到，於是出現
       UB=58,627,595  LB=-inf  gap=inf%  nodes=0
   的災難結果（正常應該是 UB≈25M、LB≈12M、nodes≈5 萬）。
   根因：lshaped_core 的 seeding 只有「總時限用完」一個時間保護
   （lshaped_core.py L1166-1169），停滯門檻放寬後就失去唯一的煞車。
   本檔在每個 case 結束後會自動偵測這個症狀並發出警告（見 sanity_warnings）。

用法
----
    python "run experiment/batch_scale_ext_vs_bbc.py"

中斷後把最新的 raw CSV 檔名填進 RESUME_FROM_CSV 即可續跑；留空 = 從頭跑。
"""
from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Any


# =============================================================================
# 參數設定區
# =============================================================================

SCALES = ["small", "medium", "large"]      # 災區 70 / 100 / 130（|J|=50、|H|=18 固定）

MODEL_NAME = "SP+MCVaR"

# 情境數。老師指示 S=50；commit 142520f 當時是 30。改這一行即可切換。
BASE_SCENARIOS = 50
BASE_TIME_PERIODS = 8
BASE_SAMPLE_RATIO = 1.0
BASE_DEMAND_MULTIPLIER = 1.0
BASE_ROAD_CAPACITY_MULTIPLIER = 1.0
BASE_HOSPITAL_CAPACITY_MULTIPLIER = 1.0

RISK_ALPHA = 0.9
RISK_LAMBDA = 0.5

TIME_LIMIT = 7200.0    # 每個 case 求解時間上限 = 2 小時
MIP_GAP = 0.01         # 達到 1% 即提早結束

# 子程序硬超時 = 求解時限 + 建模緩衝（Extensive 建模遠慢於 B&BC）
HARD_TIMEOUT_BUFFER_SEC = 1800.0
HARD_TIMEOUT_BUFFER_SEC_EXTENSIVE = 5400.0

# 健檢門檻：seeding 佔用總時限超過這個比例就警告（正常應 < 15%）
ROOT_SEED_TIME_WARN_RATIO = 0.5

COMPUTE_KPIS = False
STOP_ON_ERROR = False

EXCEL_REPLACE_RETRIES = 5
EXCEL_REPLACE_RETRY_SLEEP_SEC = 3.0

RESULT_PREFIX = "Scale_ablation_Ext_vs_BBC"
LOG_SUBDIR_NAME = "scale ext vs bbc"

# 續跑：留空 = 從頭跑全部。
RESUME_FROM_CSV: str = ""


# =============================================================================
# Setup
# =============================================================================

ROOT_DIR = Path(__file__).resolve().parents[1]
LOG_DIR = ROOT_DIR / "logs"
LOG_SUBDIR = LOG_DIR / LOG_SUBDIR_NAME
RESULT_DIR = ROOT_DIR / "experiment result"

LOCAL_PYTHON_PACKAGE_CANDIDATES = [
    ROOT_DIR / ".codex_spreadsheet" / "python_packages",
    Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime"
    / "dependencies" / "python" / "Lib" / "site-packages",
]
if os.environ.get("CODEX_PRIMARY_PYTHON_PACKAGES"):
    LOCAL_PYTHON_PACKAGE_CANDIDATES.insert(
        0, Path(os.environ["CODEX_PRIMARY_PYTHON_PACKAGES"])
    )
for package_dir in reversed(LOCAL_PYTHON_PACKAGE_CANDIDATES):
    if package_dir.exists():
        sys.path.insert(0, str(package_dir))

os.chdir(ROOT_DIR)
for _p in (str(ROOT_DIR / "model core"), str(ROOT_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import config as cfg  # noqa: E402
import logging_utils  # noqa: E402

MCVAR_MODEL_PATH = ROOT_DIR / "model portal" / "mcvar bbc.py"
EXTENSIVE_MODEL_PATH = ROOT_DIR / "model portal" / "extensive_dro.py"

DEFAULT_ROOT_SEED_ITERS = int(cfg.BENDERS_ROOT_SEED_ITERS)
DEFAULT_ROOT_CUT_ROUNDS = int(cfg.BENDERS_ROOT_CUT_ROUNDS)
DEFAULT_MIPFOCUS = int(cfg.BENDERS_MIPFOCUS)
DEFAULT_HEURISTICS = float(cfg.BENDERS_HEURISTICS)
DEFAULT_NUMERIC_FOCUS = int(cfg.BENDERS_NUMERIC_FOCUS)
DEFAULT_BRANCH_PRIORITY_ENABLED = bool(cfg.BENDERS_X_BRANCH_PRIORITY_ENABLED)
DEFAULT_BRANCH_PRIORITY = int(cfg.BENDERS_X_BRANCH_PRIORITY)

CONFIGS = [
    {"name": "Extensive", "engine": "extensive",
     "ev": False, "seed": 0, "rounds": 0, "user": False, "pareto": False},
    {"name": "BBC+WS+RS+UC+Pareto", "engine": "bbc",
     "ev": True, "seed": DEFAULT_ROOT_SEED_ITERS, "rounds": DEFAULT_ROOT_CUT_ROUNDS,
     "user": True, "pareto": True},
]

FIELDNAMES = [
    "scale", "test_id", "config", "I", "J", "H", "S", "T",
    "obj_value", "best_lb", "best_ub", "gap_pct", "cpu_s", "wall_s",
    "nodes", "iterations", "num_vars", "num_constrs",
    "opened_ccps", "idle_open_ccps", "idle_wasted_cost",
    "first_stage_decision", "first_stage_cost",
    "total_cuts", "seed_cuts", "lazy_cuts", "user_cuts",
    "root_seed_lb", "root_seed_iters_done", "root_seed_stop_reason",
    "root_seed_time_s", "root_cut_rounds_done", "oracle_solves",
    "incumbent_evals", "callback_time_s", "parallel_oracles",
    "solver_status", "sanity", "log_path", "solver_log_path", "status", "note",
]

TEXT_KEYS = {
    "scale", "test_id", "config", "first_stage_decision", "solver_status",
    "log_path", "solver_log_path", "status", "note", "root_seed_stop_reason",
    "sanity",
}
INTEGER_KEYS = {
    "I", "J", "H", "S", "T", "nodes", "iterations", "num_vars", "num_constrs",
    "opened_ccps", "idle_open_ccps", "total_cuts", "seed_cuts", "lazy_cuts", "user_cuts",
    "root_seed_iters_done", "root_cut_rounds_done", "oracle_solves",
    "incumbent_evals", "parallel_oracles",
}


def blank_row() -> dict[str, Any]:
    return {key: "NA" for key in FIELDNAMES}


def ensure_excel_dependency() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(
            "需要 openpyxl 才能輸出 Excel。請先執行："
            f" {sys.executable!r} -m pip install openpyxl"
        ) from exc


def ensure_solver_dependency() -> None:
    try:
        import gurobipy  # noqa: F401
    except ImportError as exc:
        raise RuntimeError(f"gurobipy 在 {sys.executable!r} 無法載入。") from exc


def load_portal(path: Path, module_name: str) -> Any:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"無法載入 portal：{path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


def scale_counts(scale: str) -> dict[str, int]:
    p = cfg.SCALE_PROFILES[scale]
    return {"I": int(p["n_disaster"]), "J": int(p["n_ccp"]), "H": int(p["n_hospital"])}


# =============================================================================
# 設定套用
# =============================================================================

@contextmanager
def temporary_config(case: dict[str, Any], scale: str):
    is_extensive = case.get("engine") == "extensive"
    values = {
        "SCENARIOS": BASE_SCENARIOS,
        "TIME_PERIODS": BASE_TIME_PERIODS,
        "SAMPLE_RATIO": BASE_SAMPLE_RATIO,
        "SP_SAMPLE_RATIO": BASE_SAMPLE_RATIO,
        "DEMAND_MULTIPLIER": BASE_DEMAND_MULTIPLIER,
        "ROAD_CAPACITY_MULTIPLIER": BASE_ROAD_CAPACITY_MULTIPLIER,
        "HOSPITAL_CAPACITY_MULTIPLIER": BASE_HOSPITAL_CAPACITY_MULTIPLIER,
        "SP_TIME_LIMIT": TIME_LIMIT,
        "SP_MIP_GAP": MIP_GAP,
        "EXPERIMENT_SCALE": scale,
        "BENDERS_MULTI_CUT": True,
        "BENDERS_EV_WARM_START": case["ev"],
        "BENDERS_ROOT_SEED_ITERS": case["seed"],
        "BENDERS_ROOT_CUT_ROUNDS": case["rounds"],
        "BENDERS_USE_USER_CUTS": case["user"],
        "BENDERS_PARETO_ENABLED": case["pareto"],
        "BENDERS_MIPFOCUS": 0 if is_extensive else DEFAULT_MIPFOCUS,
        "BENDERS_HEURISTICS": 0.0 if is_extensive else DEFAULT_HEURISTICS,
        "BENDERS_NUMERIC_FOCUS": 0 if is_extensive else DEFAULT_NUMERIC_FOCUS,
        "BENDERS_X_BRANCH_PRIORITY_ENABLED": (
            False if is_extensive else DEFAULT_BRANCH_PRIORITY_ENABLED
        ),
        "BENDERS_X_BRANCH_PRIORITY": 0 if is_extensive else DEFAULT_BRANCH_PRIORITY,
    }
    if hasattr(cfg, "CCP_SAMPLE_SIZE"):
        values["CCP_SAMPLE_SIZE"] = None
    original = {key: getattr(cfg, key) for key in values}
    try:
        for key, value in values.items():
            setattr(cfg, key, value)
        yield
    finally:
        for key, value in original.items():
            setattr(cfg, key, value)


@contextmanager
def scaled_instance_generation():
    original = cfg.generate_data
    cache: dict[Any, Any] = {}

    def _scaled(*args, **kwargs):
        kwargs.pop("sample_ratio", None)
        kwargs.pop("ccp_sample_size", None)
        key = (cfg.EXPERIMENT_SCALE, cfg.SCENARIOS)
        if key not in cache:
            cache[key] = original(scale=cfg.EXPERIMENT_SCALE)
        return cache[key]

    cfg.generate_data = _scaled
    try:
        yield
    finally:
        cfg.generate_data = original


# =============================================================================
# 執行單一 case
# =============================================================================

def _safe_case_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._") or "case"


def hard_timeout_for(case: dict[str, Any]) -> float:
    buf = (HARD_TIMEOUT_BUFFER_SEC_EXTENSIVE
           if case.get("engine") == "extensive" else HARD_TIMEOUT_BUFFER_SEC)
    return TIME_LIMIT + buf


def make_test_id(scale: str, case_name: str) -> str:
    return f"{scale}_{case_name}".replace("+", "_plus_")


def _case_paths(scale: str, case: dict[str, Any], run_idx: int, log_run_dir: Path):
    test_id = make_test_id(scale, case["name"])
    stem = f"{run_idx:02d}_{_safe_case_name(test_id)}"
    return (
        test_id,
        log_run_dir / f"{stem}.log",
        log_run_dir / f"{stem}_solver.log",
        log_run_dir / f".{stem}_launcher.tmp.log",
        log_run_dir / f".{stem}_row.tmp.json",
    )


def snapshot_logs() -> set[Path]:
    return {p.resolve() for p in LOG_DIR.glob("*.log")} if LOG_DIR.exists() else set()


def move_newest_log(before: set[Path], destination: Path) -> Path | None:
    created = [p for p in LOG_DIR.glob("*.log") if p.resolve() not in before]
    if not created:
        return None
    source = max(created, key=lambda p: p.stat().st_mtime)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        raise FileExistsError(f"不覆寫既有的 solver log：{destination}")
    shutil.move(str(source), str(destination))
    return destination


def _clean_int(value: Any) -> int:
    """把求解器回傳的整數變數值清成乾淨的非負整數。

    V/U/Y 都是 INTEGER 且 lb=0，但 Gurobi 會回傳 -0.0 或 -1e-11 這種
    數值誤差，直接用 f"{v:.0f}" 會印出 "-0"（看起來像負數，其實就是 0）。
    """
    v = round(float(value))
    return 0 if v == 0 else v      # 消掉 -0


def idle_open_ccps(fs: dict[str, Any] | None) -> list[str]:
    """找出「有開設但完全沒配資源」的 CCP（X=1 但 V=U=ΣY=0）。

    這種 CCP 在模型中**完全沒有作用**：
        SupplyCap: Σ_l Σ_t supply_consumption[l]·REG ≤ Σ_h Y[h,j] = 0
                   ⇒ REG = 0 ⇒ FI = 0（沒有傷患能被送進來）
    卻仍要付 ccp_fixed_opening_cost（150 萬）。
    把 X 改成 0 可以直接省下這筆錢且不違反任何限制式，
    所以只要出現這種 CCP，就代表**這個解一定不是最佳解**（時限內沒清乾淨）。
    數量 × 150 萬 = 目標值至少可以再降低的金額。
    """
    if not fs:
        return []
    supply: dict[str, float] = {}
    for (h, j), value in (fs.get("Y") or {}).items():
        supply[j] = supply.get(j, 0.0) + float(value)
    idle = []
    for j in sorted(fs.get("X") or {}):
        if float(fs["X"][j]) > 0.5:
            if (_clean_int(fs["V"][j]) == 0 and _clean_int(fs["U"][j]) == 0
                    and round(supply.get(j, 0.0), 6) == 0):
                idle.append(str(j))
    return idle


def first_stage_string(fs: dict[str, Any] | None) -> str:
    if not fs:
        return "NA"
    supply: dict[str, float] = {}
    for (h, j), value in fs["Y"].items():
        supply[j] = supply.get(j, 0.0) + float(value)
    idle = set(idle_open_ccps(fs))
    lines = []
    for j in sorted(fs["X"]):
        if float(fs["X"][j]) > 0.5:
            mark = "   <-- 開設但無任何資源（浪費固定成本）" if str(j) in idle else ""
            lines.append(
                f"CCP {j:4s} -> X: 1, Staff(V): {_clean_int(fs['V'][j])}, "
                f"Amb(U): {_clean_int(fs['U'][j])}, "
                f"MedicalSupply(Y): {supply.get(j, 0.0):.2f}{mark}"
            )
    return "\n".join(lines) if lines else "none opened"


def _dispose_model(model: Any) -> None:
    if model is None:
        return
    try:
        model.dispose()
    except Exception:  # noqa: BLE001
        pass


def sanity_warnings(row: dict[str, Any]) -> list[str]:
    """偵測「seeding 吃光時間」這類異常（2026-08-16 事故的自動防呆）。"""
    issues: list[str] = []
    if row.get("status") != "OK":
        return issues

    def num(key):
        v = row.get(key)
        if v in (None, "", "NA"):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    lb, nodes = num("best_lb"), num("nodes")
    gap, seed_t = num("gap_pct"), num("root_seed_time_s")

    if lb is None:
        issues.append("沒有下界（LB=-inf/None）")
    if nodes is not None and nodes == 0 and row.get("config") != "Extensive":
        issues.append("B&C 一個節點都沒跑到（nodes=0）")
    if gap is None:
        issues.append("gap 無法計算（多半是沒有下界）")
    if seed_t is not None and seed_t > ROOT_SEED_TIME_WARN_RATIO * TIME_LIMIT:
        issues.append(
            f"root seeding 吃掉 {seed_t:.0f}s = 總時限的 "
            f"{seed_t / TIME_LIMIT * 100:.0f}%（正常應 < 15%）"
        )
    idle_n = num("idle_open_ccps")
    if idle_n:
        waste = num("idle_wasted_cost") or 0.0
        issues.append(
            f"{idle_n:.0f} 座 CCP 開設但無任何資源（V=U=Y=0，模型中完全無作用），"
            f"白付固定成本 {waste:,.0f} → 此解必非最佳，目標值至少還能再降這麼多"
        )
    return issues


def run_one_case(portal: Any, case: dict[str, Any], scale: str,
                 run_idx: int, total: int, log_run_dir: Path) -> dict[str, Any]:
    _tid, case_log_path, solver_log_path, _l, _r = _case_paths(
        scale, case, run_idx, log_run_dir
    )
    case_log_path.parent.mkdir(parents=True, exist_ok=True)
    with logging_utils.tee_output(case_log_path):
        return _run_one_case_logged(
            portal, case, scale, run_idx, total, case_log_path, solver_log_path
        )


def _run_one_case_logged(portal, case, scale, run_idx, total,
                         case_log_path: Path, solver_log_path: Path) -> dict[str, Any]:
    engine = case.get("engine", "bbc")
    counts = scale_counts(scale)
    row = blank_row()
    row.update({
        "scale": scale, "test_id": make_test_id(scale, case["name"]),
        "config": case["name"], "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "log_path": str(case_log_path), "status": "RUNNING", "note": "",
    })
    print(f"\n[{run_idx}/{total}] scale={scale} config={case['name']}")
    before = snapshot_logs()
    wall_start = time.time()
    model = summary = None
    try:
        try:
            with temporary_config(case, scale):
                model, summary = portal.run_mcvar_model(
                    scenario_size=BASE_SCENARIOS, sample_ratio=BASE_SAMPLE_RATIO,
                    time_limit=TIME_LIMIT, mip_gap=MIP_GAP,
                    alpha=RISK_ALPHA, lam=RISK_LAMBDA, compute_kpis=COMPUTE_KPIS,
                )
        finally:
            row["wall_s"] = f"{time.time() - wall_start:.2f}"
            try:
                moved = move_newest_log(before, solver_log_path)
                row["solver_log_path"] = str(moved) if moved else "NA"
            except Exception as log_exc:  # noqa: BLE001
                row["solver_log_path"] = "NA"
                row["note"] = f"solver log move failed: {type(log_exc).__name__}: {log_exc}"
    except Exception as exc:  # noqa: BLE001
        _dispose_model(model)
        row["status"] = "FAIL"
        prefix = f"{row['note']}; " if row["note"] else ""
        row["note"] = f"{prefix}{type(exc).__name__}: {exc}"
        print(f"  -> FAIL: {row['note']}")
        if STOP_ON_ERROR:
            raise
        return row

    if model is None or summary is None:
        _dispose_model(model)
        row.update({"status": "FAIL", "note": "portal 沒有回傳模型或摘要"})
        return row

    objective = summary.get("objective")
    if objective is None:
        _dispose_model(model)
        row.update({"status": "FAIL", "note": "時限內未找到任何可行解（objective 為 None）"})
        print(f"  -> FAIL: {row['note']}")
        return row

    st = summary.get("bbc_stats", {}) or {}
    fs = summary.get("first_stage")
    opened = [j for j, v in ((fs or {}).get("X") or {}).items() if float(v) > 0.5]
    idle = idle_open_ccps(fs)
    # 每座閒置 CCP 都白付一次固定開設成本 → 目標值至少還能再降這麼多
    try:
        with temporary_config(case, scale):
            _p = cfg.generate_data(scale=scale)["deterministic_parameters"]
        unit_open_cost = float(next(iter(_p["ccp_fixed_opening_cost"].values())))
    except Exception:  # noqa: BLE001
        unit_open_cost = float(cfg.PARAMETERS["ccp_fixed_opening_cost"])
    idle_cost = len(idle) * unit_open_cost
    objective = float(objective)
    best_lb = summary.get("best_lb")
    gap = summary.get("gap_pct")
    row.update({
        "obj_value": f"{objective:.6f}", "best_ub": f"{objective:.6f}",
        "best_lb": "NA" if best_lb is None else f"{float(best_lb):.6f}",
        "gap_pct": "NA" if gap is None else f"{float(gap):.6f}",
        "cpu_s": f"{float(st.get('runtime', float('nan'))):.2f}",
        "nodes": f"{float(getattr(model, 'NodeCount', float('nan'))):.0f}",
        "iterations": f"{float(getattr(model, 'IterCount', float('nan'))):.0f}",
        "num_vars": getattr(model, "NumVars", "NA"),
        "num_constrs": getattr(model, "NumConstrs", "NA"),
        "opened_ccps": len(opened),
        "idle_open_ccps": len(idle),
        "idle_wasted_cost": f"{idle_cost:.2f}",
        "first_stage_decision": first_stage_string(fs),
        "first_stage_cost": summary.get("first_stage_cost", "NA"),
        "total_cuts": st.get("cuts_added", "NA"),
        "seed_cuts": st.get("seed_cuts_added", "NA"),
        "lazy_cuts": st.get("lazy_cuts_added", "NA"),
        "user_cuts": st.get("user_cuts_added", "NA"),
        "root_seed_lb": st.get("root_seed_lb", "NA"),
        "root_seed_iters_done": st.get("root_seed_iters_done", "NA"),
        "root_seed_stop_reason": st.get("root_seed_stop_reason", "NA"),
        "root_seed_time_s": st.get("root_seed_time", "NA"),
        "root_cut_rounds_done": st.get("root_cut_rounds_done", "NA"),
        "oracle_solves": st.get("oracle_solves", "NA"),
        "incumbent_evals": st.get("incumbent_evals", "NA"),
        "callback_time_s": st.get("callback_time", "NA"),
        "parallel_oracles": (0 if engine == "extensive"
                             else st.get("parallel_oracles",
                                         getattr(cfg, "BENDERS_PARALLEL_ORACLES", "NA"))),
        "solver_status": st.get("solver_status", "NA"),
        "status": "OK",
    })
    issues = sanity_warnings(row)
    row["sanity"] = "OK" if not issues else "; ".join(issues)
    _dispose_model(model)
    return row


def _failed_case_row(case, scale, case_log_path: Path, note: str, wall_s: float):
    counts = scale_counts(scale)
    row = blank_row()
    row.update({
        "scale": scale, "test_id": make_test_id(scale, case["name"]),
        "config": case["name"], "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS, "wall_s": f"{wall_s:.2f}",
        "solver_status": "HARD_TIMEOUT" if "hard timeout" in note.lower() else "ERROR",
        "log_path": str(case_log_path), "status": "FAIL", "note": note,
    })
    return row


def _write_child_result(path: Path, row: dict[str, Any]) -> None:
    writing = path.with_suffix(path.suffix + ".writing")
    try:
        with writing.open("w", encoding="utf-8") as f:
            json.dump(row, f, ensure_ascii=False, indent=2)
        os.replace(writing, path)
    finally:
        writing.unlink(missing_ok=True)


def _single_case_cli(argv: list[str]) -> None:
    if len(argv) != 6:
        raise SystemExit(
            "single-case arguments: SCALE CONFIG_INDEX RUN_INDEX TOTAL LOG_DIR RESULT_JSON"
        )
    scale, config_idx_raw, run_idx_raw, total_raw, log_dir_raw, result_raw = argv
    ensure_solver_dependency()
    case = CONFIGS[int(config_idx_raw)]
    log_run_dir = Path(log_dir_raw).resolve()
    result_path = Path(result_raw).resolve()
    portal = load_portal(
        EXTENSIVE_MODEL_PATH if case.get("engine") == "extensive" else MCVAR_MODEL_PATH,
        "scale_ablation_portal_child",
    )
    original_scale = cfg.EXPERIMENT_SCALE
    try:
        cfg.EXPERIMENT_SCALE = scale
        with scaled_instance_generation(), open(os.devnull, "w", encoding="utf-8") as sink:
            with redirect_stdout(sink), redirect_stderr(sink):
                row = run_one_case(portal, case, scale, int(run_idx_raw),
                                   int(total_raw), log_run_dir)
    finally:
        cfg.EXPERIMENT_SCALE = original_scale
    _write_child_result(result_path, row)


def run_one_case_subprocess(case, scale, run_idx, total, log_run_dir: Path):
    _tid, case_log_path, solver_log_path, launcher_path, result_path = _case_paths(
        scale, case, run_idx, log_run_dir
    )
    before = snapshot_logs()
    start = time.time()
    cmd = [
        sys.executable, str(Path(__file__).resolve()), "--single-case",
        scale, str(CONFIGS.index(case)), str(run_idx), str(total),
        str(log_run_dir), str(result_path),
    ]
    timed_out = False
    return_code = None
    hard_timeout = hard_timeout_for(case)
    with launcher_path.open("w", encoding="utf-8", newline="") as launcher:
        try:
            completed = subprocess.run(
                cmd, cwd=ROOT_DIR, stdout=launcher, stderr=subprocess.STDOUT,
                timeout=hard_timeout, check=False,
            )
            return_code = completed.returncode
        except subprocess.TimeoutExpired:
            timed_out = True

    wall_s = time.time() - start
    if not solver_log_path.exists():
        try:
            move_newest_log(before, solver_log_path)
        except Exception:  # noqa: BLE001
            pass

    try:
        if not timed_out and return_code == 0 and result_path.is_file():
            with result_path.open("r", encoding="utf-8") as f:
                return json.load(f)
        if timed_out:
            note = f"Hard timeout: 子程序超過 {hard_timeout:.0f} 秒"
        else:
            note = f"子程序結束碼 {return_code}"
            if return_code in (137, -9):
                note += "（很可能是記憶體不足被系統終止）"
        tail = ""
        if launcher_path.is_file():
            tail = launcher_path.read_text(encoding="utf-8", errors="replace")[-8000:]
        if tail.strip():
            note = f"{note}; launcher output: {tail.strip()}"
        row = _failed_case_row(case, scale, case_log_path, note, wall_s)
        if not case_log_path.exists():
            case_log_path.write_text(note + "\n", encoding="utf-8")
        else:
            with case_log_path.open("a", encoding="utf-8") as f:
                f.write("\n" + note + "\n")
        row["solver_log_path"] = str(solver_log_path) if solver_log_path.exists() else "NA"
        return row
    finally:
        result_path.unlink(missing_ok=True)
        launcher_path.unlink(missing_ok=True)


# =============================================================================
# 輸出
# =============================================================================

def write_results(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}_", suffix=".csv", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        with temp_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
            writer.writeheader()
            writer.writerows(rows)
        for attempt in range(EXCEL_REPLACE_RETRIES):
            try:
                os.replace(temp_path, path)
                return
            except PermissionError:
                if attempt == 0:
                    print(f"[csv] 無法寫入 {path.name}（可能正被 Excel 開啟）；重試中 …")
                time.sleep(EXCEL_REPLACE_RETRY_SLEEP_SEC)
        rescue = path.with_name(f"{path.stem}.rescue_{int(time.time())}.csv")
        shutil.copy2(temp_path, rescue)
        print(f"[csv][WARN] {path.name} 被鎖住，已另存 {rescue.name}")
    finally:
        temp_path.unlink(missing_ok=True)


def _load_prior_rows(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(f"RESUME_FROM_CSV 找不到檔案: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [{k: r.get(k, "NA") for k in FIELDNAMES} for r in csv.DictReader(f)]


def csv_row_count(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return max(0, sum(1 for row in csv.reader(f) if any(row)) - 1)


def excel_value(row: dict[str, Any], key: str) -> Any:
    value = row.get(key, "NA")
    if value in (None, "", "NA"):
        return "NA"
    if key in TEXT_KEYS:
        return str(value)
    try:
        number = float(value)
        return int(number) if key in INTEGER_KEYS and number.is_integer() else number
    except (TypeError, ValueError):
        return value


def _style_header(cell, fill, font, Alignment) -> None:
    cell.fill, cell.font = fill, font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_raw_sheet(wb, rows, fill, font, Alignment) -> None:
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet("raw_results")
    for col, name in enumerate(FIELDNAMES, 1):
        _style_header(ws.cell(1, col, name), fill, font, Alignment)
    for r, row in enumerate(rows, 2):
        for col, key in enumerate(FIELDNAMES, 1):
            cell = ws.cell(r, col, excel_value(row, key))
            if key == "first_stage_decision":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    for col in range(1, len(FIELDNAMES) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _write_settings_sheet(wb, run_id, log_run_dir, fill, font, Alignment) -> None:
    ws = wb.create_sheet("run_settings")
    rows = [
        ("run_id", run_id or "NA"),
        ("log_directory", str(log_run_dir) if log_run_dir else "NA"),
        ("scales", ", ".join(SCALES)),
        ("configs", ", ".join(c["name"] for c in CONFIGS)),
        ("model", MODEL_NAME),
        ("scenario_count", BASE_SCENARIOS),
        ("time_periods", BASE_TIME_PERIODS),
        ("time_limit_s_per_case", TIME_LIMIT),
        ("mip_gap (提早結束門檻)", MIP_GAP),
        ("risk_alpha", RISK_ALPHA),
        ("risk_lambda", RISK_LAMBDA),
        ("data_disaster_csv", cfg.DISASTER_CSV),
        ("data_ccp_csv", cfg.CCP_CSV),
        ("data_hospital_csv", cfg.HOSPITAL_CSV),
        ("ccp_upper_bound_scaling", cfg.CCP_UPPER_BOUND_SCALING),
        ("bbc_root_seed_iters", DEFAULT_ROOT_SEED_ITERS),
        ("bbc_root_seed_lb_rel_tol", float(cfg.BENDERS_ROOT_SEED_LB_REL_TOL)),
        ("bbc_root_seed_stall_rounds", int(cfg.BENDERS_ROOT_SEED_STALL_ROUNDS)),
        ("bbc_root_cut_rounds", DEFAULT_ROOT_CUT_ROUNDS),
        ("bbc_parallel_oracles", int(cfg.BENDERS_PARALLEL_ORACLES)),
        ("hard_timeout_bbc_s", TIME_LIMIT + HARD_TIMEOUT_BUFFER_SEC),
        ("hard_timeout_extensive_s", TIME_LIMIT + HARD_TIMEOUT_BUFFER_SEC_EXTENSIVE),
    ]
    ws.append(["Setting", "Value"])
    for cell in ws[1]:
        _style_header(cell, fill, font, Alignment)
    for item in rows:
        ws.append(item)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 56


def _write_summary_sheet(wb, rows, fill, font, Alignment) -> None:
    """UB / LB / Time / Gap / Nodes，列 = 規模，欄 = 2 個 config。"""
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws = wb.create_sheet("summary_table")

    sub_columns = [("UB", "best_ub"), ("LB", "best_lb"), ("Time(s)", "cpu_s"),
                   ("Gap(%)", "gap_pct"), ("Nodes", "nodes")]
    n_sub = len(sub_columns)
    left = ["Scale", "I", "J", "H"]
    n_left = len(left)
    n_cols = n_left + len(CONFIGS) * n_sub
    last = 2 + len(SCALES)

    for i in range(1, n_left + 1):
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
    for idx in range(len(CONFIGS)):
        start = n_left + 1 + idx * n_sub
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + n_sub - 1)

    for i, title in enumerate(left, 1):
        ws.cell(1, i, title)
    for idx, case in enumerate(CONFIGS):
        start = n_left + 1 + idx * n_sub
        ws.cell(1, start, case["name"])
        for offset, (sub_title, _k) in enumerate(sub_columns):
            ws.cell(2, start + offset, sub_title)

    for k, scale in enumerate(SCALES):
        r = 3 + k
        counts = scale_counts(scale)
        ws.cell(r, 1, scale)
        ws.cell(r, 2, counts["I"])
        ws.cell(r, 3, counts["J"])
        ws.cell(r, 4, counts["H"])
        for idx, case in enumerate(CONFIGS):
            match = next((x for x in rows if x.get("scale") == scale
                          and x.get("config") == case["name"]), None)
            start = n_left + 1 + idx * n_sub
            if match is None:
                values = tuple("NA" for _ in sub_columns)
            elif match.get("status") != "OK":
                values = tuple("FAIL" for _ in sub_columns)
            else:
                values = tuple(excel_value(match, key) for _t, key in sub_columns)
            for offset, value in enumerate(values):
                cell = ws.cell(r, start + offset, value)
                if isinstance(value, float):
                    key = sub_columns[offset][1]
                    cell.number_format = "0.0000" if key == "gap_pct" else "#,##0.00"

    for row in ws.iter_rows(min_row=1, max_row=last, min_col=1, max_col=n_cols):
        for cell in row:
            cell.border = border
            if cell.row <= 2:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 10
    for col in "BCD":
        ws.column_dimensions[col].width = 6
    for col in range(n_left + 1, n_cols + 1):
        key = sub_columns[(col - n_left - 1) % n_sub][1]
        ws.column_dimensions[get_column_letter(col)].width = 16 if key in ("best_ub", "best_lb") else 11
    ws.freeze_panes = f"{get_column_letter(n_left + 1)}3"
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18


def export_xlsx(rows, path: Path, run_id: str = "", log_run_dir: Path | None = None) -> Path:
    ensure_excel_dependency()
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    fill = PatternFill("solid", fgColor="2E74B5")
    font = Font(bold=True, color="FFFFFF")
    _write_raw_sheet(wb, rows, fill, font, Alignment)
    _write_settings_sheet(wb, run_id, log_run_dir, fill, font, Alignment)
    _write_summary_sheet(wb, rows, fill, font, Alignment)

    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}_", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
        last_exc: Exception | None = None
        for attempt in range(EXCEL_REPLACE_RETRIES):
            try:
                os.replace(temp_path, path)
                return path
            except PermissionError as exc:
                last_exc = exc
                if attempt == 0:
                    print(f"[excel] 無法寫入 {path.name}（可能正被 Excel 開啟）；重試中 …")
                time.sleep(EXCEL_REPLACE_RETRY_SLEEP_SEC)
        raise PermissionError(f"無法寫入 {path}，請關閉開啟該檔的 Excel") from last_exc
    finally:
        temp_path.unlink(missing_ok=True)


def export_xlsx_incremental(rows, path, run_id, log_run_dir) -> None:
    try:
        export_xlsx(rows, path, run_id, log_run_dir)
    except Exception as exc:  # noqa: BLE001
        print(f"[excel][WARN] 中途匯出失敗，實驗繼續（結果已存在 CSV）："
              f"{type(exc).__name__}: {exc}")


def expected_test_ids() -> set[str]:
    return {make_test_id(s, c["name"]) for s in SCALES for c in CONFIGS}


def print_console_summary(rows) -> None:
    print("\n--- SUMMARY（UB / LB / Time / Gap / Nodes）---")
    head = (f"{'scale':8}{'config':22}{'UB':>15}{'LB':>15}"
            f"{'Time(s)':>10}{'Gap(%)':>9}{'Nodes':>10}")
    print(head)
    print("-" * len(head))
    for scale in SCALES:
        for case in CONFIGS:
            m = next((x for x in rows if x.get("scale") == scale
                      and x.get("config") == case["name"]), None)
            if m is None:
                continue
            if m.get("status") != "OK":
                print(f"{scale:8}{case['name']:22}{'FAIL':>15}")
                continue
            def fmt(k, d=0):
                v = m.get(k)
                return f"{float(v):,.{d}f}" if v not in (None, "", "NA") else "NA"
            print(f"{scale:8}{case['name']:22}{fmt('best_ub'):>15}{fmt('best_lb'):>15}"
                  f"{fmt('cpu_s', 1):>10}{fmt('gap_pct', 2):>9}{fmt('nodes'):>10}")


# =============================================================================
# main
# =============================================================================

def main() -> None:
    ensure_solver_dependency()
    ensure_excel_dependency()

    # 時間戳只到「秒」，同一秒內啟動兩次會撞名：
    #   log 目錄用 exist_ok=False 會直接拋 FileExistsError 讓程式當場掛掉，
    #   CSV/Excel 也會覆蓋掉前一次的結果。加上後綴確保唯一。
    base_ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    timestamp = base_ts
    suffix = 1
    while (LOG_SUBDIR / timestamp).exists() or \
            (RESULT_DIR / f"{RESULT_PREFIX}_raw_{timestamp}.csv").exists():
        suffix += 1
        timestamp = f"{base_ts}_{suffix}"
    csv_path = RESULT_DIR / f"{RESULT_PREFIX}_raw_{timestamp}.csv"
    xlsx_path = RESULT_DIR / f"{RESULT_PREFIX}_{timestamp}.xlsx"
    log_run_dir = LOG_SUBDIR / timestamp
    log_run_dir.mkdir(parents=True, exist_ok=False)

    n_cases = len(SCALES) * len(CONFIGS)
    print("=" * 78)
    print("規模 ABLATION：Extensive vs B&BC(全開) × 災區 70 / 100 / 130")
    print("=" * 78)
    print(f"資料 : {cfg.DISASTER_CSV} / {cfg.CCP_CSV} / {cfg.HOSPITAL_CSV}")
    for sc in SCALES:
        c = scale_counts(sc)
        print(f"  {sc:7}: I={c['I']:3d} J={c['J']:3d} H={c['H']:3d}")
    print(f"模型 : {MODEL_NAME}   S={BASE_SCENARIOS} T={BASE_TIME_PERIODS}")
    print(f"配置 : {[c['name'] for c in CONFIGS]}")
    print(f"時限 : {TIME_LIMIT:.0f}s/case，gap 達 {MIP_GAP:.0%} 提早結束")
    print(f"cases = {len(SCALES)} scale × {len(CONFIGS)} config = {n_cases}")

    # ── B&BC seeding 參數健檢（2026-08-16 事故的預防）──
    print("\n--- B&BC root seeding 設定 ---")
    print(f"  ITERS={cfg.BENDERS_ROOT_SEED_ITERS}  "
          f"REL_TOL={cfg.BENDERS_ROOT_SEED_LB_REL_TOL:.0e}  "
          f"STALL_ROUNDS={cfg.BENDERS_ROOT_SEED_STALL_ROUNDS}  "
          f"平行 oracle={cfg.BENDERS_PARALLEL_ORACLES}")
    est_per_round = 2.43 * (BASE_SCENARIOS / 30.0) * (5.0 / max(1, int(cfg.BENDERS_PARALLEL_ORACLES)))
    est_seed = cfg.BENDERS_ROOT_SEED_ITERS * est_per_round
    print(f"  推估 seeding 最多 {est_seed:.0f}s = 總時限的 {est_seed / TIME_LIMIT * 100:.0f}%")
    if est_seed > 0.5 * TIME_LIMIT:
        print("  [!!] 警告：seeding 可能吃掉超過一半的求解時間，"
              "會導致 B&C 跑不到節點、LB=-inf。請調小 BENDERS_ROOT_SEED_ITERS。")
    else:
        print("  [OK] seeding 預算合理，B&C 有足夠時間")

    worst_h = sum(hard_timeout_for(c) for c in CONFIGS) * len(SCALES) / 3600.0
    print(f"\n最壞情況總時間 ≈ {worst_h:.0f} 小時")
    print(f"CSV   : {csv_path}\nExcel : {xlsx_path}\nLogs  : {log_run_dir}")

    export_xlsx([], xlsx_path, timestamp, log_run_dir)

    prior_ok: dict[str, dict[str, Any]] = {}
    if not RESUME_FROM_CSV:
        print(f"\n[FRESH START] RESUME_FROM_CSV 為空 → 從頭跑全部 {n_cases} 個 case")
    else:
        prior_path = Path(RESUME_FROM_CSV)
        if not prior_path.is_absolute():
            prior_path = RESULT_DIR / RESUME_FROM_CSV
        for pr in _load_prior_rows(prior_path):
            if pr.get("status") == "OK" and pr.get("test_id"):
                prior_ok[pr["test_id"]] = pr
        print(f"\n[RESUME] 從 {prior_path} 載入 {len(prior_ok)} 個已完成 case")

    rows: list[dict[str, Any]] = []
    run_idx = 0
    try:
        for scale in SCALES:
            for case in CONFIGS:
                run_idx += 1
                test_id = make_test_id(scale, case["name"])
                if test_id in prior_ok:
                    print(f"[{run_idx}/{n_cases}] SKIP（沿用上次已完成）{test_id}")
                    rows.append(prior_ok[test_id])
                else:
                    print(f"\n[{run_idx}/{n_cases}] scale={scale} config={case['name']}")
                    row = run_one_case_subprocess(case, scale, run_idx, n_cases, log_run_dir)
                    rows.append(row)
                    print(f"  -> {row['status']} UB={row['obj_value']} LB={row['best_lb']} "
                          f"time={row['cpu_s']}s gap={row['gap_pct']}% nodes={row['nodes']}")
                    if row.get("sanity") not in (None, "", "NA", "OK"):
                        print(f"  [!!] 結果異常：{row['sanity']}")
                write_results(csv_path, rows)
                export_xlsx_incremental(rows, xlsx_path, timestamp, log_run_dir)
    except KeyboardInterrupt:
        write_results(csv_path, rows)
        export_xlsx_incremental(rows, xlsx_path, timestamp, log_run_dir)
        print("\n" + "=" * 78)
        print("[INTERRUPTED] 已中斷，完成的結果都已保存。")
        print(f"  已完成 OK: {sum(r['status'] == 'OK' for r in rows)}/{n_cases}")
        print("\n續跑方式：把下面這行貼回本檔參數區，再執行一次。")
        print(f'    RESUME_FROM_CSV = "{csv_path.name}"')
        print("=" * 78)
        return

    export_xlsx(rows, xlsx_path, timestamp, log_run_dir)
    n_ok = sum(r["status"] == "OK" for r in rows)
    print("\n" + "=" * 78)
    print(f"Done: {n_ok}/{len(rows)} cases OK")
    print_console_summary(rows)

    bad = [r for r in rows if r.get("sanity") not in (None, "", "NA", "OK")]
    if bad:
        print("\n[!!] 以下 case 的結果不正常，請先看 log 再用這些數字：")
        for r in bad:
            print(f"  - {r['test_id']}: {r['sanity']}")

    print(f"\nCSV   : {csv_path}\nExcel : {xlsx_path}\nLogs  : {log_run_dir}")
    if n_ok < n_cases:
        print("\n尚有 case 未完成，續跑請把下面這行貼回參數區：")
        print(f'    RESUME_FROM_CSV = "{csv_path.name}"')


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--single-case":
        _single_case_cli(sys.argv[2:])
    else:
        main()
