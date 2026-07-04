#!/usr/bin/env python3
"""
Batch SP experiment runner.
(Phase R 重構：原 run experiment.py，僅改路徑——ROOT_DIR 改為專案根、
 SP 模型路徑指向 model portal/extensive form.py、結果 CSV 寫入 experiment result/，邏輯零改動)

This script is only a runner. It temporarily changes values in the imported
config module while it runs each experiment case, then restores them. It does
not rewrite config.py and does not change the model core logic.

Experiment modes (set EXPERIMENT_AXIS):
  "ablation"                     -> run 4 ablation cases (USE_SCENARIO_OMEGA x USE_SPATIAL_KMEANS)
  "scenario"                     -> run different scenario counts, e.g. [5, 10, 20]
  "ccp"                          -> run different CCP counts, e.g. [6, 8, 10]
  "sample_ratio"                 -> run different I/H sample ratios, e.g. [0.1, 0.25, 0.5, 1.0]
  "time_period"                  -> run different time periods, e.g. [4, 8, 12, 16]
  "demand_multiplier"            -> run different demand multipliers, e.g. [1.0, 1.2, 1.5]
  "road_capacity_multiplier"     -> run different road capacity multipliers
  "hospital_capacity_multiplier" -> run different hospital capacity multipliers
"""

from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import math
import os
import re
import shutil
import sys
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Any


# =============================================================================
# Parameter setting area
# =============================================================================

# ── Experiment axis ──────────────────────────────────────────────────────────
# 情境規模實驗（Phase 6）：固定基礎設定（全局 omega [0.8,1.2]、空間 U[0.5,1.5] 正規化、
# kmeans+omega 皆開、T=8），只掃情境數，比較求解速度；Final Gap > GAP_STOP_PCT 即早停
EXPERIMENT_AXIS   = "scenario"
EXPERIMENT_VALUES = [5, 20, 30, 50, 100, 150]

# ── Ablation cases (only used when EXPERIMENT_AXIS == "ablation") ────────────
ABLATION_CASES = [
    {"label": "A_no_omega_no_kmeans", "use_omega": False, "use_kmeans": False},
    {"label": "B_omega_only",          "use_omega": True,  "use_kmeans": False},
    {"label": "C_kmeans_only",         "use_omega": False, "use_kmeans": True},
    {"label": "D_both",                "use_omega": True,  "use_kmeans": True},
]

# ── Fixed base settings ───────────────────────────────────────────────────────
BASE_SCENARIOS              = 5       # fallback when EXPERIMENT_AXIS != "scenario"
BASE_CCP_SAMPLE_SIZE        = None    # None = use all CCPs (10 個)
BASE_SAMPLE_RATIO           = 1.0     # 全部災害點與醫院
BASE_TIME_PERIODS           = 8       # 固定 8 個 time period
BASE_DEMAND_MULTIPLIER      = 1.0
BASE_ROAD_CAPACITY_MULTIPLIER    = 1.0
BASE_HOSPITAL_CAPACITY_MULTIPLIER = 1.0

# ── Solver settings ───────────────────────────────────────────────────────────
TIME_LIMIT   = 3600.0
MIP_GAP      = 0.01
GAP_STOP_PCT = 10.0    # gap 超過此值 → 標記 STOP，並停止後續實驗

# ── Output settings ───────────────────────────────────────────────────────────
RESULT_PREFIX = "east_district_stress_test"
STOP_ON_ERROR = True   # STOP 後不繼續跑後面的實驗

# ── Log subfolder ─────────────────────────────────────────────────────────────
# 每個實驗跑完後，新生成的 log 自動移入此子資料夾
LOG_SUBDIR_NAME = "east district stress test"


# =============================================================================
# Setup
# =============================================================================
ROOT_DIR      = Path(__file__).resolve().parents[1]   # 專案根（Phase R：原為本檔所在目錄）
LOG_DIR       = ROOT_DIR / "logs"
LOG_SUBDIR    = LOG_DIR / LOG_SUBDIR_NAME
RESULT_DIR    = ROOT_DIR / "experiment result"

os.chdir(ROOT_DIR)
for _p in (str(ROOT_DIR / "model core"), str(ROOT_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import config as cfg  # noqa: E402

# ── 求解引擎分派（Phase 5）：config.SOLVER_ENGINE = "extensive" | "lshaped" ──
_ENGINE_FILE = {
    "extensive": "extensive form.py",
    "lshaped":   "benders bbc.py",
}
SOLVER_ENGINE = getattr(cfg, "SOLVER_ENGINE", "lshaped")
SP_MODEL_PATH = ROOT_DIR / "model portal" / _ENGINE_FILE.get(SOLVER_ENGINE, "extensive form.py")


FIELDNAMES = [
    "test_id",
    "axis",
    "axis_value",
    "factor",
    "use_omega",
    "use_kmeans",
    "sample_ratio",
    "I", "J", "H", "S", "T",
    "obj_value",
    "first_stage_decision",
    "best_lb",
    "best_ub",
    "cpu_s",
    "wall_s",
    "num_vars",
    "num_constrs",
    "nodes",
    "iterations",
    "gap_pct",
    "vss_pct",
    "evpi_pct",
    "demand_multiplier",
    "road_capacity_multiplier",
    "hospital_capacity_multiplier",
    # ---- B&BC 引擎統計（extensive 引擎時為 NA）----
    "engine",
    "total_cuts",
    "seed_cuts",
    "lazy_cuts",
    "user_cuts",
    "root_seed_iters",
    "root_seed_iters_done",
    "root_seed_lb",
    "root_seed_stop_reason",
    "root_seed_time_s",
    "root_cut_rounds",
    "root_cut_rounds_done",
    "parallel_oracles",
    "cache_hits",
    "cache_misses",
    "oracle_solves",
    "callback_time_s",
    "log_path",
    "status",
    "note",
]

# 論文表格用的 factor 標籤（比照 Excel 表列名：Scenario x5、Time Period x8 ...）
AXIS_FACTOR_PREFIX = {
    "scenario":                     "Scenario x",
    "time_period":                  "Time Period x",
    "ccp":                          "CCP x",
    "sample_ratio":                 "Sample ratio ",
    "demand_multiplier":            "Demand mult. ",
    "road_capacity_multiplier":     "Road cap. mult. ",
    "hospital_capacity_multiplier": "Hosp cap. mult. ",
}

AXIS_TO_SETTING = {
    "scenario":                     "scenarios",
    "ccp":                          "ccp_sample_size",
    "sample_ratio":                 "sample_ratio",
    "time_period":                  "time_periods",
    "demand_multiplier":            "demand_multiplier",
    "road_capacity_multiplier":     "road_capacity_multiplier",
    "hospital_capacity_multiplier": "hospital_capacity_multiplier",
}


# =============================================================================
# Helpers
# =============================================================================
def load_sp_module():
    if not SP_MODEL_PATH.exists():
        raise FileNotFoundError(f"Cannot find SP model file: {SP_MODEL_PATH}")
    spec = importlib.util.spec_from_file_location("sp_model_runner", SP_MODEL_PATH)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load module from {SP_MODEL_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def csv_row_count(filepath: Path) -> int:
    if not filepath.exists():
        return 0
    with filepath.open(encoding="utf-8-sig", newline="") as f:
        return max(0, sum(1 for _ in f) - 1)


def fmt_value(value: Any) -> str:
    if value is None:
        return "ALL"
    if isinstance(value, dict):
        return value.get("label", str(value))
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def blank_row() -> dict[str, Any]:
    return {key: "NA" for key in FIELDNAMES}


def write_results(csv_path: Path, rows: list[dict[str, Any]]) -> None:
    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)


def extract_first_stage(model: Any) -> str:
    """從回傳的模型（extensive form 完整模型或 B&BC master）讀一階決策字串。"""
    if model is None:
        return "NA"
    xs: dict[str, float] = {}
    vs: dict[str, float] = {}
    us: dict[str, float] = {}
    ys: dict[str, float] = {}
    try:
        for var in model.getVars():
            name = var.VarName
            if name.startswith("X["):
                xs[name[2:-1]] = var.X
            elif name.startswith("V["):
                vs[name[2:-1]] = var.X
            elif name.startswith("U["):
                us[name[2:-1]] = var.X
            elif name.startswith("Y["):
                h, j = name[2:-1].split(",")
                ys[j] = ys.get(j, 0.0) + var.X
            else:
                break   # 一階變數在最前面，遇到其他變數即可停
    except Exception:
        return "NA"
    lines = []
    for j in sorted(xs):
        if xs[j] > 0.5:
            lines.append(
                f"CCP {j:4s} -> X: 1, Staff(V): {vs.get(j, 0):2.0f}, "
                f"Amb(U): {us.get(j, 0):2.0f}, MedicalSupply(Y): {ys.get(j, 0.0):.2f}"
            )
    return "\n".join(lines) if lines else "none opened"


# Excel 欄位：18 欄論文表格 + 精選 B&BC 統計（extensive 引擎時為 NA）
XLSX_COLUMNS = [
    ("factor",               "factor"),
    ("|I| Disaster",         "I"),
    ("|J| CCP",              "J"),
    ("|H| Hosp",             "H"),
    ("|S| Scen",             "S"),
    ("|T| Per",              "T"),
    ("obj_value",            "obj_value"),
    ("First Stage Decision", "first_stage_decision"),
    ("Best LB",              "best_lb"),
    ("Best UB",              "best_ub"),
    ("CPU Time(s)",          "cpu_s"),
    ("num_vars",             "num_vars"),
    ("num_constrs",          "num_constrs"),
    ("Nodes",                "nodes"),
    ("Iteration",            "iterations"),
    ("Final Gap(%)",         "gap_pct"),
    ("VSS(%)",               "vss_pct"),
    ("EVPI(%)",              "evpi_pct"),
    ("Engine",               "engine"),
    ("Total Cuts",           "total_cuts"),
    ("Seed Cuts",            "seed_cuts"),
    ("Lazy Cuts",            "lazy_cuts"),
    ("User Cuts",            "user_cuts"),
    ("Root Seed Iters",      "root_seed_iters_done"),
    ("Seeded LB",            "root_seed_lb"),
    ("Root Seed Stop",       "root_seed_stop_reason"),
    ("Root Seed Time(s)",    "root_seed_time_s"),
    ("Root Cut Rounds",      "root_cut_rounds_done"),
    ("Parallel Oracles",     "parallel_oracles"),
    ("Cache Hits",           "cache_hits"),
    ("Cache Misses",         "cache_misses"),
    ("Oracle Solves",        "oracle_solves"),
    ("Callback Time(s)",     "callback_time_s"),
]

_XLSX_INT_KEYS = ("I", "J", "H", "S", "T", "num_vars", "num_constrs",
                  "nodes", "iterations", "total_cuts", "lazy_cuts",
                  "seed_cuts", "user_cuts", "root_seed_iters",
                  "root_seed_iters_done", "root_cut_rounds",
                  "root_cut_rounds_done", "parallel_oracles",
                  "cache_hits", "cache_misses", "oracle_solves")


def export_xlsx(rows: list[dict[str, Any]], xlsx_path: Path) -> None:
    """匯出結果至 xlsx（experiment result/）。CSV 為來源真相；
    匯出失敗（openpyxl 未裝、檔案被 Excel 開啟）不會中斷實驗。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("  [xlsx] openpyxl 未安裝，略過 Excel 匯出（pip install openpyxl）")
        return
    try:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        header_fill = PatternFill("solid", fgColor="2E74B5")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, (title, _key) in enumerate(XLSX_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r_idx, row in enumerate(rows, start=2):
            for col_idx, (_title, key) in enumerate(XLSX_COLUMNS, start=1):
                value = row.get(key, "NA")
                if key not in ("factor", "first_stage_decision", "engine"):
                    try:
                        value = float(value)
                        if value == int(value) and key in _XLSX_INT_KEYS:
                            value = int(value)
                    except (TypeError, ValueError):
                        pass
                cell = ws.cell(row=r_idx, column=col_idx, value=value)
                if key == "first_stage_decision":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["H"].width = 62   # First Stage Decision
        wb.save(xlsx_path)
    except Exception as exc:
        print(f"  [xlsx] Excel 匯出失敗（不影響 CSV）: {exc}")


def snapshot_logs() -> set[Path]:
    if not LOG_DIR.exists():
        return set()
    return {p.resolve() for p in LOG_DIR.glob("*.log")}


def newest_created_log(before: set[Path]) -> Path | None:
    """回傳 LOG_DIR（不含子資料夾）裡新出現的 log；None 表示沒有新檔案。"""
    if not LOG_DIR.exists():
        return None
    created = [p for p in LOG_DIR.glob("*.log") if p.resolve() not in before]
    if not created:
        return None
    return max(created, key=lambda p: p.stat().st_mtime)


def move_log_to_subdir(log_path: Path | None) -> Path | None:
    """把剛產生的 log 移入 LOG_SUBDIR，回傳新路徑。"""
    if log_path is None or not log_path.exists():
        return log_path
    LOG_SUBDIR.mkdir(parents=True, exist_ok=True)
    dest = LOG_SUBDIR / log_path.name
    shutil.move(str(log_path), str(dest))
    return dest


def parse_log_summary(log_path: Path | None) -> dict[str, float]:
    if log_path is None or not log_path.exists():
        return {}
    patterns = {
        "obj_value": re.compile(r"Objective Value:\s*([-+]?\d+(?:\.\d+)?)"),
        "cpu_s":     re.compile(r"CPU Time:\s*([-+]?\d+(?:\.\d+)?)\s*s"),
        "best_ub":   re.compile(r"Best UB \(Objective\):\s*([-+]?\d+(?:\.\d+)?)"),
        "best_lb":   re.compile(r"Best LB \(Bound\):\s*([-+]?\d+(?:\.\d+)?)"),
        "gap_pct":   re.compile(r"Final Gap:\s*([-+]?\d+(?:\.\d+)?)\s*%"),
        "vss_pct":   re.compile(r"VSS\(%\)\s*=\s*([-+]?\d+(?:\.\d+)?)"),
        "evpi_pct":  re.compile(r"EVPI\(%\)\s*=\s*([-+]?\d+(?:\.\d+)?)"),
    }
    values: dict[str, float] = {}
    text = log_path.read_text(encoding="utf-8", errors="replace")
    for key, pat in patterns.items():
        matches = pat.findall(text)
        if matches:
            values[key] = float(matches[-1])
    return values


def safe_model_attr(model: Any, attr: str) -> float | None:
    try:
        return float(getattr(model, attr))
    except Exception:
        return None


def model_fallback_summary(model: Any) -> dict[str, float]:
    if model is None:
        return {}
    values: dict[str, float] = {}
    for attr, key in [("Runtime", "cpu_s"), ("ObjVal", "obj_value"), ("ObjBound", "best_lb"), ("MIPGap", None)]:
        v = safe_model_attr(model, attr)
        if v is not None:
            if attr == "MIPGap":
                values["gap_pct"] = v * 100.0
            elif attr == "ObjVal":
                values["obj_value"] = v
                values["best_ub"]   = v
            else:
                values[key] = v
    return values


def format_float(value: Any, digits: int = 4) -> str:
    if value is None or value == "NA":
        return "NA"
    try:
        return f"{float(value):.{digits}f}"
    except Exception:
        return "NA"


def estimate_counts(settings: dict[str, Any]) -> dict[str, int]:
    full_i = csv_row_count(ROOT_DIR / "data" / cfg.DISASTER_CSV)
    full_j = csv_row_count(ROOT_DIR / "data" / cfg.CCP_CSV)
    full_h = csv_row_count(ROOT_DIR / "data" / cfg.HOSPITAL_CSV)
    ratio    = float(settings["sample_ratio"])
    ccp_size = settings["ccp_sample_size"]
    return {
        "I": max(1, math.ceil(full_i * ratio)) if ratio < 1.0 else full_i,
        "J": min(ccp_size, full_j) if ccp_size is not None else full_j,
        "H": max(1, math.ceil(full_h * ratio)) if ratio < 1.0 else full_h,
    }


# =============================================================================
# Settings builders
# =============================================================================
def base_case_settings() -> dict[str, Any]:
    return {
        "scenarios":                    BASE_SCENARIOS,
        "ccp_sample_size":              BASE_CCP_SAMPLE_SIZE,
        "sample_ratio":                 BASE_SAMPLE_RATIO,
        "time_periods":                 BASE_TIME_PERIODS,
        "demand_multiplier":            BASE_DEMAND_MULTIPLIER,
        "road_capacity_multiplier":     BASE_ROAD_CAPACITY_MULTIPLIER,
        "hospital_capacity_multiplier": BASE_HOSPITAL_CAPACITY_MULTIPLIER,
        "use_omega":                    None,
        "use_kmeans":                   None,
    }


def build_case(axis_value: Any) -> dict[str, Any]:
    settings = base_case_settings()
    if EXPERIMENT_AXIS == "ablation":
        settings["use_omega"]  = axis_value["use_omega"]
        settings["use_kmeans"] = axis_value["use_kmeans"]
        return settings
    if EXPERIMENT_AXIS not in AXIS_TO_SETTING:
        valid = ", ".join(sorted(AXIS_TO_SETTING))
        raise ValueError(f"Unknown EXPERIMENT_AXIS={EXPERIMENT_AXIS!r}; valid: {valid}")
    settings[AXIS_TO_SETTING[EXPERIMENT_AXIS]] = axis_value
    return settings


def validate_case(settings: dict[str, Any]) -> None:
    if not isinstance(settings["scenarios"], int) or settings["scenarios"] < 1:
        raise ValueError("scenarios must be a positive integer")
    if not isinstance(settings["time_periods"], int) or settings["time_periods"] < 1:
        raise ValueError("time_periods must be a positive integer")
    if not (0 < float(settings["sample_ratio"]) <= 1.0):
        raise ValueError("sample_ratio must be in (0, 1.0]")
    ccp_size = settings["ccp_sample_size"]
    if ccp_size is not None and (not isinstance(ccp_size, int) or ccp_size < 1):
        raise ValueError("ccp_sample_size must be None or a positive integer")


# =============================================================================
# Context managers (temporarily patch config, restore on exit)
# =============================================================================
@contextmanager
def temporary_config(settings: dict[str, Any]):
    keys: dict[str, Any] = {
        "SCENARIOS":                    settings["scenarios"],
        "TIME_PERIODS":                 settings["time_periods"],
        "SAMPLE_RATIO":                 settings["sample_ratio"],
        "SP_SAMPLE_RATIO":              settings["sample_ratio"],
        "DEMAND_MULTIPLIER":            settings["demand_multiplier"],
        "ROAD_CAPACITY_MULTIPLIER":     settings["road_capacity_multiplier"],
        "HOSPITAL_CAPACITY_MULTIPLIER": settings["hospital_capacity_multiplier"],
        "SP_TIME_LIMIT":                TIME_LIMIT,
        "SP_MIP_GAP":                   MIP_GAP,
    }
    if hasattr(cfg, "CCP_SAMPLE_SIZE"):
        keys["CCP_SAMPLE_SIZE"] = settings["ccp_sample_size"]
    if settings.get("use_omega") is not None and hasattr(cfg, "USE_SCENARIO_OMEGA"):
        keys["USE_SCENARIO_OMEGA"] = settings["use_omega"]
    if settings.get("use_kmeans") is not None and hasattr(cfg, "USE_SPATIAL_KMEANS"):
        keys["USE_SPATIAL_KMEANS"] = settings["use_kmeans"]

    original = {key: getattr(cfg, key) for key in keys}
    try:
        for key, value in keys.items():
            setattr(cfg, key, value)
        yield
    finally:
        for key, value in original.items():
            setattr(cfg, key, value)


@contextmanager
def patched_generate_data(ccp_sample_size: int | None):
    original = cfg.generate_data
    def _patched(*args, **kwargs):
        kwargs["ccp_sample_size"] = ccp_sample_size
        return original(*args, **kwargs)
    cfg.generate_data = _patched
    try:
        yield
    finally:
        cfg.generate_data = original


@contextmanager
def patched_generate_scenarios(settings: dict[str, Any]):
    original = cfg.generate_scenarios
    def _patched(*args, **kwargs):
        kwargs.setdefault("demand_multiplier",            settings["demand_multiplier"])
        kwargs.setdefault("road_capacity_multiplier",     settings["road_capacity_multiplier"])
        kwargs.setdefault("hospital_capacity_multiplier", settings["hospital_capacity_multiplier"])
        kwargs.setdefault("num_periods",                  settings["time_periods"])
        return original(*args, **kwargs)
    cfg.generate_scenarios = _patched
    try:
        yield
    finally:
        cfg.generate_scenarios = original


# =============================================================================
# Core run logic
# =============================================================================
def run_one_case(
    sp_module: Any,
    run_idx: int,
    total_runs: int,
    axis_value: Any,
) -> dict[str, Any]:
    settings = build_case(axis_value)
    validate_case(settings)
    counts   = estimate_counts(settings)

    if EXPERIMENT_AXIS == "ablation":
        label   = axis_value["label"]
        test_id = f"ablation_{label}"
    else:
        label   = fmt_value(axis_value)
        test_id = f"{EXPERIMENT_AXIS}_{label}".replace(".", "p")

    factor = AXIS_FACTOR_PREFIX.get(EXPERIMENT_AXIS, f"{EXPERIMENT_AXIS} ") + str(label)
    if EXPERIMENT_AXIS == "ablation":
        factor = f"Ablation {label}"

    row = blank_row()
    row.update({
        "test_id":                      test_id,
        "axis":                         EXPERIMENT_AXIS,
        "axis_value":                   label,
        "factor":                       factor,
        "use_omega":                    str(getattr(cfg, "USE_SCENARIO_OMEGA", "NA")),
        "use_kmeans":                   str(getattr(cfg, "USE_SPATIAL_KMEANS", "NA")),
        "sample_ratio":                 fmt_value(settings["sample_ratio"]),
        "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": settings["scenarios"],
        "T": settings["time_periods"],
        "demand_multiplier":            fmt_value(settings["demand_multiplier"]),
        "road_capacity_multiplier":     fmt_value(settings["road_capacity_multiplier"]),
        "hospital_capacity_multiplier": fmt_value(settings["hospital_capacity_multiplier"]),
        "status": "RUNNING",
        "note": "",
    })

    print(
        f"\n[{run_idx}/{total_runs}] {test_id}"
        f" | I={counts['I']}, J={counts['J']}, H={counts['H']}"
        f", S={settings['scenarios']}, T={settings['time_periods']}"
    )

    logs_before = snapshot_logs()
    wall_start  = time.time()
    model = None
    summary = None
    log_path = None

    try:
        with (
            temporary_config(settings),
            patched_generate_data(settings["ccp_sample_size"]),
            patched_generate_scenarios(settings),
        ):
            model, summary = sp_module.run_sp_model(
                scenario_size=settings["scenarios"],
                sample_ratio=settings["sample_ratio"],
                time_limit=TIME_LIMIT,
                mip_gap=MIP_GAP,
            )
    finally:
        raw_log = newest_created_log(logs_before)
        log_path = move_log_to_subdir(raw_log)   # 移入 east district stress test 子資料夾

    wall_s = time.time() - wall_start
    row["wall_s"]   = format_float(wall_s, digits=1)
    row["log_path"] = str(log_path) if log_path is not None else "NA"

    if model is None:
        row["status"] = "STOP"
        row["note"]   = "INFEASIBLE / no solution / no model returned"
        return row

    parsed = model_fallback_summary(model)
    parsed.update(parse_log_summary(log_path))

    row["cpu_s"]     = format_float(parsed.get("cpu_s"),    digits=2)
    row["gap_pct"]   = format_float(parsed.get("gap_pct"),  digits=4)
    row["obj_value"] = format_float(parsed.get("obj_value"), digits=2)
    row["best_ub"]   = format_float(parsed.get("best_ub"),  digits=2)
    row["best_lb"]   = format_float(parsed.get("best_lb"),  digits=2)

    # 模型維度與樹統計（extensive = 完整模型；lshaped = master，論文表格需註明）
    row["num_vars"]    = int(safe_model_attr(model, "NumVars") or 0)
    row["num_constrs"] = int(safe_model_attr(model, "NumConstrs") or 0)
    row["nodes"]       = int(safe_model_attr(model, "NodeCount") or 0)
    row["iterations"]  = int(safe_model_attr(model, "IterCount") or 0)
    row["first_stage_decision"] = extract_first_stage(model)

    if summary is not None:
        row["vss_pct"]  = format_float(summary.get("VSS_pct"),  digits=4)
        row["evpi_pct"] = format_float(summary.get("EVPI_pct"), digits=4)
        bbc = summary.get("bbc_stats")
        if bbc:
            row.update({
                "engine":                bbc.get("engine", "bbc"),
                "total_cuts":            bbc.get("cuts_added", "NA"),
                "seed_cuts":             bbc.get("seed_cuts_added", "NA"),
                "lazy_cuts":             bbc.get("lazy_cuts_added", "NA"),
                "user_cuts":             bbc.get("user_cuts_added", "NA"),
                "root_seed_iters":       bbc.get("root_seed_iters", "NA"),
                "root_seed_iters_done":  bbc.get("root_seed_iters_done", "NA"),
                "root_seed_lb":          format_float(bbc.get("root_seed_lb"), digits=2),
                "root_seed_stop_reason": bbc.get("root_seed_stop_reason", "NA"),
                "root_seed_time_s":      format_float(bbc.get("root_seed_time"), digits=2),
                "root_cut_rounds":       bbc.get("root_cut_rounds", "NA"),
                "root_cut_rounds_done":  bbc.get("root_cut_rounds_done", "NA"),
                "parallel_oracles":      bbc.get("parallel_oracles", "NA"),
                "cache_hits":            bbc.get("cache_hits", "NA"),
                "cache_misses":          bbc.get("cache_misses", "NA"),
                "oracle_solves":         bbc.get("oracle_solves", "NA"),
                "callback_time_s":       format_float(bbc.get("callback_time"), digits=2),
            })
            row["cpu_s"] = format_float(bbc.get("runtime"), digits=2)
        else:
            row["engine"] = "extensive"
    # log 裡也有 VSS/EVPI，parse_log_summary 已經抓進 parsed
    if "vss_pct" in parsed and row["vss_pct"] == "NA":
        row["vss_pct"]  = format_float(parsed["vss_pct"],  digits=4)
    if "evpi_pct" in parsed and row["evpi_pct"] == "NA":
        row["evpi_pct"] = format_float(parsed["evpi_pct"], digits=4)

    gap_pct = parsed.get("gap_pct")
    stops   = []
    if gap_pct is not None and gap_pct > GAP_STOP_PCT:
        stops.append(f"Gap {gap_pct:.2f}% > {GAP_STOP_PCT:.2f}% — 停止後續實驗")

    row["status"] = "STOP" if stops else "OK"
    row["note"]   = "; ".join(stops)

    print(
        f"  status={row['status']} | CPU={row['cpu_s']}s | "
        f"Gap={row['gap_pct']}% | VSS={row['vss_pct']}% | EVPI={row['evpi_pct']}%"
    )
    if row["note"]:
        print(f"  *** {row['note']}")
    return row


# =============================================================================
# Print helpers
# =============================================================================
def print_header(csv_path: Path) -> None:
    cases = ABLATION_CASES if EXPERIMENT_AXIS == "ablation" else EXPERIMENT_VALUES
    print("=" * 80)
    print("SP BATCH EXPERIMENT RUNNER  —  East District Stress Test")
    print(f"Start  : {dt.datetime.now():%Y-%m-%d %H:%M:%S}")
    print(f"Mode   : {EXPERIMENT_AXIS}  |  Values: {cases}")
    print(
        f"Base   : T={BASE_TIME_PERIODS}, sample={BASE_SAMPLE_RATIO}"
        f", CCP={fmt_value(BASE_CCP_SAMPLE_SIZE)}"
    )
    print(f"Omega  : USE_SCENARIO_OMEGA={getattr(cfg,'USE_SCENARIO_OMEGA','?')}"
          f"  USE_SPATIAL_KMEANS={getattr(cfg,'USE_SPATIAL_KMEANS','?')}")
    print(f"Solver : time_limit={TIME_LIMIT:.0f}s  mip_gap={MIP_GAP:g}"
          f"  stop_if_gap>{GAP_STOP_PCT}%")
    print(f"Engine : {SOLVER_ENGINE}  ({SP_MODEL_PATH.name})")
    print(f"Log dir: {LOG_SUBDIR}")
    print(f"Output : {csv_path}")
    print("=" * 80)


def print_summary(rows: list[dict[str, Any]], csv_path: Path) -> None:
    print("\n" + "=" * 80)
    print("SUMMARY")
    hdr = f"{'ID':>20} | {'S':>4} | {'CPU(s)':>9} | {'Gap%':>8} | {'VSS%':>8} | {'EVPI%':>8} | Status"
    print(hdr)
    print("-" * len(hdr))
    for row in rows:
        print(
            f"{row['test_id']:>20} | {str(row['S']):>4} | "
            f"{str(row['cpu_s']):>9} | {str(row['gap_pct']):>8} | "
            f"{str(row['vss_pct']):>8} | {str(row['evpi_pct']):>8} | {row['status']}"
        )
    print("=" * 80)
    print(f"Results: {csv_path}")
    print(f"End    : {dt.datetime.now():%Y-%m-%d %H:%M:%S}")


# =============================================================================
# Main
# =============================================================================
def main() -> None:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path  = RESULT_DIR / f"{RESULT_PREFIX}_{EXPERIMENT_AXIS}_{timestamp}.csv"
    xlsx_path = RESULT_DIR / f"{RESULT_PREFIX}_{EXPERIMENT_AXIS}_{timestamp}.xlsx"
    rows: list[dict[str, Any]] = []

    LOG_SUBDIR.mkdir(parents=True, exist_ok=True)

    sp_module = load_sp_module()
    print_header(csv_path)
    print(f"Excel  : {xlsx_path}")

    cases = ABLATION_CASES if EXPERIMENT_AXIS == "ablation" else EXPERIMENT_VALUES
    for run_idx, axis_value in enumerate(cases, start=1):
        try:
            row = run_one_case(sp_module, run_idx, len(cases), axis_value)
        except Exception as exc:
            label = axis_value.get("label", fmt_value(axis_value)) if isinstance(axis_value, dict) else fmt_value(axis_value)
            row = blank_row()
            row.update({
                "test_id":    f"{EXPERIMENT_AXIS}_{label}",
                "axis":       EXPERIMENT_AXIS,
                "axis_value": label,
                "status":     "STOP",
                "note":       f"ERROR: {exc}",
            })
            print(f"  ERROR: {exc}")

        rows.append(row)
        write_results(csv_path, rows)
        export_xlsx(rows, xlsx_path)   # 每組跑完即同步 Excel（experiment result/）

        if row["status"] == "STOP" and STOP_ON_ERROR:
            remaining = [
                (c.get("label", str(c)) if isinstance(c, dict) else fmt_value(c))
                for c in cases[run_idx:]
            ]
            if remaining:
                print(f"  跳過後續實驗: {', '.join(str(r) for r in remaining)}")
            break

    print_summary(rows, csv_path)


if __name__ == "__main__":
    main()
