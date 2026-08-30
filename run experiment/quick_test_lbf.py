#!/usr/bin/env python3
"""
quick_test_lbf.py  —  快速測試 BBC+WS+RS+UC+Pareto+LBF 單一設定。

只跑「BBC+WS+RS+UC+Pareto+LBF」一種 config，4 種 risk model，3 種 scale，
時間限制 2 小時（7200 秒），J=20 固定，S=50 情境。
輸出 Excel + CSV 到 experiment result/ 資料夾。

用法：
    cd <專案根目錄>
    python "run experiment/quick_test_lbf.py"

跟 batch_ablation_experiment.py 共用相同的基礎設施（portal、config、logging），
但只跑 LBF 這一個 config，不跑其他 6 個 ablation config。
"""
from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import json
import os
import subprocess
import sys
import tempfile
import time
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Any

# =============================================================================
# 參數設定（與 batch_ablation_experiment.py 保持一致）
# =============================================================================

SCALES = ["small", "medium", "large"]

BASE_SCENARIOS = 50
BASE_CCP_SAMPLE_SIZE = None
BASE_SAMPLE_RATIO = 1.0
BASE_TIME_PERIODS = 8
BASE_DEMAND_MULTIPLIER = 1.0
BASE_ROAD_CAPACITY_MULTIPLIER = 1.0
BASE_HOSPITAL_CAPACITY_MULTIPLIER = 1.0

RISK_ALPHA = 0.9
RISK_LAMBDA = 0.5
BOX_SCOPE = 0.01
ELLIPSOIDAL_SCOPE = 0.0005
POLYHEDRAL_SCOPE = 0.001

TIME_LIMIT = 7200.0     # 2 小時
MIP_GAP = 0.01          # 1% 提早結束
HARD_TIMEOUT_BUFFER_SEC = 1800.0
COMPUTE_KPIS = False

RESULT_PREFIX = "LBF_quick_test"
LOG_SUBDIR_NAME = "lbf quick test"

# =============================================================================
# Setup
# =============================================================================

ROOT_DIR = Path(__file__).resolve().parents[1]
LOG_DIR = ROOT_DIR / "logs"
LOG_SUBDIR = LOG_DIR / LOG_SUBDIR_NAME
RESULT_DIR = ROOT_DIR / "experiment result"

LOCAL_PYTHON_PACKAGE_CANDIDATES = [
    ROOT_DIR / ".codex_spreadsheet" / "python_packages",
    Path.home() / ".cache" / "codex-runtimes" / "codex-primary-runtime"
    / "dependencies" / "python" / "Lib" / "site-packages",
]
if os.environ.get("CODEX_PRIMARY_PYTHON_PACKAGES"):
    LOCAL_PYTHON_PACKAGE_CANDIDATES.insert(
        0, Path(os.environ["CODEX_PRIMARY_PYTHON_PACKAGES"])
    )
for package_dir in reversed(LOCAL_PYTHON_PACKAGE_CANDIDATES):
    if package_dir.exists():
        sys.path.insert(0, str(package_dir))

os.chdir(ROOT_DIR)
for _p in (str(ROOT_DIR / "model core"), str(ROOT_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import config as cfg  # noqa: E402

try:
    import logging_utils  # noqa: E402
except ImportError:
    logging_utils = None

# Portal paths
SP_MODEL_PATH = ROOT_DIR / "model portal" / "benders bbc.py"
DRO_MODEL_PATH = ROOT_DIR / "model portal" / "dro bbc.py"
EXTENSIVE_MODEL_PATH = ROOT_DIR / "model portal" / "extensive_dro.py"
MCVAR_MODEL_PATH = ROOT_DIR / "model portal" / "mcvar bbc.py"

DEFAULT_ROOT_SEED_ITERS = int(cfg.BENDERS_ROOT_SEED_ITERS)
DEFAULT_ROOT_CUT_ROUNDS = int(cfg.BENDERS_ROOT_CUT_ROUNDS)
DEFAULT_MIPFOCUS = int(cfg.BENDERS_MIPFOCUS)
DEFAULT_HEURISTICS = float(cfg.BENDERS_HEURISTICS)
DEFAULT_NUMERIC_FOCUS = int(cfg.BENDERS_NUMERIC_FOCUS)
DEFAULT_BRANCH_PRIORITY_ENABLED = bool(cfg.BENDERS_X_BRANCH_PRIORITY_ENABLED)
DEFAULT_BRANCH_PRIORITY = int(cfg.BENDERS_X_BRANCH_PRIORITY)

# 只跑一個 config：BBC+WS+RS+UC+Pareto+LBF
LBF_CONFIG = {
    "name": "BBC+WS+RS+UC+Pareto+LBF",
    "ev": True,
    "seed": DEFAULT_ROOT_SEED_ITERS,
    "rounds": DEFAULT_ROOT_CUT_ROUNDS,
    "user": True,
    "pareto": True,
    "lbf": True,
}

MODEL_SPECS = [
    {"name": "SP+MCVaR",        "kind": "mcvar"},
    {"name": "DRO-box",         "kind": "dro", "ambiguity_set": "box",         "scope": BOX_SCOPE},
    {"name": "DRO-ellipsoidal", "kind": "dro", "ambiguity_set": "ellipsoidal", "scope": ELLIPSOIDAL_SCOPE},
    {"name": "DRO-polyhedral",  "kind": "dro", "ambiguity_set": "polyhedral",  "scope": POLYHEDRAL_SCOPE},
]
MODELS = [spec["name"] for spec in MODEL_SPECS]
_SPEC_BY_NAME = {spec["name"]: spec for spec in MODEL_SPECS}

FIELDNAMES = [
    "scale", "test_id", "model", "config", "I", "J", "H", "S", "T",
    "obj_value", "first_stage_decision", "best_lb", "best_ub", "cpu_s",
    "wall_s", "num_vars", "num_constrs", "nodes", "iterations", "gap_pct",
    "total_cuts", "seed_cuts", "lazy_cuts", "user_cuts", "root_seed_lb",
    "root_seed_iters_done", "root_cut_rounds_done", "ev_warm_start",
    "root_seed_iters", "root_cut_rounds", "use_user_cuts", "pareto_enabled",
    "lbf_enabled",
    "multi_cut", "parallel_oracles", "oracle_solves", "incumbent_evals",
    "mip_focus", "heuristics", "numeric_focus", "branch_priority_enabled",
    "branch_priority",
    "callback_time_s", "solver_status", "log_path", "solver_log_path",
    "status", "note",
]
TEXT_KEYS = {
    "scale", "test_id", "model", "config", "first_stage_decision",
    "solver_status", "log_path", "solver_log_path", "status", "note",
}
BOOL_KEYS = {
    "ev_warm_start", "use_user_cuts", "pareto_enabled", "lbf_enabled",
    "multi_cut", "branch_priority_enabled",
}
INTEGER_KEYS = {
    "I", "J", "H", "S", "T", "num_vars", "num_constrs", "nodes",
    "iterations", "total_cuts", "seed_cuts", "lazy_cuts", "user_cuts",
    "root_seed_iters_done", "root_cut_rounds_done", "root_seed_iters",
    "root_cut_rounds", "parallel_oracles", "oracle_solves", "incumbent_evals",
    "mip_focus", "numeric_focus", "branch_priority",
}


# =============================================================================
# Utility
# =============================================================================

def scale_counts(scale: str) -> dict[str, int]:
    profile = cfg.SCALE_PROFILES[scale.lower()]
    return {"I": profile["n_disaster"], "J": profile["n_ccp"], "H": profile["n_hospital"]}


def blank_row() -> dict[str, Any]:
    return {key: "NA" for key in FIELDNAMES}


def load_portal(path: Path, module_name: str) -> Any:
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load portal: {path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def first_stage_string(fs: dict | None) -> str:
    if not fs:
        return "NA"
    xs = [j for j, v in fs.get("X", {}).items() if round(v)]
    return f"Open={sorted(xs)} ({len(xs)})"


@contextmanager
def temporary_config(case: dict[str, Any]):
    values = {
        "SCENARIOS": BASE_SCENARIOS,
        "TIME_PERIODS": BASE_TIME_PERIODS,
        "SAMPLE_RATIO": BASE_SAMPLE_RATIO,
        "SP_SAMPLE_RATIO": BASE_SAMPLE_RATIO,
        "DEMAND_MULTIPLIER": BASE_DEMAND_MULTIPLIER,
        "ROAD_CAPACITY_MULTIPLIER": BASE_ROAD_CAPACITY_MULTIPLIER,
        "HOSPITAL_CAPACITY_MULTIPLIER": BASE_HOSPITAL_CAPACITY_MULTIPLIER,
        "SP_TIME_LIMIT": TIME_LIMIT,
        "SP_MIP_GAP": MIP_GAP,
        "BENDERS_MULTI_CUT": True,
        "BENDERS_EV_WARM_START": case["ev"],
        "BENDERS_ROOT_SEED_ITERS": case["seed"],
        "BENDERS_ROOT_CUT_ROUNDS": case["rounds"],
        "BENDERS_USE_USER_CUTS": case["user"],
        "BENDERS_PARETO_ENABLED": case["pareto"],
        "BENDERS_LBF_ENABLED": case.get("lbf", False),
        "BENDERS_INCUMBENT_EARLY_TERMINATION": True,
        "BENDERS_DISPLAY_INTERVAL": 30,  # 每 30 秒印一行 LB/UB/Gap/Time
        "BENDERS_MIPFOCUS": DEFAULT_MIPFOCUS,
        "BENDERS_HEURISTICS": DEFAULT_HEURISTICS,
        "BENDERS_NUMERIC_FOCUS": DEFAULT_NUMERIC_FOCUS,
        "BENDERS_X_BRANCH_PRIORITY_ENABLED": DEFAULT_BRANCH_PRIORITY_ENABLED,
        "BENDERS_X_BRANCH_PRIORITY": DEFAULT_BRANCH_PRIORITY,
    }
    if hasattr(cfg, "CCP_SAMPLE_SIZE"):
        values["CCP_SAMPLE_SIZE"] = BASE_CCP_SAMPLE_SIZE
    original = {key: getattr(cfg, key) for key in values}
    try:
        for key, value in values.items():
            setattr(cfg, key, value)
        yield
    finally:
        for key, value in original.items():
            setattr(cfg, key, value)


@contextmanager
def scaled_instance_generation():
    original = cfg.generate_data
    cache: dict[str, Any] = {}

    def _scaled(*args, **kwargs):
        kwargs.pop("sample_ratio", None)
        kwargs.pop("ccp_sample_size", None)
        scale = cfg.EXPERIMENT_SCALE
        if scale not in cache:
            cache[scale] = original(scale=scale)
        return cache[scale]

    cfg.generate_data = _scaled
    try:
        yield
    finally:
        cfg.generate_data = original


def excel_value(row: dict[str, Any], key: str) -> Any:
    value = row.get(key, "NA")
    if value in (None, "", "NA"):
        return "NA"
    if key in TEXT_KEYS:
        return str(value)
    if key in BOOL_KEYS:
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes"}
    try:
        number = float(value)
        return int(number) if key in INTEGER_KEYS and number.is_integer() else number
    except (TypeError, ValueError):
        return value


# =============================================================================
# Run one case
# =============================================================================

def run_one_case(scale: str, model_name: str) -> dict[str, Any]:
    case = LBF_CONFIG
    counts = scale_counts(scale)
    spec = _SPEC_BY_NAME[model_name]
    if spec["kind"] == "mcvar":
        portal = load_portal(MCVAR_MODEL_PATH, "lbf_mcvar_portal")
    else:
        portal = load_portal(DRO_MODEL_PATH, "lbf_dro_portal")

    test_id = f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
    row = blank_row()
    row.update({
        "scale": scale, "test_id": test_id, "model": model_name, "config": case["name"],
        "I": counts["I"], "J": counts["J"], "H": counts["H"],
        "S": BASE_SCENARIOS, "T": BASE_TIME_PERIODS,
        "ev_warm_start": case["ev"], "root_seed_iters": case["seed"],
        "root_cut_rounds": case["rounds"], "use_user_cuts": case["user"],
        "pareto_enabled": case["pareto"], "lbf_enabled": case.get("lbf", False),
        "multi_cut": True,
        "mip_focus": DEFAULT_MIPFOCUS,
        "heuristics": DEFAULT_HEURISTICS,
        "numeric_focus": DEFAULT_NUMERIC_FOCUS,
        "branch_priority_enabled": DEFAULT_BRANCH_PRIORITY_ENABLED,
        "branch_priority": DEFAULT_BRANCH_PRIORITY,
        "status": "RUNNING", "note": "",
    })

    wall_start = time.time()
    model = summary = None
    try:
        with temporary_config(case), scaled_instance_generation():
            if spec["kind"] == "mcvar":
                model, summary = portal.run_mcvar_model(
                    scenario_size=BASE_SCENARIOS, sample_ratio=BASE_SAMPLE_RATIO,
                    time_limit=TIME_LIMIT, mip_gap=MIP_GAP,
                    alpha=RISK_ALPHA, lam=RISK_LAMBDA, compute_kpis=COMPUTE_KPIS,
                )
            else:
                model, summary = portal.run_dro_model(
                    ambiguity_set=spec["ambiguity_set"], scenario_size=BASE_SCENARIOS,
                    sample_ratio=BASE_SAMPLE_RATIO, time_limit=TIME_LIMIT,
                    mip_gap=MIP_GAP, alpha=RISK_ALPHA, lam=RISK_LAMBDA,
                    scope=spec["scope"], compute_kpis=COMPUTE_KPIS,
                )
    except Exception as exc:  # noqa: BLE001
        row["wall_s"] = f"{time.time() - wall_start:.2f}"
        if model is not None:
            try:
                model.dispose()
            except Exception:
                pass
        row["status"] = "FAIL"
        row["note"] = f"{type(exc).__name__}: {exc}"
        print(f"  -> FAIL: {row['note']}")
        return row

    row["wall_s"] = f"{time.time() - wall_start:.2f}"

    if model is None or summary is None:
        if model is not None:
            try:
                model.dispose()
            except Exception:
                pass
        row["status"], row["note"] = "FAIL", "no feasible solution"
        return row

    st = summary.get("bbc_stats", {})
    fs = summary.get("first_stage")
    objective = float(summary["objective"])
    row.update({
        "obj_value": f"{objective:.6f}",
        "first_stage_decision": first_stage_string(fs),
        "best_lb": f"{float(summary['best_lb']):.6f}",
        "best_ub": f"{objective:.6f}",
        "cpu_s": f"{float(st.get('runtime', float('nan'))):.2f}",
        "num_vars": getattr(model, "NumVars", "NA"),
        "num_constrs": getattr(model, "NumConstrs", "NA"),
        "nodes": f"{float(getattr(model, 'NodeCount', float('nan'))):.0f}",
        "iterations": f"{float(getattr(model, 'IterCount', float('nan'))):.0f}",
        "gap_pct": f"{float(summary['gap_pct']):.6f}",
        "total_cuts": st.get("cuts_added", "NA"),
        "seed_cuts": st.get("seed_cuts_added", "NA"),
        "lazy_cuts": st.get("lazy_cuts_added", "NA"),
        "user_cuts": st.get("user_cuts_added", "NA"),
        "root_seed_lb": st.get("root_seed_lb", "NA"),
        "root_seed_iters_done": st.get("root_seed_iters_done", "NA"),
        "root_cut_rounds_done": st.get("root_cut_rounds_done", "NA"),
        "parallel_oracles": st.get("parallel_oracles", getattr(cfg, "BENDERS_PARALLEL_ORACLES", "NA")),
        "oracle_solves": st.get("oracle_solves", "NA"),
        "incumbent_evals": st.get("incumbent_evals", "NA"),
        "callback_time_s": st.get("callback_time", "NA"),
        "solver_status": st.get("solver_status", "NA"),
        "status": "OK",
    })
    try:
        model.dispose()
    except Exception:
        pass
    print(f"  -> OK obj={row['obj_value']} time={row['cpu_s']}s gap={row['gap_pct']}%")
    return row


# =============================================================================
# CSV / Excel output
# =============================================================================

def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}_", suffix=".csv", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        with temp_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
            writer.writeheader()
            writer.writerows(rows)
        os.replace(temp_path, path)
    finally:
        temp_path.unlink(missing_ok=True)


def export_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[WARN] openpyxl 不可用，跳過 Excel 輸出")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "LBF_results"

    fill = PatternFill("solid", fgColor="2E74B5")
    font = Font(bold=True, color="FFFFFF")

    # Header
    for col, key in enumerate(FIELDNAMES, 1):
        cell = ws.cell(1, col, key)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    for out_row, row in enumerate(rows, 2):
        for col, key in enumerate(FIELDNAMES, 1):
            cell = ws.cell(out_row, col, excel_value(row, key))
            cell.alignment = Alignment(
                horizontal="left" if key in TEXT_KEYS else "right",
                vertical="top",
                wrap_text=key in {"first_stage_decision", "note"},
            )
            if key in INTEGER_KEYS:
                cell.number_format = "#,##0"
            elif key not in TEXT_KEYS and key not in BOOL_KEYS:
                cell.number_format = "#,##0.0000"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}{max(1, len(rows) + 1)}"
    ws.row_dimensions[1].height = 34
    for col, key in enumerate(FIELDNAMES, 1):
        width = 14
        if key == "first_stage_decision":
            width = 62
        elif key in {"test_id", "log_path", "solver_log_path", "note"}:
            width = 34
        elif key in {"config", "solver_status"}:
            width = 30
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Summary sheet ──
    ws2 = wb.create_sheet("summary")
    headers = ["Scale", "I", "J", "H", "S", "Model", "Objective", "Best LB",
               "Gap(%)", "CPU(s)", "Wall(s)", "Cuts", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(1, col, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for out_row, row in enumerate(rows, 2):
        ws2.cell(out_row, 1, row.get("scale", "NA"))
        ws2.cell(out_row, 2, excel_value(row, "I"))
        ws2.cell(out_row, 3, excel_value(row, "J"))
        ws2.cell(out_row, 4, excel_value(row, "H"))
        ws2.cell(out_row, 5, excel_value(row, "S"))
        ws2.cell(out_row, 6, row.get("model", "NA"))
        ws2.cell(out_row, 7, excel_value(row, "obj_value")).number_format = "#,##0.00"
        ws2.cell(out_row, 8, excel_value(row, "best_lb")).number_format = "#,##0.00"
        ws2.cell(out_row, 9, excel_value(row, "gap_pct")).number_format = "0.0000"
        ws2.cell(out_row, 10, excel_value(row, "cpu_s")).number_format = "#,##0.00"
        ws2.cell(out_row, 11, excel_value(row, "wall_s")).number_format = "#,##0.00"
        ws2.cell(out_row, 12, excel_value(row, "total_cuts"))
        ws2.cell(out_row, 13, row.get("status", "NA"))
    ws2.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # ── Settings sheet ──
    ws3 = wb.create_sheet("settings")
    settings = [
        ("config", "BBC+WS+RS+UC+Pareto+LBF"),
        ("scales", ", ".join(SCALES)),
        ("models", ", ".join(MODELS)),
        ("S (scenarios)", BASE_SCENARIOS),
        ("T (time periods)", BASE_TIME_PERIODS),
        ("time_limit_s", TIME_LIMIT),
        ("mip_gap", MIP_GAP),
        ("risk_alpha", RISK_ALPHA),
        ("risk_lambda", RISK_LAMBDA),
        ("EV warm start", True),
        ("Root seed iters", DEFAULT_ROOT_SEED_ITERS),
        ("Root cut rounds", DEFAULT_ROOT_CUT_ROUNDS),
        ("User cuts", True),
        ("Pareto (Papadakos)", True),
        ("LBF (Jensen lower bound)", True),
        ("Incumbent early termination", True),
        ("MIP Focus", DEFAULT_MIPFOCUS),
        ("Heuristics", DEFAULT_HEURISTICS),
        ("Numeric Focus", DEFAULT_NUMERIC_FOCUS),
        ("Branch Priority", DEFAULT_BRANCH_PRIORITY),
        ("Parallel Oracles", int(cfg.BENDERS_PARALLEL_ORACLES)),
    ]
    for col in [1, 2]:
        ws3.column_dimensions[get_column_letter(col)].width = 30
    ws3.append(["Setting", "Value"])
    for cell in ws3[1]:
        cell.fill = fill
        cell.font = font
    for s, v in settings:
        ws3.append([s, v])
    for sc in SCALES:
        c = scale_counts(sc)
        ws3.append([f"  {sc}", f"I={c['I']} J={c['J']} H={c['H']}"])

    # Save
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.stem}_", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
        for attempt in range(5):
            try:
                os.replace(temp_path, path)
                return
            except PermissionError:
                if attempt == 0:
                    print(f"[excel] 無法寫入 {path.name}，重試中 …")
                time.sleep(3.0)
        raise PermissionError(f"無法寫入 {path}")
    finally:
        temp_path.unlink(missing_ok=True)


# =============================================================================
# Main
# =============================================================================

def main() -> None:
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = RESULT_DIR / f"{RESULT_PREFIX}_raw_{timestamp}.csv"
    xlsx_path = RESULT_DIR / f"{RESULT_PREFIX}_{timestamp}.xlsx"

    n_cases = len(SCALES) * len(MODELS)

    print("=" * 72)
    print("LBF QUICK TEST — BBC+WS+RS+UC+Pareto+LBF")
    print("=" * 72)
    print(f"data  : {cfg.DISASTER_CSV} / {cfg.CCP_CSV} / {cfg.HOSPITAL_CSV}")
    print(f"scales: {SCALES}")
    for sc in SCALES:
        c = scale_counts(sc)
        print(f"  {sc:6}: I={c['I']:3d} J={c['J']:3d} H={c['H']:3d}")
    print(f"models: {MODELS}")
    print(f"S={BASE_SCENARIOS}  T={BASE_TIME_PERIODS}")
    print(f"time_limit={TIME_LIMIT:.0f}s  mip_gap={MIP_GAP:.0%}")
    print(f"LBF=True  Incumbent_Early_Term=True")
    print(f"parallel_oracles={int(cfg.BENDERS_PARALLEL_ORACLES)}")
    print(f"DisplayInterval=30s（每 30 秒印一行 Gurobi node log：LB / UB / Gap / Time）")
    print(f"總共 {n_cases} 個 case（3 scale × 4 model × 1 config）")
    worst_h = n_cases * (TIME_LIMIT + HARD_TIMEOUT_BUFFER_SEC) / 3600.0
    print(f"最壞情況總時間 ≈ {worst_h:.0f} 小時")
    print(f"CSV   : {csv_path}")
    print(f"Excel : {xlsx_path}")
    print("=" * 72)

    # ── 驗證 J=20 ──
    for sc in SCALES:
        c = scale_counts(sc)
        assert c["J"] == 20, f"ERROR: {sc} J={c['J']} != 20 ！請檢查 config.py SCALE_PROFILES"
    print("[CHECK] 所有 scale 的 J=20 ✓")
    print(f"[CHECK] S={BASE_SCENARIOS} ✓")
    print()

    rows: list[dict[str, Any]] = []
    run_idx = 0

    try:
        for scale in SCALES:
            original_scale = cfg.EXPERIMENT_SCALE
            cfg.EXPERIMENT_SCALE = scale
            try:
                for model_name in MODELS:
                    run_idx += 1
                    print(f"\n[{run_idx}/{n_cases}] scale={scale} model={model_name} "
                          f"config=BBC+WS+RS+UC+Pareto+LBF")
                    row = run_one_case(scale, model_name)
                    rows.append(row)
                    # 每個 case 結束就落地
                    write_csv(csv_path, rows)
                    try:
                        export_xlsx(rows, xlsx_path)
                    except Exception as exc:
                        print(f"[WARN] Excel 即時匯出失敗: {exc}")
            finally:
                cfg.EXPERIMENT_SCALE = original_scale
    except KeyboardInterrupt:
        write_csv(csv_path, rows)
        try:
            export_xlsx(rows, xlsx_path)
        except Exception:
            pass
        print("\n[INTERRUPTED] 已中斷。已完成的結果都已保存。")
        print(f"  已完成: {sum(r['status'] == 'OK' for r in rows)}/{n_cases}")
        print(f"  CSV: {csv_path}")
        return

    # Final export
    export_xlsx(rows, xlsx_path)

    print("\n" + "=" * 72)
    n_ok = sum(r["status"] == "OK" for r in rows)
    print(f"Done: {n_ok}/{len(rows)} cases OK")
    print()

    # Print summary table
    print(f"{'Scale':<8} {'Model':<20} {'Obj':>14} {'Gap%':>8} {'CPU(s)':>10} {'Status':<6}")
    print("-" * 72)
    for row in rows:
        print(f"{row.get('scale', 'NA'):<8} "
              f"{row.get('model', 'NA'):<20} "
              f"{row.get('obj_value', 'NA'):>14} "
              f"{row.get('gap_pct', 'NA'):>8} "
              f"{row.get('cpu_s', 'NA'):>10} "
              f"{row.get('status', 'NA'):<6}")

    print(f"\nCSV   : {csv_path}")
    print(f"Excel : {xlsx_path}")


if __name__ == "__main__":
    main()
