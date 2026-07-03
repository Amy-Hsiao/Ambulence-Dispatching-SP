#!/usr/bin/env python3
"""
Uncertainty-source sensitivity batch runner (three axes).
(Phase R 重構：原 run_sensitivity_spatial.py，僅改路徑——ROOT_DIR 改為專案根、
 SP 模型路徑指向 model portal/extensive form.py、結果寫入 experiment result/，邏輯零改動)

Axis 1  "spatial_only" : USE_SPATIAL_KMEANS=True,  USE_SCENARIO_OMEGA=False
    Only the K-means spatial demand multipliers are random,
    drawn from Uniform[1-d, 1+d], size-weighted normalized (mean = 1).
    -> isolates the value of knowing WHERE demand concentrates.

Axis 2  "global_only"  : USE_SCENARIO_OMEGA=True,  USE_SPATIAL_KMEANS=False
    Only the global scenario multipliers (demand / road / hospital omega)
    are random, drawn from Uniform[1-d, 1+d].
    -> isolates the value of knowing HOW LARGE the disaster is.

Axis 3  "both"         : USE_SCENARIO_OMEGA=True,  USE_SPATIAL_KMEANS=True
    Global omega fixed at config default [0.8, 1.2]; the spatial multiplier
    range is swept (normalized). This is the paper's main model evaluated at
    each spatial-uncertainty level -> shows interaction of the two sources.

Sweep: d in {0.2, 0.5, 0.7}  ->  ranges [0.8,1.2], [0.5,1.5], [0.3,1.7].

Implementation notes
--------------------
* config.generate_scenarios freezes master_seed=MASTER_SEED as a default
  argument at import time; patched_generate_scenarios() injects it explicitly.
* Every launch re-runs ALL cases from scratch (no resume/skip); the CSV and
  xlsx are rewritten after each completed case.
* CSV column order matches the paper's Excel table:
  factor | I | J | H | S | T | obj_value | First Stage Decision | Best LB |
  Best UB | CPU Time(s) | num_vars | num_constrs | Nodes | Iteration |
  Final Gap(%) | VSS | EVPI   (+ traceability extras at the end)
"""

from __future__ import annotations

import csv
import datetime as dt
import importlib.util
import os
import shutil
import sys
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Any

# =============================================================================
# Parameter setting area
# =============================================================================
DELTAS = [0.2, 0.5, 0.7]          # multiplier range = [1 - d, 1 + d]
SEEDS  = [42]                     # 之後要加 seed 改成 [42, 123, ...]（每次執行全部重跑）

AXES = [
    # (axis_label,      use_omega, use_kmeans)
    ("spatial_only",    False,     True),    # 只有 K-means 空間乘數
    ("global_only",     True,      False),   # 只有全局 scenario 乘數
    ("both",            True,      True),    # 兩者都開：全局 omega 固定 [0.8,1.2]，掃空間乘數範圍（= 主模型）
]

NORMALIZE_SPATIAL = True          # 空間乘數依群集規模加權正規化（僅影響 spatial_only / both 軸）

TIME_LIMIT   = 3600.0
MIP_GAP      = 0.01

RESULT_CSV_NAME  = "sensitivity_uncertainty_results.csv"   # 每次執行整檔重寫
RESULT_XLSX_NAME = "sensitivity_uncertainty_results.xlsx"  # Excel 匯出（欄位對齊論文表格）
LOG_SUBDIR_NAME  = "east district stress test"

# =============================================================================
# Setup
# =============================================================================
ROOT_DIR      = Path(__file__).resolve().parents[1]   # 專案根（Phase R：原為本檔所在目錄）
LOG_DIR       = ROOT_DIR / "logs"
LOG_SUBDIR    = LOG_DIR / LOG_SUBDIR_NAME
SP_MODEL_PATH = ROOT_DIR / "model portal" / "extensive form.py"
RESULT_DIR    = ROOT_DIR / "experiment result"
CSV_PATH      = RESULT_DIR / RESULT_CSV_NAME
XLSX_PATH     = RESULT_DIR / RESULT_XLSX_NAME

os.chdir(ROOT_DIR)
for _p in (str(ROOT_DIR / "model core"), str(ROOT_DIR)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import config as cfg  # noqa: E402

# 欄位順序對齊論文 Excel 表
FIELDNAMES = [
    "factor",
    "I", "J", "H", "S", "T",
    "obj_value",
    "first_stage_decision",
    "best_lb",
    "best_ub",
    "cpu_s",
    "num_vars",
    "num_constrs",
    "nodes",
    "iterations",
    "gap_pct",
    "vss_pct",
    "evpi_pct",
    # ---- 附加欄位（Excel 表不需要，保留供追溯）----
    "test_id",
    "axis",
    "delta",
    "omega_low",
    "omega_high",
    "seed",
    "normalized",
    "use_omega",
    "use_kmeans",
    "ev_obj",
    "eev_obj",
    "ws_obj",
    "vss_abs",
    "evpi_abs",
    "n_open_ccp",
    "wall_s",
    "log_path",
    "status",
    "note",
]


# =============================================================================
# Helpers
# =============================================================================
def load_sp_module():
    if not SP_MODEL_PATH.exists():
        raise FileNotFoundError(f"Cannot find SP model file: {SP_MODEL_PATH}")
    spec = importlib.util.spec_from_file_location("sp_model_runner", SP_MODEL_PATH)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load module from {SP_MODEL_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def fmt(value: Any, digits: int = 4) -> str:
    if value is None:
        return "NA"
    try:
        return f"{float(value):.{digits}f}"
    except Exception:
        return "NA"


def blank_row() -> dict[str, Any]:
    return {key: "NA" for key in FIELDNAMES}


def case_id(axis: str, delta: float, seed: int) -> str:
    return f"{axis}_d{delta:g}_seed{seed}".replace(".", "p")


def load_existing_rows() -> dict[str, dict[str, Any]]:
    """Read previous results keyed by test_id (kept for ad-hoc analysis; not used by main)."""
    if not CSV_PATH.exists():
        return {}
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        return {row["test_id"]: dict(row) for row in csv.DictReader(f) if row.get("test_id")}


def write_results(rows_by_id: dict[str, dict[str, Any]], order: list[str]) -> None:
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        for tid in order:
            if tid in rows_by_id:
                writer.writerow({k: rows_by_id[tid].get(k, "NA") for k in FIELDNAMES})


# Excel 欄位（與論文表格完全一致）: (Excel 標題, 內部欄位名)
XLSX_COLUMNS = [
    ("factor",               "factor"),
    ("|I| Disaster",         "I"),
    ("|J| CCP",              "J"),
    ("|H| Hosp",             "H"),
    ("|S| Scen",             "S"),
    ("|T| Per",              "T"),
    ("obj_value",            "obj_value"),
    ("First Stage Decision", "first_stage_decision"),
    ("Best LB",              "best_lb"),
    ("Best UB",              "best_ub"),
    ("CPU Time(s)",          "cpu_s"),
    ("num_vars",             "num_vars"),
    ("num_constrs",          "num_constrs"),
    ("Nodes",                "nodes"),
    ("Iteration",            "iterations"),
    ("Final Gap(%)",         "gap_pct"),
    ("VSS(%)",               "vss_pct"),
    ("EVPI(%)",              "evpi_pct"),
]


def export_xlsx(rows_by_id: dict[str, dict[str, Any]], order: list[str]) -> None:
    """Export results to xlsx with the paper's table columns. CSV remains the
    source of truth; failure to export (e.g. openpyxl missing or file open in
    Excel) never interrupts the batch."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("  [xlsx] openpyxl 未安裝，略過 Excel 匯出（pip install openpyxl）")
        return
    try:
        RESULT_DIR.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "sensitivity"
        header_fill = PatternFill("solid", fgColor="2E74B5")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, (title, _key) in enumerate(XLSX_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r = 2
        for tid in order:
            row = rows_by_id.get(tid)
            if row is None:
                continue
            for col_idx, (_title, key) in enumerate(XLSX_COLUMNS, start=1):
                value = row.get(key, "NA")
                # numeric conversion where possible (keeps Excel sortable)
                if key not in ("factor", "first_stage_decision"):
                    try:
                        value = float(value)
                        if value == int(value) and key in ("I", "J", "H", "S", "T",
                                                           "num_vars", "num_constrs",
                                                           "nodes", "iterations"):
                            value = int(value)
                    except (TypeError, ValueError):
                        pass
                cell = ws.cell(row=r, column=col_idx, value=value)
                if key == "first_stage_decision":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["H"].width = 62   # First Stage Decision
        wb.save(XLSX_PATH)
    except Exception as exc:
        print(f"  [xlsx] Excel 匯出失敗（不影響 CSV）: {exc}")


def snapshot_logs() -> set[Path]:
    if not LOG_DIR.exists():
        return set()
    return {p.resolve() for p in LOG_DIR.glob("*.log")}


def move_new_log(before: set[Path]) -> Path | None:
    if not LOG_DIR.exists():
        return None
    created = [p for p in LOG_DIR.glob("*.log") if p.resolve() not in before]
    if not created:
        return None
    newest = max(created, key=lambda p: p.stat().st_mtime)
    LOG_SUBDIR.mkdir(parents=True, exist_ok=True)
    dest = LOG_SUBDIR / newest.name
    shutil.move(str(newest), str(dest))
    return dest


def extract_first_stage(model: Any) -> tuple[str, int]:
    """Extract RP first-stage decision from the solved Gurobi model.

    Returns (multi-line string in the log's display format, number of open CCPs).
    """
    if model is None:
        return "NA", 0
    xs: dict[str, float] = {}
    vs: dict[str, float] = {}
    us: dict[str, float] = {}
    ys: dict[str, float] = {}
    for var in model.getVars():
        name = var.VarName
        if name.startswith("X["):
            xs[name[2:-1]] = var.X
        elif name.startswith("V["):
            vs[name[2:-1]] = var.X
        elif name.startswith("U["):
            us[name[2:-1]] = var.X
        elif name.startswith("Y["):
            h, j = name[2:-1].split(",")
            ys[j] = ys.get(j, 0.0) + var.X
        else:
            # first-stage vars are added before recourse vars; stop early
            break
    lines = []
    for j in sorted(xs):
        if xs[j] > 0.5:
            lines.append(
                f"CCP {j:4s} -> X: 1, Staff(V): {vs.get(j, 0):2.0f}, "
                f"Amb(U): {us.get(j, 0):2.0f}, MedicalSupply(Y): {ys.get(j, 0.0):.2f}"
            )
    return ("\n".join(lines) if lines else "none opened"), len(lines)


# =============================================================================
# Config patching
# =============================================================================
@contextmanager
def temporary_config(axis: str, use_omega: bool, use_kmeans: bool,
                     delta: float, seed: int):
    keys: dict[str, Any] = {
        "MASTER_SEED":             seed,
        "USE_SCENARIO_OMEGA":      use_omega,
        "USE_SPATIAL_KMEANS":      use_kmeans,
        "NORMALIZE_SPATIAL_OMEGA": NORMALIZE_SPATIAL,
        "SP_TIME_LIMIT":           TIME_LIMIT,
        "SP_MIP_GAP":              MIP_GAP,
    }
    if axis in ("spatial_only", "both"):
        # 掃空間乘數範圍；"both" 時全局 omega 維持 config 預設 [0.8, 1.2]
        keys["SCENARIO_SPATIAL_OMEGA_LOW"]  = 1.0 - delta
        keys["SCENARIO_SPATIAL_OMEGA_HIGH"] = 1.0 + delta
    elif axis == "global_only":
        keys["SCENARIO_OMEGA_LOW"]  = 1.0 - delta
        keys["SCENARIO_OMEGA_HIGH"] = 1.0 + delta
    else:
        raise ValueError(f"Unknown axis: {axis}")

    original = {key: getattr(cfg, key) for key in keys}
    try:
        for key, value in keys.items():
            setattr(cfg, key, value)
        yield
    finally:
        for key, value in original.items():
            setattr(cfg, key, value)


@contextmanager
def patched_generate_scenarios(seed: int):
    """Inject master_seed explicitly (the function's default arg is frozen at import)."""
    original = cfg.generate_scenarios

    def _patched(*args, **kwargs):
        kwargs.setdefault("master_seed", seed)
        return original(*args, **kwargs)

    cfg.generate_scenarios = _patched
    try:
        yield
    finally:
        cfg.generate_scenarios = original


# =============================================================================
# Core run logic
# =============================================================================
AXIS_FACTOR_LABEL = {
    "spatial_only": "K-means only",
    "global_only":  "Scenario multiplier only",
    "both":         "Both (omega + K-means)",
}


def run_one_case(sp_module: Any, axis: str, use_omega: bool, use_kmeans: bool,
                 delta: float, seed: int,
                 run_idx: int, total_runs: int) -> dict[str, Any]:
    test_id = case_id(axis, delta, seed)
    row = blank_row()
    row.update({
        "factor":     f"{AXIS_FACTOR_LABEL[axis]} U[{1.0 - delta:g},{1.0 + delta:g}]",
        "test_id":    test_id,
        "axis":       axis,
        "delta":      f"{delta:g}",
        "omega_low":  f"{1.0 - delta:g}",
        "omega_high": f"{1.0 + delta:g}",
        "seed":       seed,
        "normalized": str(NORMALIZE_SPATIAL if axis in ("spatial_only", "both") else "NA"),
        "use_omega":  str(use_omega),
        "use_kmeans": str(use_kmeans),
        "status":     "RUNNING",
        "note":       "",
    })

    print(f"\n[{run_idx}/{total_runs}] {test_id}  "
          f"({AXIS_FACTOR_LABEL[axis]}, U[{1.0 - delta:g}, {1.0 + delta:g}])")

    logs_before = snapshot_logs()
    wall_start  = time.time()
    model = summary = None
    error_note = ""

    try:
        with temporary_config(axis, use_omega, use_kmeans, delta, seed), \
             patched_generate_scenarios(seed):
            model, summary = sp_module.run_sp_model(
                time_limit=TIME_LIMIT,
                mip_gap=MIP_GAP,
            )
    except Exception as exc:
        error_note = f"ERROR: {exc}"
        print(f"  {error_note}")

    log_path = move_new_log(logs_before)
    row["wall_s"]   = fmt(time.time() - wall_start, 1)
    row["log_path"] = str(log_path) if log_path is not None else "NA"

    if model is None or summary is None:
        row["status"] = "FAIL"
        row["note"]   = error_note or "INFEASIBLE / no solution returned"
        return row

    rp  = summary.get("RP", {})
    ev  = summary.get("EV", {})
    eev = summary.get("EEV", {})
    ws  = summary.get("WS", {})

    first_stage_txt, n_open = extract_first_stage(model)

    row.update({
        "obj_value":   fmt(rp.get("best_ub"), 2),
        "best_ub":     fmt(rp.get("best_ub"), 2),
        "best_lb":     fmt(rp.get("best_lb"), 2),
        "num_vars":    int(getattr(model, "NumVars", 0)),
        "num_constrs": int(getattr(model, "NumConstrs", 0)),
        "nodes":       int(getattr(model, "NodeCount", 0)),
        "iterations":  int(getattr(model, "IterCount", 0)),
        "gap_pct":     fmt(rp.get("gap"), 4),
        "ev_obj":      fmt(ev.get("best_ub"), 2),
        "eev_obj":     fmt(eev.get("best_ub"), 2),
        "ws_obj":      fmt(ws.get("best_ub_weighted"), 2),
        "vss_abs":     fmt(summary.get("VSS"), 2),
        "evpi_abs":    fmt(summary.get("EVPI"), 2),
        "vss_pct":     fmt(summary.get("VSS_pct"), 4),
        "evpi_pct":    fmt(summary.get("EVPI_pct"), 4),
        "n_open_ccp":  n_open,
        "first_stage_decision": first_stage_txt,
        "cpu_s":       fmt(getattr(model, "Runtime", None), 2),
        "status":      "OK",
        "note":        "; ".join(summary.get("warnings", [])),
    })
    print(f"  OK | VSS%={row['vss_pct']} | EVPI%={row['evpi_pct']} | "
          f"gap={row['gap_pct']}% | open CCPs={n_open}")
    return row


def fill_dimensions(row: dict[str, Any]) -> None:
    """Fill I/J/H/S/T columns from current config (base setting, fixed for all runs)."""
    row["S"] = cfg.SCENARIOS
    row["T"] = cfg.TIME_PERIODS
    try:
        data_dir = ROOT_DIR / "data"

        def _count(name: str) -> int:
            p = data_dir / name
            if not p.exists():
                return -1
            with p.open(encoding="utf-8-sig", newline="") as f:
                return max(0, sum(1 for _ in f) - 1)

        row["I"] = _count(cfg.DISASTER_CSV)
        row["J"] = _count(cfg.CCP_CSV)
        row["H"] = _count(cfg.HOSPITAL_CSV)
    except Exception:
        pass


# =============================================================================
# Summary
# =============================================================================
def print_final_summary(rows_by_id: dict[str, dict[str, Any]]) -> None:
    print("\n" + "=" * 78)
    print("UNCERTAINTY-SOURCE SENSITIVITY — VSS% / EVPI% by axis and range")
    hdr = (f"{'axis':>26} | {'range':>11} | {'VSS%':>9} | {'EVPI%':>9} | "
           f"{'gap%':>7} | status")
    print(hdr)
    print("-" * len(hdr))
    for axis, _uo, _uk in AXES:
        for delta in DELTAS:
            for seed in SEEDS:
                tid = case_id(axis, delta, seed)
                r = rows_by_id.get(tid)
                if r is None:
                    continue
                print(f"{AXIS_FACTOR_LABEL[axis]:>26} | "
                      f"[{1-delta:g},{1+delta:g}]".rjust(12) +
                      f" | {str(r.get('vss_pct','NA')):>9} | "
                      f"{str(r.get('evpi_pct','NA')):>9} | "
                      f"{str(r.get('gap_pct','NA')):>7} | {r.get('status','NA')}")
    print("=" * 78)
    print(f"Results CSV: {CSV_PATH}")


# =============================================================================
# Main
# =============================================================================
def main() -> None:
    LOG_SUBDIR.mkdir(parents=True, exist_ok=True)
    RESULT_DIR.mkdir(parents=True, exist_ok=True)
    sp_module = load_sp_module()

    cases = [(axis, uo, uk, d, s)
             for (axis, uo, uk) in AXES for d in DELTAS for s in SEEDS]
    order = [case_id(axis, d, s) for (axis, _uo, _uk, d, s) in cases]
    rows_by_id: dict[str, dict[str, Any]] = {}   # 每次執行全部重跑，不讀舊結果
    todo = cases

    print("=" * 78)
    print("UNCERTAINTY-SOURCE SENSITIVITY RUNNER")
    print(f"Start   : {dt.datetime.now():%Y-%m-%d %H:%M:%S}")
    print(f"Axes    : {[a for a, _, _ in AXES]}")
    print(f"Deltas  : {DELTAS}   Seeds: {SEEDS}   ({len(cases)} runs total)")
    print(f"Fixed   : S={cfg.SCENARIOS}, T={cfg.TIME_PERIODS}, "
          f"normalize_spatial={NORMALIZE_SPATIAL}")
    print(f"Solver  : time_limit={TIME_LIMIT:.0f}s, mip_gap={MIP_GAP:g}")
    print(f"Rerun   : all {len(todo)} cases run from scratch (no resume)")
    print(f"Logs    : {LOG_SUBDIR}")
    print(f"Output  : {CSV_PATH}")
    print(f"Excel   : {XLSX_PATH}")
    print("=" * 78)

    for run_idx, (axis, uo, uk, delta, seed) in enumerate(todo, start=1):
        row = run_one_case(sp_module, axis, uo, uk, delta, seed, run_idx, len(todo))
        fill_dimensions(row)
        rows_by_id[row["test_id"]] = row
        write_results(rows_by_id, order)   # write after every run (resume safety)
        export_xlsx(rows_by_id, order)     # keep the Excel file in sync too

    print_final_summary(rows_by_id)
    print(f"End     : {dt.datetime.now():%Y-%m-%d %H:%M:%S}")


if __name__ == "__main__":
    main()
