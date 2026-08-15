#!/usr/bin/env python3
"""Ablation runner 的全流程 dry run（不需要 Gurobi、不會真的求解）。

把 run_one_case_subprocess 換成假的求解器，跑完整的 main() 流程，驗證：
  1. 72 個 case 的矩陣正確、順序正確
  2. CSV 與 Excel 都成功產生在 experiment result/
  3. Excel 六個分頁齊全、summary_table 結構正確（含 Obj Value）
  4. 有 case 失敗時不會中止整批，且 Excel 仍可正常輸出
  5. 中斷後用 RESUME_FROM_CSV 續跑，只補跑缺的 case
  6. Ctrl+C 中斷時結果有保存、並印出續跑指令

輸出寫到暫存資料夾，不會汙染真正的 experiment result/。

用法：
    python tests/dry_run_ablation.py
"""
from __future__ import annotations

import importlib.util
import shutil
import sys
import tempfile
import types
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "model core"))
sys.path.insert(0, str(ROOT))

# Gurobi 在檢查用環境不一定裝得起來；dry run 不需要真的求解。
sys.modules.setdefault("gurobipy", types.ModuleType("gurobipy"))

spec = importlib.util.spec_from_file_location(
    "ABL", ROOT / "run experiment" / "batch_ablation_experiment.py"
)
ABL = importlib.util.module_from_spec(spec)
spec.loader.exec_module(ABL)

PASS, FAIL = [], []


def check(name: str, condition: bool, detail: str = "") -> None:
    (PASS if condition else FAIL).append(name)
    mark = "OK  " if condition else "FAIL"
    print(f"  [{mark}] {name}" + (f"  — {detail}" if detail else ""))


def fake_runner(fail_these: set[str] | None = None):
    """假的單一 case 執行器：回傳結構正確的 row，可指定某些 case 失敗。"""
    fail_these = fail_these or set()

    def _run(model_name, case, counts, scale, run_idx, total, log_run_dir):
        test_id = f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
        log_path = log_run_dir / f"{run_idx:02d}_{ABL._safe_case_name(test_id)}.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text("dry run\n", encoding="utf-8")
        row = ABL.blank_row()
        row.update({
            "scale": scale, "test_id": test_id, "model": model_name,
            "config": case["name"], "I": counts["I"], "J": counts["J"],
            "H": counts["H"], "S": ABL.BASE_SCENARIOS, "T": ABL.BASE_TIME_PERIODS,
            "log_path": str(log_path), "solver_log_path": "NA",
        })
        if test_id in fail_these:
            row.update({"status": "FAIL", "note": "dry-run injected failure",
                        "solver_status": "ERROR"})
            return row
        # 同一 (scale, model) 下六個 config 目標值一致（符合 objective_warnings 預期）
        base = 1_000_000.0 + 1000 * ABL.SCALES.index(scale) + ABL.MODELS.index(model_name)
        row.update({
            "obj_value": f"{base:.6f}", "best_lb": f"{base * 0.995:.6f}",
            "best_ub": f"{base:.6f}", "cpu_s": f"{100.0 + run_idx:.2f}",
            "wall_s": f"{110.0 + run_idx:.2f}", "num_vars": "123456",
            "num_constrs": "65432", "nodes": str(500 + run_idx),
            "iterations": "42", "gap_pct": "0.500000",
            "total_cuts": "10", "seed_cuts": "3", "lazy_cuts": "5",
            "user_cuts": "2", "root_seed_lb": f"{base * 0.9:.6f}",
            "root_seed_iters_done": "3", "root_cut_rounds_done": "2",
            "ev_warm_start": case["ev"], "root_seed_iters": case["seed"],
            "root_cut_rounds": case["rounds"], "use_user_cuts": case["user"],
            "pareto_enabled": case["pareto"],
            "multi_cut": case.get("engine") != "extensive",
            "parallel_oracles": "5", "oracle_solves": "150",
            "incumbent_evals": "20", "mip_focus": "3", "heuristics": "0.05",
            "numeric_focus": "1", "branch_priority_enabled": True,
            "branch_priority": "10", "callback_time_s": "12.5",
            "solver_status": "OPTIMAL", "status": "OK", "note": "",
            "first_stage_decision": "CCP 1 -> X: 1, Staff(V): 100, Amb(U): 20, MedicalSupply(Y): 300.00",
        })
        return row

    return _run


def inspect_outputs(result_dir: Path, label: str) -> tuple[Path, Path]:
    csvs = sorted(result_dir.glob("*_raw_*.csv"))
    xlsxs = sorted(p for p in result_dir.glob("*.xlsx") if not p.name.startswith("."))
    check(f"[{label}] 產生 raw CSV", len(csvs) == 1, str(csvs[-1].name) if csvs else "無")
    check(f"[{label}] 產生 Excel", len(xlsxs) == 1, str(xlsxs[-1].name) if xlsxs else "無")
    return csvs[-1], xlsxs[-1]


def verify_excel(xlsx: Path, label: str, expected_rows: int) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=False)
    try:
        expected_sheets = ["raw_results", "run_settings", *ABL.SCALES, "summary_table"]
        check(f"[{label}] Excel 分頁齊全", wb.sheetnames == expected_sheets,
              str(wb.sheetnames))
        check(f"[{label}] raw_results 列數 = {expected_rows}+1",
              wb["raw_results"].max_row == expected_rows + 1,
              str(wb["raw_results"].max_row))

        n_left, n_sub = 5, len(ABL.SUMMARY_SUBCOLUMNS)
        ws = wb["summary_table"]
        check(f"[{label}] summary_table 大小 = "
              f"{2 + len(ABL.SCALES) * len(ABL.MODELS)} 列 × {n_left + len(ABL.CONFIGS) * n_sub} 欄",
              ws.max_row == 2 + len(ABL.SCALES) * len(ABL.MODELS)
              and ws.max_column == n_left + len(ABL.CONFIGS) * n_sub,
              f"{ws.max_row} × {ws.max_column}")

        obj_headers = [ws.cell(2, n_left + 1 + i * n_sub).value for i in range(len(ABL.CONFIGS))]
        check(f"[{label}] summary_table 每個 config 都有 Obj Value 欄",
              obj_headers == ["Obj Value"] * len(ABL.CONFIGS), str(obj_headers))

        obj_values = [ws.cell(3, n_left + 1 + i * n_sub).value for i in range(len(ABL.CONFIGS))]
        check(f"[{label}] summary_table 的 Obj Value 有實際數值",
              all(isinstance(v, (int, float)) or v in ("FAIL", "NA") for v in obj_values),
              str(obj_values[:3]))

        # 每個 scale 分頁的列數
        for scale in ABL.SCALES:
            per_scale = len(ABL.MODELS) * len(ABL.CONFIGS)
            check(f"[{label}] 分頁 {scale} 列數 = {per_scale}+1",
                  wb[scale].max_row == per_scale + 1, str(wb[scale].max_row))

        settings = {r[0].value: r[1].value for r in wb["run_settings"].iter_rows(min_row=2)
                    if r[0].value}
        check(f"[{label}] run_settings 記錄了資料檔與縮放設定",
              "data_ccp_csv" in settings and "ccp_upper_bound_scaling" in settings,
              f"ccp={settings.get('data_ccp_csv')}, mode={settings.get('ccp_upper_bound_scaling')}")
    finally:
        wb.close()


def scenario_full_run(tmp: Path) -> Path:
    print("\n--- 情境 A：完整 72 個 case 全部成功 ---")
    result_dir = tmp / "A"
    ABL.RESULT_DIR = result_dir
    ABL.LOG_SUBDIR = tmp / "A_logs"
    ABL.RESUME_FROM_CSV = ""
    ABL.run_one_case_subprocess = fake_runner()
    ABL.main()
    csv_path, xlsx = inspect_outputs(result_dir, "A")
    verify_excel(xlsx, "A", 72)
    check("[A] CSV 資料列數 = 72", ABL.csv_row_count(csv_path) == 72,
          str(ABL.csv_row_count(csv_path)))
    return csv_path


def scenario_with_failures(tmp: Path) -> Path:
    print("\n--- 情境 B：有 5 個 case 失敗（模擬 OOM / solver 例外）---")
    result_dir = tmp / "B"
    ABL.RESULT_DIR = result_dir
    ABL.LOG_SUBDIR = tmp / "B_logs"
    ABL.RESUME_FROM_CSV = ""
    failing = {
        "large_SP_plus_MCVaR_Extensive", "large_DRO-box_Extensive",
        "large_DRO-ellipsoidal_Extensive", "medium_DRO-box_Extensive",
        "small_DRO-polyhedral_BBC",
    }
    ABL.run_one_case_subprocess = fake_runner(failing)
    ABL.main()
    csv_path, xlsx = inspect_outputs(result_dir, "B")
    verify_excel(xlsx, "B", 72)
    rows = ABL._load_prior_rows(csv_path)
    n_fail = sum(1 for r in rows if r["status"] == "FAIL")
    check("[B] 失敗 case 不會中止整批（仍有 72 列）", len(rows) == 72, str(len(rows)))
    check("[B] 失敗 case 被記為 FAIL", n_fail == 5, f"{n_fail} 個")
    return csv_path


def scenario_resume(tmp: Path, prior_csv: Path) -> None:
    print("\n--- 情境 C：用情境 B 的 CSV 續跑（只補跑 5 個失敗的）---")
    result_dir = tmp / "C"
    ABL.RESULT_DIR = result_dir
    ABL.LOG_SUBDIR = tmp / "C_logs"
    # 把上次的 CSV 複製進新的 result dir，模擬使用者填檔名的情境
    result_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(prior_csv, result_dir / prior_csv.name)
    ABL.RESUME_FROM_CSV = prior_csv.name

    reran: list[str] = []
    base = fake_runner()

    def counting(model_name, case, counts, scale, run_idx, total, log_run_dir):
        test_id = f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
        reran.append(test_id)
        return base(model_name, case, counts, scale, run_idx, total, log_run_dir)

    ABL.run_one_case_subprocess = counting
    ABL.main()
    ABL.RESUME_FROM_CSV = ""

    check("[C] 只重跑上次失敗的 5 個 case", len(reran) == 5, f"實際重跑 {len(reran)} 個")
    csvs = sorted(p for p in result_dir.glob("*_raw_*.csv") if p.name != prior_csv.name)
    rows = ABL._load_prior_rows(csvs[-1])
    check("[C] 續跑後湊回完整 72 列", len(rows) == 72, str(len(rows)))
    check("[C] 續跑後全部 OK", all(r["status"] == "OK" for r in rows),
          f"{sum(1 for r in rows if r['status'] == 'OK')}/72")


def scenario_keyboard_interrupt(tmp: Path) -> None:
    print("\n--- 情境 D：跑到第 10 個 case 時 Ctrl+C ---")
    result_dir = tmp / "D"
    ABL.RESULT_DIR = result_dir
    ABL.LOG_SUBDIR = tmp / "D_logs"
    ABL.RESUME_FROM_CSV = ""
    base = fake_runner()
    seen = {"n": 0}

    def interrupting(model_name, case, counts, scale, run_idx, total, log_run_dir):
        seen["n"] += 1
        if seen["n"] > 10:
            raise KeyboardInterrupt
        return base(model_name, case, counts, scale, run_idx, total, log_run_dir)

    ABL.run_one_case_subprocess = interrupting
    try:
        ABL.main()
        interrupted_cleanly = True
    except KeyboardInterrupt:
        interrupted_cleanly = False
    check("[D] Ctrl+C 被攔截，沒有往外拋出", interrupted_cleanly)
    csvs = sorted(result_dir.glob("*_raw_*.csv"))
    check("[D] 中斷時 CSV 已保存", bool(csvs))
    if csvs:
        rows = ABL._load_prior_rows(csvs[-1])
        check("[D] 已完成的 10 個 case 都在檔案裡", len(rows) == 10, f"{len(rows)} 列")


def scenario_excel_locked(tmp: Path) -> None:
    """模擬「使用者在 Excel 開著輸出檔」→ os.replace 丟 PermissionError。

    中途匯出失敗只能警告，絕不可讓跑了好幾小時的整批實驗掛掉。
    """
    print("\n--- 情境 E：實驗途中輸出檔被 Excel 鎖住 ---")
    result_dir = tmp / "E"
    ABL.RESULT_DIR = result_dir
    ABL.LOG_SUBDIR = tmp / "E_logs"
    ABL.RESUME_FROM_CSV = ""
    ABL.run_one_case_subprocess = fake_runner()

    import os as _os
    real_replace = _os.replace
    state = {"n": 0}

    def flaky_replace(src, dst, *a, **kw):
        if not str(dst).endswith(".xlsx"):
            return real_replace(src, dst, *a, **kw)
        state["n"] += 1
        # 第 1 次是 main() 開頭的 preflight（檔名帶新 timestamp，實務上不可能被鎖）。
        # 第 2~40 次是實驗途中的即時匯出 → 模擬使用者開著 Excel。
        # 之後解除鎖定，確認最終匯出仍會成功。
        if 2 <= state["n"] < 40:
            raise PermissionError(13, "被其他程序使用中")
        return real_replace(src, dst, *a, **kw)

    ABL.EXCEL_REPLACE_RETRY_SLEEP_SEC = 0.0  # 測試不要真的等
    _os.replace = flaky_replace
    try:
        ABL.main()
        survived = True
    except Exception as exc:  # noqa: BLE001
        survived = False
        print(f"      例外：{type(exc).__name__}: {exc}")
    finally:
        _os.replace = real_replace
        ABL.EXCEL_REPLACE_RETRY_SLEEP_SEC = 3.0

    check("[E] Excel 被鎖住時實驗沒有中止", survived)
    csvs = sorted(result_dir.glob("*_raw_*.csv"))
    check("[E] CSV 仍完整寫出 72 列", bool(csvs) and ABL.csv_row_count(csvs[-1]) == 72,
          str(ABL.csv_row_count(csvs[-1])) if csvs else "無 CSV")
    xlsxs = [p for p in result_dir.glob("*.xlsx") if not p.name.startswith(".")]
    check("[E] 鎖住解除後最終仍產生 Excel", bool(xlsxs),
          xlsxs[-1].name if xlsxs else "無")


def main() -> int:
    print("=" * 78)
    print("ABLATION RUNNER 全流程 DRY RUN（不需 Gurobi，不會真的求解）")
    print("=" * 78)
    print(f"矩陣：{len(ABL.SCALES)} scale × {len(ABL.MODELS)} model "
          f"× {len(ABL.CONFIGS)} config = {len(ABL.expected_test_ids())} cases")
    print(f"模型：{ABL.MODELS}")
    print(f"時限：{ABL.TIME_LIMIT}s，gap 門檻 {ABL.MIP_GAP}")
    print(f"硬超時：B&BC {ABL.hard_timeout_for({}):.0f}s / "
          f"Extensive {ABL.hard_timeout_for({'engine': 'extensive'}):.0f}s")

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        import contextlib
        import io

        # main() 的畫面輸出很長，只在失敗時才需要，先吞掉
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            csv_a = scenario_full_run(tmp)
            csv_b = scenario_with_failures(tmp)
            scenario_resume(tmp, csv_b)
            scenario_keyboard_interrupt(tmp)
            scenario_excel_locked(tmp)
        # 只印出 check() 的結果行
        for line in buf.getvalue().splitlines():
            if line.startswith("  [") or line.startswith("---") or line.startswith("\n---"):
                print(line)

    print("\n" + "=" * 78)
    print(f"通過 {len(PASS)} 項，失敗 {len(FAIL)} 項")
    if FAIL:
        for name in FAIL:
            print(f"  [FAIL] {name}")
        return 1
    print("全部通過 — runner 可以正常跑完並輸出結果。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
