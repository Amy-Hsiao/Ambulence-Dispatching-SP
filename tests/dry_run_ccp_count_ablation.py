#!/usr/bin/env python3
"""batch_ccp_count_ablation.py 的全流程 dry run（不需 Gurobi，不會真的求解）。

驗證：
  1. 24 個 case 的矩陣與順序正確
  2. 參數 preflight 正確（per-CCP 上限不隨 |J| 改變、不變量通過）
  3. CSV / Excel 四個分頁齊全，summary_table 是 UB/LB/Time/Gap/Nodes
  4. 有 case 失敗時不會中止整批
  5. 中斷可續跑、Excel 被鎖住不會讓實驗掛掉

用法：python tests/dry_run_ccp_count_ablation.py
"""
from __future__ import annotations

import contextlib
import importlib.util
import io
import sys
import tempfile
import types
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "model core"))
sys.path.insert(0, str(ROOT))
sys.modules.setdefault("gurobipy", types.ModuleType("gurobipy"))

spec = importlib.util.spec_from_file_location(
    "CCPABL", ROOT / "run experiment" / "batch_ccp_count_ablation.py"
)
M = importlib.util.module_from_spec(spec)
spec.loader.exec_module(M)

PASS, FAIL = [], []


def check(name: str, cond: bool, detail: str = "") -> None:
    (PASS if cond else FAIL).append(name)
    print(f"  [{'OK  ' if cond else 'FAIL'}] {name}" + (f"  — {detail}" if detail else ""))


def fake_runner(fail_these: set[str] | None = None):
    fail_these = fail_these or set()

    def _run(case, scale, ccp_count, run_idx, total, log_run_dir):
        test_id = M.make_test_id(scale, ccp_count, case["name"])
        log_path = log_run_dir / f"{run_idx:02d}_{M._safe_case_name(test_id)}.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text("dry run\n", encoding="utf-8")
        counts = M.scale_counts(scale, ccp_count)
        row = M.blank_row()
        row.update({
            "scale": scale, "test_id": test_id, "ccp_count": ccp_count,
            "config": case["name"], "I": counts["I"], "J": counts["J"],
            "H": counts["H"], "S": M.BASE_SCENARIOS, "T": M.BASE_TIME_PERIODS,
            "log_path": str(log_path), "solver_log_path": "NA",
        })
        if test_id in fail_these:
            row.update({"status": "FAIL", "note": "injected", "solver_status": "ERROR"})
            return row
        # |J| 越小應該越好解 → 用遞減的時間/節點模擬
        idx = M.CCP_COUNTS.index(ccp_count)
        ub = 25_000_000.0 + 100_000 * idx
        row.update({
            "first_stage_space": 2 ** ccp_count,
            "obj_value": f"{ub:.6f}", "best_ub": f"{ub:.6f}",
            "best_lb": f"{ub * 0.7:.6f}", "gap_pct": f"{30.0 - 5 * idx:.6f}",
            "cpu_s": f"{7200.0 - 1000 * idx:.2f}", "wall_s": "7300.00",
            "nodes": str(50000 // (idx + 1)), "iterations": "123456",
            "num_vars": "5165050", "num_constrs": "1272740",
            "opened_ccps": "8", "first_stage_decision": "CCP 1 -> X: 1, ...",
            "first_stage_cost": "14952000", "total_cuts": "21708",
            "seed_cuts": "4530", "lazy_cuts": "16728", "user_cuts": "450",
            "root_seed_lb": "11000000", "root_seed_iters_done": "600",
            "root_seed_stop_reason": "max_iters", "root_seed_time_s": "1460.0",
            "root_cut_rounds_done": "15", "oracle_solves": "22050",
            "incumbent_evals": "420", "callback_time_s": "4824.33",
            "parallel_oracles": "10", "ccp_staff_ub": "57",
            "ccp_ambulance_ub": "10", "total_staff_pool": "299",
            "total_amb_pool": "72", "solver_status": "TIME_LIMIT",
            "status": "OK", "note": "",
        })
        return row

    return _run


def verify_excel(xlsx: Path, label: str, expected_rows: int) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, data_only=False)
    try:
        expected = ["raw_results", "run_settings", "parameters", "summary_table"]
        check(f"[{label}] Excel 分頁齊全", wb.sheetnames == expected, str(wb.sheetnames))
        check(f"[{label}] raw_results 列數", wb["raw_results"].max_row == expected_rows + 1,
              str(wb["raw_results"].max_row))

        ws = wb["summary_table"]
        n_left, n_sub = 5, 5
        check(f"[{label}] summary_table 大小",
              ws.max_row == 2 + len(M.SCALES) * len(M.CCP_COUNTS)
              and ws.max_column == n_left + len(M.CONFIGS) * n_sub,
              f"{ws.max_row} × {ws.max_column}")
        subs = [ws.cell(2, n_left + 1 + i).value for i in range(n_sub)]
        check(f"[{label}] summary 子欄位 = UB/LB/Time/Gap/Nodes",
              subs == ["UB", "LB", "Time(s)", "Gap(%)", "Nodes"], str(subs))
        cfg_names = [ws.cell(1, n_left + 1 + i * n_sub).value for i in range(len(M.CONFIGS))]
        check(f"[{label}] summary 兩個 config 標題",
              cfg_names == [c["name"] for c in M.CONFIGS], str(cfg_names))
        # |J| 欄應該是 50/40/30/20 依序
        js = [ws.cell(3 + i, 3).value for i in range(len(M.CCP_COUNTS))]
        check(f"[{label}] summary 的 |J| 欄", js == M.CCP_COUNTS, str(js))

        params = wb["parameters"]
        check(f"[{label}] parameters 分頁有 {len(M.SCALES) * len(M.CCP_COUNTS)} 列",
              params.max_row >= 1 + len(M.SCALES) * len(M.CCP_COUNTS))
    finally:
        wb.close()


def main() -> int:
    print("=" * 78)
    print("batch_ccp_count_ablation.py — 全流程 DRY RUN")
    print("=" * 78)
    n = len(M.SCALES) * len(M.CCP_COUNTS) * len(M.CONFIGS)
    print(f"矩陣：{len(M.SCALES)} scale × {len(M.CCP_COUNTS)} |J| × "
          f"{len(M.CONFIGS)} config = {n} cases")
    print(f"|J| = {M.CCP_COUNTS}")
    print(f"config = {[c['name'] for c in M.CONFIGS]}")
    print(f"時限 = {M.TIME_LIMIT:.0f}s，縮放模式 = {M.CCP_SCALING_MODE}\n")

    check("矩陣共 24 個 case", n == 24, str(n))
    check("只有 2 個 config（Extensive + BBC 全開）", len(M.CONFIGS) == 2)
    check("config 內容正確",
          [c["name"] for c in M.CONFIGS] == ["Extensive", "BBC+WS+RS+UC+Pareto"])
    check("|J| = 50/40/30/20", M.CCP_COUNTS == [50, 40, 30, 20])
    check("時限 7200s", M.TIME_LIMIT == 7200.0)
    check("預設從頭跑", M.RESUME_FROM_CSV == "")
    check("expected_test_ids 共 24 個", len(M.expected_test_ids()) == 24)

    print("\n--- 參數 preflight（per-CCP 上限應不隨 |J| 改變）---")
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        param_rows = M.preflight_parameters()
    for line in buf.getvalue().splitlines():
        print("   " + line)
    small = [r for r in param_rows if r["scale"] == "small"]
    check("small 的單CCP醫護上限不隨 |J| 改變",
          len({r["ccp_staff_ub"] for r in small}) == 1,
          str(sorted({r["ccp_staff_ub"] for r in small})))
    check("small 的單CCP救護車上限不隨 |J| 改變",
          len({r["ccp_amb_ub"] for r in small}) == 1,
          str(sorted({r["ccp_amb_ub"] for r in small})))
    spaces = [r["first_stage_space"] for r in small]
    check("一階搜尋空間隨 |J| 遞減（這才是唯一變因）",
          spaces == sorted(spaces, reverse=True),
          " → ".join(f"2^{M.CCP_COUNTS[i]}" for i in range(len(spaces))))
    check("一階搜尋空間 = 2^|J|",
          spaces == [2 ** j for j in M.CCP_COUNTS],
          f"{spaces[0]:.3e} … {spaces[-1]:.3e}")
    check("所有組合都通過設計不變量",
          all(r["inv_pool_binding"] and r["inv_amb_deployable"] for r in param_rows))

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        print("\n--- 情境 A：24 個 case 全部成功 ---")
        M.RESULT_DIR = tmp / "A"
        M.LOG_SUBDIR = tmp / "A_logs"
        M.RESUME_FROM_CSV = ""
        M.run_one_case_subprocess = fake_runner()
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            M.main()
        csvs = sorted((tmp / "A").glob("*_raw_*.csv"))
        xlsxs = [p for p in (tmp / "A").glob("*.xlsx") if not p.name.startswith(".")]
        check("[A] 產生 CSV", len(csvs) == 1, csvs[-1].name if csvs else "無")
        check("[A] CSV 有 24 列", M.csv_row_count(csvs[-1]) == 24)
        check("[A] 產生 Excel", len(xlsxs) == 1)
        verify_excel(xlsxs[-1], "A", 24)

        print("\n--- 情境 B：4 個 case 失敗 ---")
        M.RESULT_DIR = tmp / "B"
        M.LOG_SUBDIR = tmp / "B_logs"
        failing = {M.make_test_id("large", j, "Extensive") for j in [50, 40]} | \
                  {M.make_test_id("medium", 50, "Extensive"),
                   M.make_test_id("small", 20, "BBC+WS+RS+UC+Pareto")}
        M.run_one_case_subprocess = fake_runner(failing)
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            M.main()
        csv_b = sorted((tmp / "B").glob("*_raw_*.csv"))[-1]
        rows_b = M._load_prior_rows(csv_b)
        check("[B] 失敗不中止整批（仍 24 列）", len(rows_b) == 24, str(len(rows_b)))
        check("[B] 4 個記為 FAIL",
              sum(1 for r in rows_b if r["status"] == "FAIL") == 4)
        xlsx_b = [p for p in (tmp / "B").glob("*.xlsx") if not p.name.startswith(".")][-1]
        verify_excel(xlsx_b, "B", 24)

        print("\n--- 情境 C：續跑只補跑失敗的 4 個 ---")
        M.RESULT_DIR = tmp / "C"
        M.LOG_SUBDIR = tmp / "C_logs"
        (tmp / "C").mkdir(parents=True, exist_ok=True)
        import shutil as _sh
        _sh.copy2(csv_b, tmp / "C" / csv_b.name)
        M.RESUME_FROM_CSV = csv_b.name
        reran: list[str] = []
        base = fake_runner()

        def counting(case, scale, ccp_count, run_idx, total, log_run_dir):
            reran.append(M.make_test_id(scale, ccp_count, case["name"]))
            return base(case, scale, ccp_count, run_idx, total, log_run_dir)

        M.run_one_case_subprocess = counting
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            M.main()
        M.RESUME_FROM_CSV = ""
        check("[C] 只重跑 4 個", len(reran) == 4, f"實際 {len(reran)}")
        csv_c = sorted(p for p in (tmp / "C").glob("*_raw_*.csv") if p.name != csv_b.name)[-1]
        rows_c = M._load_prior_rows(csv_c)
        check("[C] 續跑後 24 列全 OK",
              len(rows_c) == 24 and all(r["status"] == "OK" for r in rows_c))

        print("\n--- 情境 D：Ctrl+C ---")
        M.RESULT_DIR = tmp / "D"
        M.LOG_SUBDIR = tmp / "D_logs"
        seen = {"n": 0}
        base2 = fake_runner()

        def interrupting(case, scale, ccp_count, run_idx, total, log_run_dir):
            seen["n"] += 1
            if seen["n"] > 6:
                raise KeyboardInterrupt
            return base2(case, scale, ccp_count, run_idx, total, log_run_dir)

        M.run_one_case_subprocess = interrupting
        buf = io.StringIO()
        try:
            with contextlib.redirect_stdout(buf):
                M.main()
            survived = True
        except KeyboardInterrupt:
            survived = False
        check("[D] Ctrl+C 被攔截", survived)
        csv_d = sorted((tmp / "D").glob("*_raw_*.csv"))
        check("[D] 已完成的 6 個有保存",
              bool(csv_d) and M.csv_row_count(csv_d[-1]) == 6,
              str(M.csv_row_count(csv_d[-1])) if csv_d else "無")

        print("\n--- 情境 E：Excel 被鎖住 ---")
        M.RESULT_DIR = tmp / "E"
        M.LOG_SUBDIR = tmp / "E_logs"
        M.run_one_case_subprocess = fake_runner()
        import os as _os
        real_replace = _os.replace
        state = {"n": 0}

        def flaky(src, dst, *a, **kw):
            if not str(dst).endswith(".xlsx"):
                return real_replace(src, dst, *a, **kw)
            state["n"] += 1
            if 2 <= state["n"] < 15:      # 第 1 次是 preflight，之後模擬被鎖
                raise PermissionError(13, "被其他程序使用中")
            return real_replace(src, dst, *a, **kw)

        M.EXCEL_REPLACE_RETRY_SLEEP_SEC = 0.0
        _os.replace = flaky
        try:
            buf = io.StringIO()
            with contextlib.redirect_stdout(buf):
                M.main()
            survived = True
        except Exception as exc:  # noqa: BLE001
            survived = False
            print(f"      例外：{type(exc).__name__}: {exc}")
        finally:
            _os.replace = real_replace
            M.EXCEL_REPLACE_RETRY_SLEEP_SEC = 3.0
        check("[E] 被鎖住時實驗沒中止", survived)
        csv_e = sorted((tmp / "E").glob("*_raw_*.csv"))
        check("[E] CSV 仍完整 24 列",
              bool(csv_e) and M.csv_row_count(csv_e[-1]) == 24)

    print("\n" + "=" * 78)
    print(f"通過 {len(PASS)} 項，失敗 {len(FAIL)} 項")
    for name in FAIL:
        print(f"  [FAIL] {name}")
    if not FAIL:
        print("全部通過 — 程式可以正常跑完並輸出結果。")
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())
