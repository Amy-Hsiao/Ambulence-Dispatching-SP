#!/usr/bin/env python3
"""batch_scale_ext_vs_bbc.py 的全流程 dry run（不需 Gurobi）。

用法：python tests/dry_run_scale_ext_vs_bbc.py
"""
from __future__ import annotations

import contextlib
import importlib.util
import io
import shutil
import sys
import tempfile
import types
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "model core"))
sys.path.insert(0, str(ROOT))
sys.modules.setdefault("gurobipy", types.ModuleType("gurobipy"))

spec = importlib.util.spec_from_file_location(
    "SCABL", ROOT / "run experiment" / "batch_scale_ext_vs_bbc.py"
)
M = importlib.util.module_from_spec(spec)
spec.loader.exec_module(M)

PASS, FAIL = [], []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"  [{'OK  ' if cond else 'FAIL'}] {name}" + (f"  — {detail}" if detail else ""))


def fake_runner(fail_these=None, disaster_these=None):
    """disaster_these：模擬 2026-08-16 的 seeding 吃光時間事故。"""
    fail_these = fail_these or set()
    disaster_these = disaster_these or set()

    def _run(case, scale, run_idx, total, log_run_dir):
        test_id = M.make_test_id(scale, case["name"])
        log_path = log_run_dir / f"{run_idx:02d}_{M._safe_case_name(test_id)}.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text("dry run\n", encoding="utf-8")
        counts = M.scale_counts(scale)
        row = M.blank_row()
        row.update({
            "scale": scale, "test_id": test_id, "config": case["name"],
            "I": counts["I"], "J": counts["J"], "H": counts["H"],
            "S": M.BASE_SCENARIOS, "T": M.BASE_TIME_PERIODS,
            "log_path": str(log_path), "solver_log_path": "NA",
        })
        if test_id in fail_these:
            row.update({"status": "FAIL", "note": "injected", "solver_status": "ERROR"})
            return row
        if test_id in disaster_these:
            row.update({
                "status": "OK", "obj_value": "58627595.123717",
                "best_ub": "58627595.123717", "best_lb": "NA", "gap_pct": "NA",
                "cpu_s": "7202.02", "nodes": "0", "root_seed_time_s": "7190.0",
                "root_seed_iters_done": "600", "root_seed_stop_reason": "time_limit",
                "solver_status": "TIME_LIMIT",
            })
            row["sanity"] = "; ".join(M.sanity_warnings(row)) or "OK"
            return row
        idx = M.SCALES.index(scale)
        ub = 25_000_000.0 + 2_000_000 * idx
        row.update({
            "obj_value": f"{ub:.6f}", "best_ub": f"{ub:.6f}",
            "best_lb": f"{ub * 0.5:.6f}", "gap_pct": f"{50.0 + idx:.6f}",
            "cpu_s": "7200.00", "wall_s": "7300.00",
            "nodes": str(60000 - 10000 * idx), "iterations": "5590485",
            "num_vars": "5165050", "num_constrs": "1272740",
            "opened_ccps": "6", "first_stage_decision": "CCP 1 -> X: 1, ...",
            "first_stage_cost": "14952000", "total_cuts": "23263",
            "seed_cuts": "4530", "lazy_cuts": "18223", "user_cuts": "900",
            "root_seed_lb": "11793084", "root_seed_iters_done": "151",
            "root_seed_stop_reason": "lb_rel_improve_below_0.000500_for_10_rounds",
            "root_seed_time_s": "366.56", "root_cut_rounds_done": "15",
            "oracle_solves": "23460", "incumbent_evals": "420",
            "callback_time_s": "4604.57", "parallel_oracles": "5",
            "solver_status": "TIME_LIMIT", "status": "OK", "note": "",
        })
        row["sanity"] = "; ".join(M.sanity_warnings(row)) or "OK"
        return row

    return _run


def verify_excel(xlsx: Path, label: str, expected_rows: int):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    try:
        check(f"[{label}] Excel 分頁齊全",
              wb.sheetnames == ["raw_results", "run_settings", "summary_table"],
              str(wb.sheetnames))
        check(f"[{label}] raw_results 列數",
              wb["raw_results"].max_row == expected_rows + 1,
              str(wb["raw_results"].max_row))
        ws = wb["summary_table"]
        n_left, n_sub = 4, 5
        check(f"[{label}] summary_table 大小 = {2 + len(M.SCALES)} 列 × "
              f"{n_left + len(M.CONFIGS) * n_sub} 欄",
              ws.max_row == 2 + len(M.SCALES)
              and ws.max_column == n_left + len(M.CONFIGS) * n_sub,
              f"{ws.max_row} × {ws.max_column}")
        subs = [ws.cell(2, n_left + 1 + i).value for i in range(n_sub)]
        check(f"[{label}] 子欄位 = UB/LB/Time/Gap/Nodes",
              subs == ["UB", "LB", "Time(s)", "Gap(%)", "Nodes"], str(subs))
        names = [ws.cell(1, n_left + 1 + i * n_sub).value for i in range(len(M.CONFIGS))]
        check(f"[{label}] 兩個 config 標題", names == [c["name"] for c in M.CONFIGS],
              str(names))
        rows_i = [ws.cell(3 + i, 2).value for i in range(len(M.SCALES))]
        check(f"[{label}] 災區數欄 = 70/100/130", rows_i == [70, 100, 130], str(rows_i))
        js = [ws.cell(3 + i, 3).value for i in range(len(M.SCALES))]
        expected_j = [M.scale_counts(s)["J"] for s in M.SCALES]
        check(f"[{label}] |J| 三規模相同且等於 config 設定 {expected_j[0]}",
              js == expected_j and len(set(js)) == 1, str(js))
    finally:
        wb.close()


def main() -> int:
    print("=" * 78)
    print("batch_scale_ext_vs_bbc.py — 全流程 DRY RUN")
    print("=" * 78)
    n = len(M.SCALES) * len(M.CONFIGS)
    print(f"矩陣：{len(M.SCALES)} scale × {len(M.CONFIGS)} config = {n} cases\n")

    check("共 6 個 case", n == 6, str(n))
    check("災區 70/100/130",
          [M.scale_counts(s)["I"] for s in M.SCALES] == [70, 100, 130])
    check("只有 Extensive + BBC 全開",
          [c["name"] for c in M.CONFIGS] == ["Extensive", "BBC+WS+RS+UC+Pareto"])
    check("時限 7200s", M.TIME_LIMIT == 7200.0)
    check("預設從頭跑", M.RESUME_FROM_CSV == "")

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)

        print("\n--- 情境 A：6 個 case 全部正常 ---")
        M.RESULT_DIR = tmp / "A"; M.LOG_SUBDIR = tmp / "A_logs"; M.RESUME_FROM_CSV = ""
        M.run_one_case_subprocess = fake_runner()
        with contextlib.redirect_stdout(io.StringIO()):
            M.main()
        csvs = sorted((tmp / "A").glob("*_raw_*.csv"))
        xlsxs = [p for p in (tmp / "A").glob("*.xlsx") if not p.name.startswith(".")]
        check("[A] 產生 CSV 且 6 列", len(csvs) == 1 and M.csv_row_count(csvs[-1]) == 6)
        check("[A] 產生 Excel", len(xlsxs) == 1)
        verify_excel(xlsxs[-1], "A", 6)
        rows = M._load_prior_rows(csvs[-1])
        check("[A] 全部 sanity = OK", all(r["sanity"] == "OK" for r in rows))
        check("[A] 24 個 log 欄位都有值", all(r["log_path"] not in ("", "NA") for r in rows))

        print("\n--- 情境 B：模擬 seeding 吃光時間的事故（要被自動抓到）---")
        M.RESULT_DIR = tmp / "B"; M.LOG_SUBDIR = tmp / "B_logs"
        disaster = {M.make_test_id(s, "BBC+WS+RS+UC+Pareto") for s in M.SCALES}
        M.run_one_case_subprocess = fake_runner(disaster_these=disaster)
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            M.main()
        out = buf.getvalue()
        rows_b = M._load_prior_rows(sorted((tmp / "B").glob("*_raw_*.csv"))[-1])
        bad = [r for r in rows_b if r["sanity"] not in ("OK", "NA", "")]
        check("[B] 3 個異常 case 被 sanity 抓到", len(bad) == 3, f"{len(bad)} 個")
        check("[B] 畫面有印出「結果異常」警告", "結果異常" in out or "結果不正常" in out)
        check("[B] 異常內容含 nodes=0 與 seeding 佔比",
              any("nodes=0" in r["sanity"] and "seeding" in r["sanity"] for r in bad))

        print("\n--- 情境 C：case 失敗不中止 + 續跑 ---")
        M.RESULT_DIR = tmp / "C"; M.LOG_SUBDIR = tmp / "C_logs"
        failing = {M.make_test_id("large", "Extensive")}
        M.run_one_case_subprocess = fake_runner(fail_these=failing)
        with contextlib.redirect_stdout(io.StringIO()):
            M.main()
        csv_c = sorted((tmp / "C").glob("*_raw_*.csv"))[-1]
        rows_c = M._load_prior_rows(csv_c)
        check("[C] 失敗不中止（仍 6 列）", len(rows_c) == 6)
        check("[C] 1 個記為 FAIL", sum(1 for r in rows_c if r["status"] == "FAIL") == 1)

        M.RESULT_DIR = tmp / "D"; M.LOG_SUBDIR = tmp / "D_logs"
        (tmp / "D").mkdir(parents=True, exist_ok=True)
        shutil.copy2(csv_c, tmp / "D" / csv_c.name)
        M.RESUME_FROM_CSV = csv_c.name
        reran = []
        base = fake_runner()

        def counting(case, scale, run_idx, total, log_run_dir):
            reran.append(M.make_test_id(scale, case["name"]))
            return base(case, scale, run_idx, total, log_run_dir)

        M.run_one_case_subprocess = counting
        with contextlib.redirect_stdout(io.StringIO()):
            M.main()
        M.RESUME_FROM_CSV = ""
        check("[C] 續跑只補跑 1 個", len(reran) == 1, f"實際 {len(reran)}")
        # 找出續跑後那份完整的 CSV（同一秒啟動時檔名會加後綴，故逐一檢查）
        complete = [c for c in sorted((tmp / "D").glob("*_raw_*.csv"))
                    if M.csv_row_count(c) == 6
                    and all(r["status"] == "OK" for r in M._load_prior_rows(c))]
        check("[C] 續跑後產生一份 6 列全 OK 的 CSV", bool(complete),
              f"D 目錄有 {len(list((tmp / 'D').glob('*_raw_*.csv')))} 份 CSV")
        check("[C] 同一秒啟動兩次不會撞名/崩潰",
              len({c.name for c in (tmp / "D").glob("*_raw_*.csv")}) ==
              len(list((tmp / "D").glob("*_raw_*.csv"))))

        print("\n--- 情境 D：Ctrl+C ---")
        M.RESULT_DIR = tmp / "E"; M.LOG_SUBDIR = tmp / "E_logs"
        seen = {"n": 0}
        b2 = fake_runner()

        def interrupting(case, scale, run_idx, total, log_run_dir):
            seen["n"] += 1
            if seen["n"] > 3:
                raise KeyboardInterrupt
            return b2(case, scale, run_idx, total, log_run_dir)

        M.run_one_case_subprocess = interrupting
        try:
            with contextlib.redirect_stdout(io.StringIO()):
                M.main()
            survived = True
        except KeyboardInterrupt:
            survived = False
        check("[D] Ctrl+C 被攔截", survived)
        csv_e = sorted((tmp / "E").glob("*_raw_*.csv"))
        check("[D] 已完成的 3 個有保存",
              bool(csv_e) and M.csv_row_count(csv_e[-1]) == 3)

    print("\n" + "=" * 78)
    print(f"通過 {len(PASS)} 項，失敗 {len(FAIL)} 項")
    for x in FAIL:
        print(f"  [FAIL] {x}")
    if not FAIL:
        print("全部通過 — 程式可以正常跑完並輸出結果。")
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())
