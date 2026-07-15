import ast
import importlib.util
import tempfile
import unittest
from pathlib import Path
from unittest import mock


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "run experiment" / "batch_ablation_experiment.py"
SPEC = importlib.util.spec_from_file_location("batch_ablation_under_test", SCRIPT)
ABL = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(ABL)


class _FakeModel:
    NumVars = 10
    NumConstrs = 20
    NodeCount = 3
    IterCount = 4

    def __init__(self):
        self.disposed = False

    def dispose(self):
        self.disposed = True


class _FakePortal:
    @staticmethod
    def _result():
        model = _FakeModel()
        fs = {
            "X": {"J01": 1.0},
            "V": {"J01": 2.0},
            "U": {"J01": 3.0},
            "Y": {("H01", "J01"): 4.0},
        }
        summary = {
            "objective": 100.0,
            "best_lb": 99.0,
            "gap_pct": 1.0,
            "first_stage": fs,
            "bbc_stats": {
                "runtime": 1.5,
                "solver_status": "OPTIMAL",
                "cuts_added": 2,
                "seed_cuts_added": 0,
                "lazy_cuts_added": 2,
                "user_cuts_added": 0,
                "parallel_oracles": 5,
            },
        }
        return model, summary

    def run_sp_model(self, **_kwargs):
        print("fake SP solve")
        return self._result()

    def run_dro_model(self, **_kwargs):
        print("fake DRO solve")
        return self._result()


class BatchAblationTests(unittest.TestCase):
    def test_extensive_portal_uses_cpu_parallel_only_parameters(self):
        portal_path = ROOT / "model portal" / "extensive_dro.py"
        tree = ast.parse(portal_path.read_text(encoding="utf-8"))
        function = next(
            node for node in tree.body
            if isinstance(node, ast.FunctionDef)
            and node.name == "_configure_cpu_parallel_only"
        )
        namespace = {}
        exec(compile(ast.Module(body=[function], type_ignores=[]), str(portal_path), "exec"), namespace)

        class FakeModel:
            def __init__(self):
                self.params = {}

            def setParam(self, name, value):
                self.params[name] = value

        model = FakeModel()
        settings = namespace["_configure_cpu_parallel_only"](model)
        self.assertEqual({
            "Threads": 0,
            "ConcurrentMIP": 1,
            "Presolve": 0,
            "Cuts": 0,
            "Heuristics": 0.0,
            "Symmetry": 0,
            "MIPFocus": 0,
        }, settings)
        self.assertEqual(settings, model.params)

    def test_exact_36_case_matrix_and_configuration_ladder(self):
        self.assertEqual(36, len(ABL.expected_test_ids()))
        expected = [
            ("Extensive", False, 0, 0, False, False),
            ("BBC", False, 0, 0, False, False),
            ("BBC+WS", True, 0, 0, False, False),
            ("BBC+WS+RS", True, ABL.DEFAULT_ROOT_SEED_ITERS, 0, False, False),
            (
                "BBC+WS+RS+UC", True, ABL.DEFAULT_ROOT_SEED_ITERS,
                ABL.DEFAULT_ROOT_CUT_ROUNDS, True, False,
            ),
            (
                "BBC+WS+RS+UC+Pareto", True, ABL.DEFAULT_ROOT_SEED_ITERS,
                ABL.DEFAULT_ROOT_CUT_ROUNDS, True, True,
            ),
        ]
        actual = [
            (c["name"], c["ev"], c["seed"], c["rounds"], c["user"], c["pareto"])
            for c in ABL.CONFIGS
        ]
        self.assertEqual(expected, actual)

    def test_temporary_config_enables_common_bbc_tuning_and_restores(self):
        keys = [
            "BENDERS_MIPFOCUS", "BENDERS_HEURISTICS",
            "BENDERS_NUMERIC_FOCUS", "BENDERS_X_BRANCH_PRIORITY_ENABLED",
            "BENDERS_X_BRANCH_PRIORITY",
        ]
        original = {key: getattr(ABL.cfg, key) for key in keys}
        with ABL.temporary_config(ABL.CONFIGS[1]):
            self.assertEqual(ABL.DEFAULT_MIPFOCUS, ABL.cfg.BENDERS_MIPFOCUS)
            self.assertEqual(ABL.DEFAULT_HEURISTICS, ABL.cfg.BENDERS_HEURISTICS)
            self.assertEqual(ABL.DEFAULT_NUMERIC_FOCUS, ABL.cfg.BENDERS_NUMERIC_FOCUS)
            self.assertEqual(
                ABL.DEFAULT_BRANCH_PRIORITY_ENABLED,
                ABL.cfg.BENDERS_X_BRANCH_PRIORITY_ENABLED,
            )
            self.assertEqual(
                ABL.DEFAULT_BRANCH_PRIORITY,
                ABL.cfg.BENDERS_X_BRANCH_PRIORITY,
            )
        with ABL.temporary_config(ABL.CONFIGS[0]):
            self.assertEqual(0, ABL.cfg.BENDERS_MIPFOCUS)
            self.assertEqual(0.0, ABL.cfg.BENDERS_HEURISTICS)
            self.assertEqual(0, ABL.cfg.BENDERS_NUMERIC_FOCUS)
            self.assertFalse(ABL.cfg.BENDERS_X_BRANCH_PRIORITY_ENABLED)
            self.assertEqual(0, ABL.cfg.BENDERS_X_BRANCH_PRIORITY)
        self.assertEqual(original, {key: getattr(ABL.cfg, key) for key in keys})

    def test_each_case_has_a_canonical_log(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            log_dir = Path(temp_dir)
            portal = _FakePortal()
            row = ABL.run_one_case(
                {"SP": portal, "DRO-box": portal, "ext": portal},
                "SP", ABL.CONFIGS[1], {"I": 20, "J": 10, "H": 6},
                "small", 1, 36, log_dir,
            )
            self.assertEqual("OK", row["status"])
            log_path = Path(row["log_path"])
            self.assertTrue(log_path.is_file())
            self.assertIn("fake SP solve", log_path.read_text(encoding="utf-8"))
            self.assertIn("01_small_SP_BBC", log_path.name)
            self.assertEqual(ABL.DEFAULT_MIPFOCUS, row["mip_focus"])
            self.assertEqual(ABL.DEFAULT_HEURISTICS, row["heuristics"])
            self.assertEqual(ABL.DEFAULT_NUMERIC_FOCUS, row["numeric_focus"])
            self.assertTrue(row["branch_priority_enabled"])
            self.assertEqual(ABL.DEFAULT_BRANCH_PRIORITY, row["branch_priority"])

    def test_subprocess_hard_timeout_still_returns_row_and_log(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with mock.patch.object(ABL, "TIME_LIMIT", 0.01), mock.patch.object(
                ABL.subprocess,
                "run",
                side_effect=ABL.subprocess.TimeoutExpired(["python"], 0.01),
            ):
                row = ABL.run_one_case_subprocess(
                    "SP", ABL.CONFIGS[1], {"I": 20, "J": 10, "H": 6},
                    "small", 1, 36, Path(temp_dir),
                )
            self.assertEqual("FAIL", row["status"])
            self.assertEqual("HARD_TIMEOUT", row["solver_status"])
            self.assertTrue(Path(row["log_path"]).is_file())
            self.assertIn("Hard timeout", row["note"])

    def test_complete_excel_and_csv_are_readable(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir)
            rows = []
            for scale in ABL.SCALES:
                counts = ABL.scale_counts(scale)
                for model_name in ABL.MODELS:
                    for case in ABL.CONFIGS:
                        row = ABL.blank_row()
                        test_id = f"{scale}_{model_name}_{case['name']}".replace("+", "_plus_")
                        log_path = output / f"{test_id}.log"
                        log_path.write_text("test log", encoding="utf-8")
                        row.update({
                            "scale": scale, "test_id": test_id,
                            "model": model_name, "config": case["name"],
                            "I": counts["I"], "J": counts["J"], "H": counts["H"],
                            "S": 30, "T": 8, "obj_value": 100.0,
                            "first_stage_decision": "test", "best_lb": 99.0,
                            "best_ub": 100.0, "cpu_s": 1.0, "wall_s": 1.1,
                            "num_vars": 10, "num_constrs": 20, "nodes": 3,
                            "iterations": 4, "gap_pct": 1.0,
                            "ev_warm_start": case["ev"],
                            "root_seed_iters": case["seed"],
                            "root_cut_rounds": case["rounds"],
                            "use_user_cuts": case["user"],
                            "pareto_enabled": case["pareto"],
                            "multi_cut": True, "parallel_oracles": 5,
                            "mip_focus": ABL.DEFAULT_MIPFOCUS,
                            "heuristics": ABL.DEFAULT_HEURISTICS,
                            "numeric_focus": ABL.DEFAULT_NUMERIC_FOCUS,
                            "branch_priority_enabled": ABL.DEFAULT_BRANCH_PRIORITY_ENABLED,
                            "branch_priority": ABL.DEFAULT_BRANCH_PRIORITY,
                            "solver_status": "OPTIMAL", "log_path": str(log_path),
                            "status": "OK", "note": "",
                        })
                        rows.append(row)

            csv_path = output / "results.csv"
            xlsx_path = output / "results.xlsx"
            ABL.write_results(csv_path, rows)
            ABL.export_xlsx(rows, xlsx_path, "test_run", output)
            ABL.validate_final_outputs(rows, csv_path, xlsx_path)

            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True, data_only=False)
            try:
                self.assertEqual(
                    ["raw_results", "run_settings", "small", "medium", "large", "summary_table"],
                    wb.sheetnames,
                )
                self.assertEqual(37, wb["raw_results"].max_row)
                self.assertEqual(len(ABL.FIELDNAMES), wb["raw_results"].max_column)
                for scale in ABL.SCALES:
                    self.assertEqual(13, wb[scale].max_row)
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
